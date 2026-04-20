from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.utils.cache import add_never_cache_headers
from django.views.decorators.csrf import csrf_protect
from django.urls import reverse
from django.urls.exceptions import NoReverseMatch
from django.db.models import Sum, Q, Avg, Subquery, IntegerField
from django.template.loader import get_template
from django.http import Http404, HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, StreamingHttpResponse
from datetime import datetime, timedelta, time
from django.db.models import Case, When, Value, CharField
from django.db.models.functions import Concat
from django.utils import timezone
from django.db.models import Sum
from django.db.models.functions import TruncMonth, TruncQuarter
from django.db.models import Count
from django.db.models.functions import ExtractYear
from .models import (
    User, Budget, Proposal, Project, BudgetTransaction, Task, AuditLog, Notification,
    BudgetDocument, ProposalDocument, ProjectDocument, ExpenseDocument, ProjectExpense,
    FormTemplate, ExtensionRequest, PersonalTask, Message, GroupChat, GroupChatMember,
    GroupChatMessage, Announcement, SystemHealth, BackupStatus, MaintenanceSchedule,
    BudgetAllocation, EquipmentCategory, EquipmentItem, ProjectEquipment, TrancheRelease,
    CalendarEvent, DigitalSignature, Mention, ProjectMilestone, Translation, UserPreference,
    DeletedMessage, DeletedConversation, DeletedGroupChat, DeletedGroupChatMessage
)
from .forms import MessageForm
from .validators import (
    validate_profile_picture, validate_document_upload, validate_image_upload,
    validate_file_extension, validate_file_size, validate_password_strength,
    validate_positive_decimal, validate_positive_integer, sanitize_string,
    ALLOWED_DOCUMENT_EXTENSIONS, ALLOWED_IMAGE_EXTENSIONS, MAX_DOCUMENT_SIZE
)
import datetime
from datetime import datetime
from django.core.exceptions import ValidationError
from django.core.cache import cache
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
import csv
from collections import defaultdict
from django.views.decorators.http import require_POST, require_GET
from django.http import JsonResponse
import json
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.utils.dateparse import parse_date
from django.contrib.auth import update_session_auth_hash
from decimal import Decimal
import math
import io
import base64
import uuid
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from geopy.geocoders import Nominatim # type: ignore
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend for server-side plotting
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator

import os
import time as pytime
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Frame, PageTemplate
from django.conf import settings
from urllib.parse import quote

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

#Remove cache
def add_no_cache_headers(response):
    add_never_cache_headers(response)
    return response


def apply_profile_picture_update(request, user):
    remove_flag = request.POST.get('remove_profile_picture') == 'true'
    cropped_data = request.POST.get('cropped_profile_picture')

    if remove_flag:
        if user.profile_picture:
            user.profile_picture.delete(save=False)
        user.profile_picture = None
        return True

    if cropped_data:
        try:
            header, data = cropped_data.split(';base64,')
            ext = header.split('/')[-1]
            decoded = ContentFile(
                base64.b64decode(data),
                name=f"profile_{user.id}_{uuid.uuid4().hex}.{ext}"
            )
            validate_profile_picture(decoded)
            user.profile_picture = decoded
            return True
        except (ValueError, ValidationError) as e:
            error_msg = e.messages[0] if hasattr(e, 'messages') and e.messages else str(e)
            messages.error(request, f"Profile picture error: {error_msg}")
            return False

    if 'profile_picture' in request.FILES:
        profile_picture = request.FILES['profile_picture']
        try:
            validate_profile_picture(profile_picture)
        except ValidationError as e:
            error_msg = e.messages[0] if e.messages else str(e)
            messages.error(request, f"Profile picture error: {error_msg}")
            return False
        user.profile_picture = profile_picture

    return True


def _invalidate_admin_comm_hub_cache_for_users(user_ids):
    """Invalidate short-lived communication hub cache for the given user IDs."""
    for user_id in {uid for uid in user_ids if uid}:
        cache.delete(f'admin_comm_hub_payload_{user_id}_v1')


def _invalidate_comm_hub_cache_for_group_chat(group_chat):
    """Invalidate communication hub cache for active members of a group chat."""
    member_ids = GroupChatMember.objects.filter(
        group_chat=group_chat,
        is_active=True,
    ).values_list('user_id', flat=True)
    _invalidate_admin_comm_hub_cache_for_users(member_ids)


def _get_announcement_target_user_ids(announcement):
    """Return user IDs affected by an announcement for cache invalidation."""
    explicit_user_ids = list(announcement.target_users.values_list('id', flat=True))
    if explicit_user_ids:
        return explicit_user_ids

    if announcement.target_roles:
        return list(
            User.objects.filter(
                role__in=announcement.target_roles,
                status='active',
            ).values_list('id', flat=True)
        )

    return list(User.objects.filter(status='active').values_list('id', flat=True))


def _build_user_conversations(user):
    """Build conversation list with minimized DB queries and delete-awareness."""
    deleted_message_ids = DeletedMessage.objects.filter(user=user).values('message_id')
    deleted_convo_cutoffs = dict(
        DeletedConversation.objects.filter(user=user).values_list('partner_id', 'delete_before')
    )

    message_base_qs = Message.objects.filter(
        Q(sender=user) | Q(recipient=user)
    ).exclude(id__in=Subquery(deleted_message_ids))

    latest_message_by_partner = {}
    ordered_messages = message_base_qs.select_related('sender', 'recipient').order_by('-created_at')

    for message in ordered_messages:
        partner = message.recipient if message.sender_id == user.id else message.sender
        if not partner or partner.id == user.id:
            continue

        cutoff = deleted_convo_cutoffs.get(partner.id)
        if cutoff and message.created_at <= cutoff:
            continue

        if partner.id not in latest_message_by_partner:
            latest_message_by_partner[partner.id] = message

    unread_counts = defaultdict(int)
    unread_messages = message_base_qs.filter(recipient=user, is_read=False).only('sender_id', 'created_at')
    for unread in unread_messages:
        cutoff = deleted_convo_cutoffs.get(unread.sender_id)
        if cutoff and unread.created_at <= cutoff:
            continue
        unread_counts[unread.sender_id] += 1

    conversations = []
    for partner_id, latest_message in latest_message_by_partner.items():
        partner = latest_message.recipient if latest_message.sender_id == user.id else latest_message.sender
        conversations.append({
            'partner': partner,
            'latest_message': latest_message,
            'unread_count': unread_counts.get(partner_id, 0),
            'last_activity': latest_message.created_at,
        })

    return conversations


def _get_auth_splash_base_url():
    """Return auth splash URL with safe fallback when URL name is unavailable."""
    try:
        return reverse('auth_splash_url')
    except NoReverseMatch:
        return '/auth/splash/'


# Login
@csrf_protect
def index_view(request):
    if request.method == "POST":
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)

        if user is not None:
            login(request, user)

            # Philippine time
            philippine_time = timezone.now() + timedelta(hours=8)

            AuditLog.objects.create(
                user=user,
                action="login",
                model_name="User",
                object_id=str(user.pk),
                old_data=None,
                new_data={
                    "email": user.email,
                    "role": user.role,
                    "status": "Login successful",
                    "ip_address": request.META.get("REMOTE_ADDR"),
                },
                timestamp=philippine_time
            )

            # Redirect based on role
            target_url = reverse('index_url')
            if user.is_superuser or user.role == 'admin':
                target_url = reverse('administrator_dashboard_url')
            elif user.role == 'dost_staff':
                target_url = reverse('staff_dashboard_url')
            elif user.role == 'proponent':
                target_url = reverse('proponent_dashboard_url')
            elif user.role == 'beneficiary':
                target_url = reverse('beneficiary_dashboard_url')
            else:
                messages.error(request, 'Role not recognized.')
                return redirect('index_url')

            splash_base = _get_auth_splash_base_url()
            return redirect(f"{splash_base}?mode=login&next={quote(target_url, safe='/?:=&')}")

        else:
            messages.error(request, 'Invalid email or password.')
            return render(request, 'index.html', {'messages': messages.get_messages(request)})

    return render(request, 'index.html')

# Logout
def logout_view(request):
    if request.user.is_authenticated:
        # ───────────────────────────────────────
        # CREATE AUDIT LOG FOR LOGOUT
        AuditLog.objects.create(
            user=request.user,
            action="logout",
            model_name="User",
            object_id=str(request.user.pk),
            old_data={
                "email": request.user.email,
                "role": request.user.role,
            },
            new_data={
                "status": "User logged out",
                "ip_address": request.META.get("REMOTE_ADDR"),
            }
        )
        # ───────────────────────────────────────

    logout(request)
    login_url = reverse('index_url')
    splash_base = _get_auth_splash_base_url()
    return redirect(f"{splash_base}?mode=logout&next={quote(login_url, safe='/?:=&')}")


def auth_splash_view(request):
    """One-time transition splash shown after login/logout before redirect."""
    mode = request.GET.get('mode', 'login')
    next_url = request.GET.get('next', reverse('index_url'))

    if not next_url.startswith('/'):
        next_url = reverse('index_url')

    if mode not in ['login', 'logout']:
        mode = 'login'

    context = {
        'mode': mode,
        'next_url': next_url,
    }
    return render(request, 'auth_splash.html', context)



# Haversine formula
def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371  # Earth radius in km
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c



@login_required
def mark_notification_read_view(request, pk):
    try:
        notif = Notification.objects.get(pk=pk, receiver=request.user)
        notif.status = "read"
        notif.save(update_fields=['status'])
        return JsonResponse({"success": True})
    except Notification.DoesNotExist:
        return JsonResponse({"success": False})


@login_required
def delete_notification_view(request, pk):
    """Delete a notification for the current user"""
    if request.method == 'POST':
        try:
            notif = Notification.objects.get(pk=pk, receiver=request.user)
            notif.delete()
            return JsonResponse({"success": True})
        except Notification.DoesNotExist:
            return JsonResponse({"success": False, "error": "Notification not found"})
    return JsonResponse({"success": False, "error": "Invalid method"})
    

@login_required
def clear_all_notifications_view(request):
    """Clear all notifications for the current user"""
    if request.method == 'POST':
        try:
            Notification.objects.filter(receiver=request.user).delete()
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid method"})


@login_required
def get_notification_count_view(request):
    """API endpoint to get current notification count for polling"""
    try:
        payload = _build_notification_payload(request.user)
        payload["success"] = True
        return JsonResponse(payload)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


def _build_notification_payload(user):
    unread_count = Notification.objects.filter(
        receiver=user,
        status='unread'
    ).count()

    notifications = Notification.objects.filter(
        receiver=user
    ).order_by('-timestamp')[:10]

    notifications_data = []
    for notif in notifications:
        notifications_data.append({
            'id': notif.id,
            'message': notif.message,
            'category': notif.category,
            'status': notif.status,
            'timestamp': notif.timestamp.strftime('%b %d, %Y %H:%M'),
            'link': notif.link or '',
        })

    return {
        "unread_count": unread_count,
        "notifications": notifications_data,
    }


@login_required
def notifications_stream_view(request):
    """Server-Sent Events endpoint for near real-time notification updates."""

    def event_stream():
        # Immediate first payload so the UI updates without waiting for a poll interval.
        last_payload = None
        initial_payload = _build_notification_payload(request.user)
        last_payload = json.dumps(initial_payload)
        yield f"event: notification\ndata: {last_payload}\n\n"

        heartbeat_counter = 0
        while True:
            try:
                current_payload = json.dumps(_build_notification_payload(request.user))
                if current_payload != last_payload:
                    last_payload = current_payload
                    yield f"event: notification\ndata: {current_payload}\n\n"

                heartbeat_counter += 1
                if heartbeat_counter >= 15:
                    heartbeat_counter = 0
                    yield ": keepalive\n\n"

                pytime.sleep(1)
            except GeneratorExit:
                break
            except Exception:
                # End this stream; client auto-reconnect handles recovery.
                break

    response = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response


# Administrator
@login_required
def administrator_dashboard_view(request):
    # -----------------------------
    # Municipality Coordinates for Biliran
    # -----------------------------
    MUNICIPALITY_COORDS = {
        'Almeria': {'lat': 11.6167, 'lng': 124.4333},
        'Biliran': {'lat': 11.5833, 'lng': 124.4667},
        'Cabucgayan': {'lat': 11.4667, 'lng': 124.5500},
        'Caibiran': {'lat': 11.5500, 'lng': 124.5833},
        'Culaba': {'lat': 11.6500, 'lng': 124.5500},
        'Kawayan': {'lat': 11.6667, 'lng': 124.5000},
        'Maripipi': {'lat': 11.7833, 'lng': 124.3333},
        'Naval': {'lat': 11.5667, 'lng': 124.4000},
    }

    # -----------------------------
    # Projects Data for Map
    # -----------------------------
    projects_qs = Project.objects.all()

    projects = []
    for p in projects_qs:
        municipality = p.mun or ''
        coords = MUNICIPALITY_COORDS.get(municipality, {})
        
        # Use project's actual coordinates if available, otherwise use municipality center with offset
        if p.latitude is not None and p.longitude is not None:
            lat = p.latitude
            lng = p.longitude
        elif coords:
            # Add small random offset to prevent markers from stacking on same point
            import random
            lat = coords.get('lat') + random.uniform(-0.01, 0.01)
            lng = coords.get('lng') + random.uniform(-0.01, 0.01)
        else:
            lat = None
            lng = None
        
        projects.append({
            'id': p.id,
            'title': p.project_title or '',
            'description': p.remarks or p.project_description or '',
            'municipality': municipality,
            'province': p.province or 'Biliran',
            'beneficiary': p.beneficiary or '',
            'proponent': p.proponent_details or '',
            'status': (p.status or 'new').lower().replace(' ', '_').replace('-', '_'),
            'funds': float(p.funds) if p.funds else 0,
            'latitude': lat,
            'longitude': lng,
        })

    proposals_qs = Proposal.objects.select_related('submitted_by').all()
    proposals = []
    for proposal in proposals_qs:
        proposals.append({
            'id': proposal.id,
            'title': proposal.title or '',
            'municipality': proposal.municipality or '',
            'province': proposal.province or 'Biliran',
            'status': (proposal.status or 'pending').lower().replace(' ', '_').replace('-', '_'),
            'submission_date': proposal.submission_date.isoformat() if proposal.submission_date else '',
            'submitted_by': (
                f"{proposal.submitted_by.first_name} {proposal.submitted_by.last_name}".strip()
                if proposal.submitted_by else ''
            ),
            'proposed_amount': float(proposal.proposed_amount) if proposal.proposed_amount else 0,
        })

    # -----------------------------
    # Total Budget
    # -----------------------------
    total_budget = Budget.objects.aggregate(total=Sum('total_equipment_value'))['total'] or Decimal('0.00')

    # -----------------------------
    # Total Spent (calculated from budget allocations, not transactions)
    # This correctly reflects all funds that have been allocated to projects
    # -----------------------------
    total_delivered = Budget.objects.aggregate(total=Sum('delivered_equipment_value'))['total'] or Decimal('0.00')
    total_spent = total_delivered
    total_remaining = total_budget - total_spent
    total_spent = total_spent.quantize(Decimal('0.01'))
    total_remaining = total_remaining.quantize(Decimal('0.01'))

    # -----------------------------
    # Utilization Rate
    # -----------------------------
    all_projects_count = Project.objects.count()
    completed_projects_count = Project.objects.filter(status__iexact='completed').count()

    utilization_rate = (completed_projects_count / all_projects_count * 100) if all_projects_count > 0 else 0
    utilization_rate = round(utilization_rate, 2)

    # -----------------------------
    # Chart 1: Project Status (normalized into fixed buckets)
    # -----------------------------
    project_status_counts = Project.objects.values('status').annotate(total=Count('id'))
    status_data = {'Proposal': 0, 'Ongoing': 0, 'Completed': 0, 'Terminated': 0, 'Unknown': 0}

    def normalize_status(raw_status):
        if not raw_status:
            return 'Unknown'
        s = str(raw_status).strip().lower()
        # map common variants into the dashboard buckets
        if s in ('proposal', 'proposed'):
            return 'Proposal'
        if s in ('ongoing', 'on-going', 'in_progress', 'in progress', 'inprogress'):
            return 'Ongoing'
        if s in ('completed', 'complete'):
            return 'Completed'
        if s in ('terminated', 'cancelled', 'canceled'):
            return 'Terminated'
        return 'Unknown'

    for entry in project_status_counts:
        raw = entry.get('status')
        label = normalize_status(raw)
        # accumulate counts into the normalized bucket
        status_data[label] = status_data.get(label, 0) + entry.get('total', 0)

    project_status_labels = list(status_data.keys())
    project_status_values = list(status_data.values())

    # -----------------------------
    # Chart 2: User Roles (Users Module)
    # -----------------------------
    user_role_counts = User.objects.values('role').annotate(total=Count('id'))
    user_role_labels = []
    user_role_values = []
    for entry in user_role_counts:
        # Format 'dost_staff' -> 'Dost Staff'
        label = entry['role'].replace('_', ' ').title()
        user_role_labels.append(label)
        user_role_values.append(entry['total'])

    # -----------------------------
    # Chart 3: Proposal Status (Proposals Module)
    # -----------------------------
    proposal_status_counts = Proposal.objects.values('status').annotate(total=Count('id'))
    proposal_status_labels = []
    proposal_status_values = []
    for entry in proposal_status_counts:
        label = entry['status'].replace('_', ' ').title()
        proposal_status_labels.append(label)
        proposal_status_values.append(entry['total'])

    # -----------------------------
    # Chart 4: Extension Request Status
    # -----------------------------
    extension_status_counts = ExtensionRequest.objects.values('status').annotate(total=Count('id'))
    extension_status_labels = []
    extension_status_values = []
    for entry in extension_status_counts:
        label = entry['status'].replace('_', ' ').title()
        extension_status_labels.append(label)
        extension_status_values.append(entry['total'])

    # -----------------------------
    # Chart 5: Proposals Over Time (Last 12 Months)
    # -----------------------------
    from django.db.models.functions import TruncMonth
    twelve_months_ago = timezone.now() - timedelta(days=365)
    proposals_by_month = Proposal.objects.filter(
        submission_date__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('submission_date')
    ).values('month').annotate(count=Count('id')).order_by('month')
    
    proposals_month_labels = []
    proposals_month_values = []
    for entry in proposals_by_month:
        if entry['month']:
            proposals_month_labels.append(entry['month'].strftime('%b %Y'))
            proposals_month_values.append(entry['count'])

    # -----------------------------
    # Chart 6: Extension Requests Over Time (Last 12 Months)
    # -----------------------------
    extensions_by_month = ExtensionRequest.objects.filter(
        date_submitted__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('date_submitted')
    ).values('month').annotate(count=Count('id')).order_by('month')
    
    extensions_month_labels = []
    extensions_month_values = []
    for entry in extensions_by_month:
        if entry['month']:
            extensions_month_labels.append(entry['month'].strftime('%b %Y'))
            extensions_month_values.append(entry['count'])

    # -----------------------------
    # Summary Stats
    # -----------------------------
    total_users = User.objects.count()
    total_proposals = Proposal.objects.count()
    total_projects = Project.objects.count()
    pending_extensions = ExtensionRequest.objects.filter(status='pending').count()

    # -----------------------------
    # Task Completion Stats
    # -----------------------------
    task_pending = Task.objects.filter(status='pending').count()
    task_in_progress = Task.objects.filter(status='in_progress').count()
    task_completed = Task.objects.filter(status='completed').count()
    task_overdue = Task.objects.filter(status='overdue').count()

    # -----------------------------
    # Recent Activity Feed
    # -----------------------------
    recent_activities = AuditLog.objects.select_related('user').order_by('-timestamp')[:10]

    context = {
        'projects_json': json.dumps(projects),
        'proposals_json': json.dumps(proposals),
        'total_budget': int(total_budget),
        'total_spent': int(total_spent),
        'total_remaining': int(total_remaining),
        'utilization_rate': utilization_rate,
        
        # Project Chart
        'project_status_labels': json.dumps(project_status_labels),
        'project_status_values': json.dumps(project_status_values),
        
        # User Chart
        'user_role_labels': json.dumps(user_role_labels),
        'user_role_values': json.dumps(user_role_values),
        
        # Proposal Chart
        'proposal_status_labels': json.dumps(proposal_status_labels),
        'proposal_status_values': json.dumps(proposal_status_values),
        
        # Extension Request Chart
        'extension_status_labels': json.dumps(extension_status_labels),
        'extension_status_values': json.dumps(extension_status_values),
        
        # Trends Charts
        'proposals_month_labels': json.dumps(proposals_month_labels),
        'proposals_month_values': json.dumps(proposals_month_values),
        'extensions_month_labels': json.dumps(extensions_month_labels),
        'extensions_month_values': json.dumps(extensions_month_values),
        
        # Summary Stats
        'total_users': total_users,
        'total_proposals': total_proposals,
        'total_projects': total_projects,
        'pending_extensions': pending_extensions,
        
        # Task Completion Stats
        'task_pending': task_pending,
        'task_in_progress': task_in_progress,
        'task_completed': task_completed,
        'task_overdue': task_overdue,
        
        # Activity Feed
        'recent_activities': recent_activities,
    }

    return render(request, 'administrator/dashboard.html', context)


@login_required
def administrator_quick_actions_view(request):
    if request.user.role != 'admin' and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    context = {
        'pending_extensions': ExtensionRequest.objects.filter(status='pending').count(),
        'pending_proposals': Proposal.objects.filter(status='pending').count(),
        'ongoing_projects': Project.objects.filter(status='ongoing').count(),
    }
    return render(request, 'administrator/quick_actions.html', context)


@login_required
@require_POST
def assistant_ai_chat_view(request):
    """Return AI-assisted response using google-genai with local context only."""
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON payload.', 'fallback': True}, status=400)

    question = (payload.get('question') or '').strip()
    context = payload.get('context') or {}

    if not question:
        return JsonResponse({'ok': False, 'error': 'Question is required.', 'fallback': True}, status=400)

    api_key = os.environ.get('GOOGLE_API_KEY')
    if not api_key:
        return JsonResponse({'ok': False, 'error': 'AI key not configured.', 'fallback': True}, status=200)

    model_name = os.environ.get('GOOGLE_GENAI_MODEL', 'gemini-2.0-flash')
    system_prompt = (
        "You are the TaskPro Assistant. Answer strictly using the provided system summary. "
        "If the summary does not include the requested data, say it is not available. "
        "Keep answers concise and actionable. Do not invent records."
    )
    prompt = (
        f"Question: {question}\n"
        f"System summary (JSON): {json.dumps(context)}"
    )

    try:
        from google import genai

        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=model_name,
            contents=[system_prompt, prompt],
        )
        answer = (getattr(response, 'text', None) or '').strip()
        if not answer:
            return JsonResponse({'ok': False, 'error': 'Empty response from AI.', 'fallback': True}, status=200)
        return JsonResponse({'ok': True, 'answer': answer, 'source': 'ai'})
    except Exception as exc:
        return JsonResponse({'ok': False, 'error': f'AI error: {exc}', 'fallback': True}, status=200)
# ----------------------------
# Administrator User Management Views
# ----------------------------
def administrator_users_view(request):
    users = User.objects.all().order_by('date_joined')
    return render(request, 'administrator/users.html', {'users': users})

# -------------------------
# Add User
# -------------------------
def administrator_users_add_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        middle_name = request.POST.get('middle_name')
        last_name = request.POST.get('last_name')
        suffix = request.POST.get('suffix')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        contact_number = request.POST.get('contact_number')
        address = request.POST.get('address')
        sex = request.POST.get('sex')
        civil_status = request.POST.get('civil_status')
        role = request.POST.get('role')
        profile_picture = request.FILES.get('profile_picture')

        # Use email as username
        username = email

        # Validate profile picture if provided
        if profile_picture:
            try:
                validate_profile_picture(profile_picture)
            except ValidationError as e:
                messages.error(request, f"Profile picture error: {e.messages[0] if e.messages else str(e)}")
                return redirect('administrator_users_url')

        # Validate password strength
        try:
            validate_password_strength(password)
        except ValidationError as e:
            messages.error(request, f"Password error: {e.messages[0] if e.messages else str(e)}")
            return redirect('administrator_users_url')

        # Validate passwords match
        if password != confirm_password:
            messages.error(request, "Passwords do not match!")
            return redirect('administrator_users_url')

        # Duplicate check
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists!")
            return redirect('administrator_users_url')

        # Create user
        user = User.objects.create(
            username=username,
            email=email,
            first_name=first_name,
            middle_name=middle_name,
            last_name=last_name,
            suffix=suffix,
            contact_number=contact_number,
            address=address,
            sex=sex,
            civil_status=civil_status,
            role=role,
            profile_picture=profile_picture,
            password=make_password(password),
        )

        # -------------------------------
        # AUDIT LOG: CREATE USER
        AuditLog.objects.create(
            user=request.user,
            action="create",
            model_name="User",
            object_id=str(user.pk),
            old_data=None,
            new_data={
                "username": user.username,
                "email": user.email,
                "role": user.role,
                "created_by": request.user.username,
                "ip_address": request.META.get("REMOTE_ADDR"),
            }
        )
        # -------------------------------

        messages.success(request, f"User '{user.email}' added successfully!")
        return redirect('administrator_users_url')

    return redirect('administrator_users_url')


# -------------------------
# Update User
# -------------------------
def administrator_users_update_view(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':

        # Save OLD DATA before changes
        old_data = {
            "email": user.email,
            "role": user.role,
            "address": user.address,
            "contact_number": user.contact_number,
        }

        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        middle_name = request.POST.get('middle_name')
        last_name = request.POST.get('last_name')
        suffix = request.POST.get('suffix')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        contact_number = request.POST.get('contact_number')
        address = request.POST.get('address')
        sex = request.POST.get('sex')
        civil_status = request.POST.get('civil_status')
        role = request.POST.get('role')
        profile_picture = request.FILES.get('profile_picture')

        # Use email as username
        username = email

        # Duplicate check
        if User.objects.filter(email=email).exclude(id=user.id).exists():
            messages.error(request, "Email already exists!")
            return redirect('administrator_users_url')

        # Update fields
        user.username = username
        user.email = email
        user.first_name = first_name
        user.middle_name = middle_name
        user.last_name = last_name
        user.suffix = suffix
        user.contact_number = contact_number
        user.address = address
        user.sex = sex
        user.civil_status = civil_status
        user.role = role
        
        # TNA fields (for beneficiaries)
        if role == 'beneficiary':
            tna_status = request.POST.get('tna_status')
            tna_completion_date = request.POST.get('tna_completion_date')
            tna_notes = request.POST.get('tna_notes')
            
            if tna_status:
                user.tna_status = tna_status
            if tna_completion_date:
                user.tna_completion_date = tna_completion_date
            else:
                user.tna_completion_date = None
            if tna_notes:
                user.tna_notes = tna_notes

        if profile_picture:
            user.profile_picture = profile_picture

        # Password update
        if password:
            if password != confirm_password:
                messages.error(request, "Passwords do not match!")
                return redirect('administrator_users_url')
            user.password = make_password(password)

        user.save()

        # -------------------------------
        # AUDIT LOG: UPDATE USER
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=str(user.pk),
            old_data=old_data,
            new_data={
                "email": user.email,
                "role": user.role,
                "updated_by": request.user.username,
                "ip_address": request.META.get("REMOTE_ADDR"),
            }
        )
        # -------------------------------

        messages.success(request, f"User '{user.email}' updated successfully!")
        return redirect('administrator_users_url')

    return redirect('administrator_users_url')

# --------------------------
# Administrator: Budget List
# --------------------------
def administrator_budgets_view(request):
    from django.db.models import Sum, Q
    from decimal import Decimal
    import json
    
    # Get fund filter from request
    selected_fund = request.GET.get('fund', 'all')
    
    # Get all budgets or filter by fund
    if selected_fund == 'all':
        budgets = Budget.objects.all().order_by('-date_created')
    else:
        budgets = Budget.objects.filter(fund_source=selected_fund).order_by('-date_created')
    
    # Get all budgets for summary statistics (always show totals for all funds)
    all_budgets = Budget.objects.all()
    
    # Summary statistics (filtered or all funds based on selection)
    if selected_fund == 'all':
        # Show totals for all funds
        total_budget = Budget.objects.aggregate(total=Sum('total_equipment_value'))['total'] or Decimal('0.00')
        total_delivered = Budget.objects.aggregate(total=Sum('delivered_equipment_value'))['total'] or Decimal('0.00')
        total_spent = total_delivered
        total_remaining = total_budget - total_spent
        utilization_rate = round((float(total_spent) / float(total_budget) * 100), 2) if total_budget > 0 else 0
    else:
        # Show totals for selected fund only
        filtered_budgets = Budget.objects.filter(fund_source=selected_fund)
        total_budget = filtered_budgets.aggregate(total=Sum('total_equipment_value'))['total'] or Decimal('0.00')
        total_delivered = filtered_budgets.aggregate(total=Sum('delivered_equipment_value'))['total'] or Decimal('0.00')
        total_spent = total_delivered
        total_remaining = total_budget - total_spent
        utilization_rate = round((float(total_spent) / float(total_budget) * 100), 2) if total_budget > 0 else 0
    
    # Budget breakdown by fund source (filtered or all)
    fund_source_data = {}
    for budget in budgets:
        source = budget.fund_source
        if source not in fund_source_data:
            fund_source_data[source] = {'total': 0, 'remaining': 0, 'spent': 0}
        fund_source_data[source]['total'] += float(budget.total_amount)
        fund_source_data[source]['remaining'] += float(budget.remaining_amount)
        fund_source_data[source]['spent'] += float(budget.total_amount - budget.remaining_amount)
    
    # Budget breakdown by fiscal year (filtered or all)
    fiscal_year_data = {}
    for budget in budgets:
        year = str(budget.fiscal_year)
        if year not in fiscal_year_data:
            fiscal_year_data[year] = {'total': 0, 'remaining': 0, 'spent': 0}
        fiscal_year_data[year]['total'] += float(budget.total_amount)
        fiscal_year_data[year]['remaining'] += float(budget.remaining_amount)
        fiscal_year_data[year]['spent'] += float(budget.total_amount - budget.remaining_amount)
    
    # Status breakdown (filtered or all) - equipment-focused statuses
    status_counts = {
        'pending_procurement': 0, 
        'available': 0, 
        'partially_allocated': 0, 
        'fully_allocated': 0, 
        'completed': 0,
        'archived': 0,
        # Legacy status mappings for existing data
        'active': 0,
        'exhausted': 0
    }
    for budget in budgets:
        if budget.status in status_counts:
            status_counts[budget.status] += 1
    
    # Get all unique fund sources for the filter dropdown
    all_fund_sources = Budget.objects.values_list('fund_source', flat=True).distinct().order_by('fund_source')
    
    # Detailed budget allocation data - showing where each budget is used (filtered or all)
    budget_allocations = []
    for budget in budgets:
        # Get projects using this budget
        projects = Project.objects.filter(budget=budget).select_related('proposal')
        proposals = Proposal.objects.filter(budget=budget, status='approved')
        
        allocation_details = []
        
        # Add equipment allocations
        equipment_allocations = BudgetAllocation.objects.filter(budget=budget).select_related('equipment_item', 'allocated_by')
        for eq_alloc in equipment_allocations:
            allocation_details.append({
                'type': 'equipment',
                'id': eq_alloc.id,
                'title': f"{eq_alloc.equipment_item.name} ({eq_alloc.allocated_quantity} {eq_alloc.equipment_item.unit})",
                'amount': float(eq_alloc.allocated_value),
                'location': 'Equipment Allocation',
                'status': eq_alloc.status,
                'beneficiary': eq_alloc.allocated_by.full_name if eq_alloc.allocated_by else 'System'
            })
        
        # Add project allocations
        for project in projects:
            allocation_details.append({
                'type': 'project',
                'id': project.id,
                'title': project.project_title,
                'amount': float(project.funds or 0),
                'location': f"{project.mun or ''} {project.province or ''}".strip() or project.beneficiary_address or 'N/A',
                'status': project.status,
                'beneficiary': project.beneficiary or 'N/A'
            })
        
        # Add approved proposal allocations (not yet converted to projects)
        for proposal in proposals:
            # Only include proposals that haven't been converted to projects yet
            if not hasattr(proposal, 'project') or proposal.project is None:
                allocation_details.append({
                    'type': 'proposal',
                    'id': proposal.id,
                    'title': proposal.title,
                    'amount': float(proposal.approved_amount or 0),
                    'location': f"{proposal.municipality or ''} {proposal.province or ''}".strip() or proposal.location or 'N/A',
                    'status': 'approved',
                    'beneficiary': proposal.beneficiary.full_name if proposal.beneficiary else 'N/A'
                })
        
        budget_allocations.append({
            'budget_id': budget.id,
            'fund_source': budget.fund_source,
            'fiscal_year': budget.fiscal_year,
            'status': budget.status,
            'total_amount': float(budget.total_amount),
            'remaining_amount': float(budget.remaining_amount),
            'allocated_amount': float(budget.total_amount - budget.remaining_amount),
            'allocations': allocation_details,
            'allocation_count': len(allocation_details)
        })
    
    context = {
        'budgets': budgets,
        'total_budget': int(total_budget),
        'total_spent': int(total_spent),
        'total_remaining': int(total_remaining),
        'utilization_rate': utilization_rate,
        'budget_count': budgets.count(),
        'fund_source_labels': json.dumps(list(fund_source_data.keys())),
        'fund_source_totals': json.dumps([d['total'] for d in fund_source_data.values()]),
        'fund_source_spent': json.dumps([d['spent'] for d in fund_source_data.values()]),
        'fund_source_remaining': json.dumps([d['remaining'] for d in fund_source_data.values()]),
        'fiscal_year_labels': json.dumps(list(fiscal_year_data.keys())),
        'fiscal_year_spent': json.dumps([d['spent'] for d in fiscal_year_data.values()]),
        'fiscal_year_remaining': json.dumps([d['remaining'] for d in fiscal_year_data.values()]),
        'status_counts': json.dumps(status_counts),
        'budget_allocations': budget_allocations,
        'selected_fund': selected_fund,
        'all_fund_sources': all_fund_sources,
    }
    
    return render(request, 'administrator/budgets.html', context)

# --------------------------
# Administrator: Add Budget
# --------------------------
def administrator_budgets_add_view(request):
    if request.method == 'POST':
        fiscal_year = request.POST.get('fiscal_year')
        fund_source = request.POST.get('fund_source')
        total_amount = request.POST.get('total_amount') or '0.00'
        delivered_amount = request.POST.get('delivered_amount') or '0.00'
        budget_documents = request.FILES.getlist('budget_documents')  # Multiple files
        
        # Validate documents before processing
        for doc in budget_documents:
            try:
                validate_document_upload(doc)
            except ValidationError as e:
                messages.error(request, f"Document '{doc.name}': {e.messages[0] if e.messages else str(e)}")
                return redirect('administrator_budgets_url')
        
        # Validate amounts
        try:
            total_equipment_value = validate_positive_decimal(total_amount, 'Total Amount')
            delivered_equipment_value = validate_positive_decimal(delivered_amount, 'Delivered Amount')
        except ValidationError as e:
            messages.error(request, e.messages[0] if e.messages else str(e))
            return redirect('administrator_budgets_url')
        
        try:
            budget = Budget.objects.create(
                fiscal_year=fiscal_year,
                fund_source=fund_source,
                total_equipment_value=total_equipment_value,
                delivered_equipment_value=delivered_equipment_value,
                created_by=request.user
            )
            
            # Handle multiple document uploads
            for doc in budget_documents:
                BudgetDocument.objects.create(
                    budget=budget,
                    file=doc,
                    uploaded_by=request.user
                )

            # -------------------------------
            # AUDIT LOG: CREATE BUDGET
            from .models import AuditLog

            AuditLog.objects.create(
                user=request.user,
                action="create",
                model_name="Budget",
                object_id=str(budget.pk),
                old_data=None,
                new_data={
                    "fiscal_year": budget.fiscal_year,
                    "fund_source": budget.fund_source,
                    "total_equipment_value": str(budget.total_equipment_value),
                    "delivered_equipment_value": str(budget.delivered_equipment_value),
                    "documents_count": budget.documents.count(),
                    "created_by": request.user.username,
                    "ip_address": request.META.get("REMOTE_ADDR"),
                }
            )
            # -------------------------------

            messages.success(request, 'Budget added successfully.')
        except ValidationError as e:
            messages.error(request, f"Error: {e}")
        return redirect('administrator_budgets_url')

# --------------------------
# Administrator: Edit Budget
# --------------------------
def administrator_budgets_update_view(request, budget_id):
    budget = get_object_or_404(Budget, id=budget_id)
    if request.method == 'POST':

        # Save OLD DATA before changes
        old_data = {
            "fiscal_year": budget.fiscal_year,
            "fund_source": budget.fund_source,
            "total_equipment_value": str(budget.total_equipment_value),
            "delivered_equipment_value": str(budget.delivered_equipment_value),
            "documents_count": budget.documents.count(),
        }

        fiscal_year = request.POST.get('fiscal_year')
        fund_source = request.POST.get('fund_source')
        total_amount = request.POST.get('total_amount') or '0.00'
        delivered_amount = request.POST.get('delivered_amount') or '0.00'
        budget_documents = request.FILES.getlist('budget_documents')  # Multiple files
        
        # Calculate equipment values
        total_equipment_value = Decimal(total_amount)
        delivered_equipment_value = Decimal(delivered_amount)
        
        try:
            budget.fiscal_year = fiscal_year
            budget.fund_source = fund_source
            budget.total_equipment_value = total_equipment_value
            budget.delivered_equipment_value = delivered_equipment_value
            budget.full_clean()
            budget.save()
            
            # Handle multiple document uploads (add to existing)
            for doc in budget_documents:
                BudgetDocument.objects.create(
                    budget=budget,
                    file=doc,
                    uploaded_by=request.user
                )

            # -------------------------------
            # AUDIT LOG: UPDATE BUDGET
            from .models import AuditLog

            AuditLog.objects.create(
                user=request.user,
                action="update",
                model_name="Budget",
                object_id=str(budget.pk),
                old_data=old_data,
                new_data={
                    "fiscal_year": budget.fiscal_year,
                    "fund_source": budget.fund_source,
                    "total_equipment_value": str(budget.total_equipment_value),
                    "delivered_equipment_value": str(budget.delivered_equipment_value),
                    "documents_count": budget.documents.count(),
                    "updated_by": request.user.username,
                    "ip_address": request.META.get("REMOTE_ADDR"),
                }
            )
            # -------------------------------

            messages.success(request, 'Budget updated successfully.')
        except ValidationError as e:
            messages.error(request, f"Error: {e}")
        return redirect('administrator_budgets_url')

def administrator_budgets_delete_view(request, budget_id):
    budget = get_object_or_404(Budget, id=budget_id)

    # Save old data before deletion
    old_data = {
        "fiscal_year": budget.fiscal_year,
        "fund_source": budget.fund_source,
        "total_equipment_value": str(budget.total_equipment_value),
        "delivered_equipment_value": str(budget.delivered_equipment_value),
        "budget_document": budget.budget_document.url if budget.budget_document else "",
        "created_by": budget.created_by.username if budget.created_by else None,
    }

    try:
        # Check for related objects before attempting deletion
        related_proposals = budget.proposals.all()
        related_projects = budget.projects.all()

        if related_proposals.exists() or related_projects.exists():
            # Build detailed error message
            related_items = []
            if related_proposals.exists():
                proposal_titles = [f"'{p.title}'" for p in related_proposals[:3]]  # Show first 3
                if related_proposals.count() > 3:
                    proposal_titles.append(f"and {related_proposals.count() - 3} more")
                related_items.append(f"Proposals: {', '.join(proposal_titles)}")

            if related_projects.exists():
                project_titles = [f"'{p.project_title}'" for p in related_projects[:3]]  # Show first 3
                if related_projects.count() > 3:
                    project_titles.append(f"and {related_projects.count() - 3} more")
                related_items.append(f"Projects: {', '.join(project_titles)}")

            error_msg = f"Cannot delete budget '{budget.fund_source}' because it is referenced by: {'; '.join(related_items)}. "
            error_msg += "Please reassign or remove these related items first, or contact an administrator."

            messages.error(request, error_msg)
            return redirect('administrator_budgets_url')

        budget.delete()

        # -------------------------------
        # AUDIT LOG: DELETE BUDGET
        AuditLog.objects.create(
            user=request.user,
            action="delete",
            model_name="Budget",
            object_id=str(budget_id),
            old_data=old_data,
            new_data=None,
        )
        # -------------------------------

        messages.success(request, f"Budget '{old_data['fund_source']}' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting budget: {e}")

    return redirect('administrator_budgets_url')


# --------------------------
# Delete Document Views
# --------------------------
@login_required
def delete_budget_document(request, document_id):
    """Delete a single budget document"""
    doc = get_object_or_404(BudgetDocument, id=document_id)
    budget_id = doc.budget.id
    doc.delete()
    messages.success(request, 'Document deleted successfully.')
    return redirect('administrator_budgets_url')


@login_required
def delete_proposal_document(request, document_id):
    """Delete a single proposal document"""
    doc = get_object_or_404(ProposalDocument, id=document_id)
    proposal_id = doc.proposal.id
    doc.delete()
    messages.success(request, 'Document deleted successfully.')
    return redirect('administrator_proposals_url')


@login_required
def delete_project_document(request, document_id):
    """Delete a single project document"""
    doc = get_object_or_404(ProjectDocument, id=document_id)
    project_id = doc.project.id
    doc.delete()
    messages.success(request, 'Document deleted successfully.')
    return redirect('administrator_projects_url')


@login_required 
def delete_expense_document(request, document_id):
    """Delete a single expense document"""
    doc = get_object_or_404(ExpenseDocument, id=document_id)
    expense_id = doc.expense.id
    doc.delete()
    messages.success(request, 'Receipt deleted successfully.')
    return redirect('administrator_projects_url')


# -------------------------
# List Proposals
# -------------------------
@login_required
def administrator_proposals_view(request):
    proposals = Proposal.objects.select_related('submitted_by', 'budget', 'proponent', 'beneficiary', 'project').all().order_by('submission_date')
    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])
    
    # Get proponents and beneficiaries for the dropdowns
    proponents = User.objects.filter(role='proponent', status='active')
    beneficiaries = User.objects.filter(role='beneficiary', status='active')
    
    # Get all signatures for proposals
    proposal_signatures = DigitalSignature.objects.filter(content_type='proposal')
    signed_proposal_ids = set(proposal_signatures.values_list('object_id', flat=True))

    # Chart Data: Proposal Status Distribution
    status_counts = Proposal.objects.values('status').annotate(count=Count('id'))
    status_map = {'pending': 'Pending', 'for_review': 'For Review', 'approved': 'Approved', 'rejected': 'Rejected', 'needs_revision': 'Needs Revision'}
    proposal_status_labels = [status_map.get(s['status'], s['status'] or 'Unknown') for s in status_counts]
    proposal_status_values = [s['count'] for s in status_counts]

    # Chart Data: Top Proposals by Amount
    top_proposals = Proposal.objects.order_by('-proposed_amount')[:10]
    proposal_amount_labels = []
    proposal_amount_values = []
    for p in top_proposals:
        title = p.title[:25] + '...' if p.title and len(p.title) > 25 else (p.title or 'No Title')
        proposal_amount_labels.append(title)
        proposal_amount_values.append(float(p.proposed_amount) if p.proposed_amount else 0.0)

    context = {
        'proposals': proposals,
        'budgets': budgets,
        'proponents': proponents,
        'beneficiaries_list': beneficiaries,
        'signed_proposal_ids': signed_proposal_ids,
        'proposal_status_labels': json.dumps(proposal_status_labels),
        'proposal_status_values': json.dumps(proposal_status_values),
        'proposal_amount_labels': json.dumps(proposal_amount_labels),
        'proposal_amount_values': json.dumps(proposal_amount_values),
    }
    return render(request, 'administrator/proposals.html', context)

# -------------------------
# Add Proposal
# -------------------------
def administrator_proposals_add_view(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        proposed_amount = request.POST.get('proposed_amount') or 0
        budget_id = request.POST.get('budget')
        documents = request.FILES.getlist('documents')  # Multiple files
        
        # Validate documents before processing
        for doc in documents:
            try:
                validate_document_upload(doc)
            except ValidationError as e:
                messages.error(request, f"Document '{doc.name}': {e.messages[0] if e.messages else str(e)}")
                return redirect('administrator_proposals_url')
        
        # Validate proposed amount
        try:
            proposed_amount = validate_positive_decimal(proposed_amount, 'Proposed Amount')
        except ValidationError as e:
            messages.error(request, e.messages[0] if e.messages else str(e))
            return redirect('administrator_proposals_url')
        
        # New fields
        proponent_id = request.POST.get('proponent')
        beneficiary_id = request.POST.get('beneficiary')
        location = request.POST.get('location', '')
        municipality = request.POST.get('municipality', '')
        province = request.POST.get('province', '')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')

        budget = Budget.objects.get(pk=budget_id) if budget_id else None
        proponent = User.objects.get(pk=proponent_id) if proponent_id else None
        beneficiary = User.objects.get(pk=beneficiary_id) if beneficiary_id else None

        proposal = Proposal(
            title=title,
            description=description,
            submitted_by=request.user,
            proposed_amount=Decimal(str(proposed_amount)),
            budget=budget,
            proponent=proponent,
            beneficiary=beneficiary,
            location=location,
            municipality=municipality,
            province=province,
            latitude=latitude if latitude else None,
            longitude=longitude if longitude else None,
        )

        try:
            proposal.full_clean()
            proposal.save()
            
            # Handle multiple document uploads
            for doc in documents:
                ProposalDocument.objects.create(
                    proposal=proposal,
                    file=doc,
                    uploaded_by=request.user
                )

            # -------------------------
            # Audit Log
            # -------------------------
            from .models import AuditLog  # adjust path if needed

            AuditLog.objects.create(
                user=request.user,
                action="create",
                model_name="Proposal",
                object_id=str(proposal.pk),
                old_data=None,
                new_data={
                    "title": proposal.title,
                    "description": proposal.description,
                    "proposed_amount": str(proposal.proposed_amount),
                    "budget": budget.fund_source if budget else None,
                    "documents_count": proposal.documents.count(),
                    "proponent": proponent.full_name() if proponent else None,
                    "beneficiary": beneficiary.full_name() if beneficiary else None,
                    "location": location,
                    "municipality": municipality,
                    "province": province,
                }
            )
            # -------------------------

            messages.success(request, 'Proposal added successfully.')
        except ValidationError as e:
            messages.error(request, e)
        
    return redirect('administrator_proposals_url')

# -------------------------
# Update Proposal
# -------------------------
def administrator_proposals_update_view(request, pk):
    proposal = get_object_or_404(Proposal, pk=pk)
    if request.method == 'POST':
        proposal.title = request.POST.get('title')
        proposal.description = request.POST.get('description')

        proposed_amount_raw = (request.POST.get('proposed_amount') or '0').replace(',', '').strip()
        approved_amount_raw = (request.POST.get('approved_amount') or '').replace(',', '').strip()
        try:
            proposal.proposed_amount = Decimal(proposed_amount_raw or '0')
            proposal.approved_amount = Decimal(approved_amount_raw) if approved_amount_raw else None
        except Exception:
            messages.error(request, 'Invalid amount format. Please enter a valid numeric value.')
            return redirect('administrator_proposals_url')

        budget_id = request.POST.get('budget')
        if budget_id:
            try:
                proposal.budget = Budget.objects.get(pk=budget_id)
            except Budget.DoesNotExist:
                messages.error(request, 'Selected budget was not found.')
                return redirect('administrator_proposals_url')
        else:
            proposal.budget = None

        proposal.status = request.POST.get('status')
        
        # Handle multiple document uploads (add to existing)
        documents = request.FILES.getlist('documents')
        for doc in documents:
            ProposalDocument.objects.create(
                proposal=proposal,
                file=doc,
                uploaded_by=request.user
            )
        
        # New fields
        proponent_id = request.POST.get('proponent')
        beneficiary_id = request.POST.get('beneficiary')
        try:
            proposal.proponent = User.objects.get(pk=proponent_id) if proponent_id else None
        except User.DoesNotExist:
            messages.error(request, 'Selected proponent was not found.')
            return redirect('administrator_proposals_url')

        try:
            proposal.beneficiary = User.objects.get(pk=beneficiary_id) if beneficiary_id else None
        except User.DoesNotExist:
            messages.error(request, 'Selected beneficiary was not found.')
            return redirect('administrator_proposals_url')

        proposal.beneficiaries = request.POST.get('beneficiaries', '')
        proposal.location = request.POST.get('location', '')
        proposal.municipality = request.POST.get('municipality', '')
        proposal.province = request.POST.get('province', '')

        try:
            proposal.full_clean()
            proposal.save()
            messages.success(request, 'Proposal updated successfully.')
        except ValidationError as e:
            messages.error(request, e)

    return redirect('administrator_proposals_url')


@login_required
def administrator_proposals_approve_view(request, pk):
    proposal = get_object_or_404(Proposal, pk=pk)

    if request.method == 'POST':
        # Capture old data for audit
        old_data = {
            "title": proposal.title,
            "description": proposal.description,
            "proposed_amount": str(proposal.proposed_amount),
            "approved_amount": str(proposal.approved_amount) if proposal.approved_amount else None,
            "status": proposal.status,
            "budget": proposal.budget.fund_source if proposal.budget else None,
            "document": proposal.document.url if proposal.document else None,
            "processed_by": proposal.processed_by.get_full_name() if proposal.processed_by else None,
            "beneficiary": proposal.beneficiary.get_full_name() if proposal.beneficiary else None,
        }

        # Update status to approved
        proposal.status = 'approved'

        # Update approved amount if provided
        approved_amount = request.POST.get('approved_amount')
        if approved_amount:
            proposal.approved_amount = Decimal(approved_amount)

        # Parse project start and end dates
        project_start_str = request.POST.get('project_start_date')
        project_end_str = request.POST.get('project_end_date')
        project_start = datetime.strptime(project_start_str, '%Y-%m-%d').date() if project_start_str else None
        project_end = datetime.strptime(project_end_str, '%Y-%m-%d').date() if project_end_str else None

        try:
            proposal.full_clean()

            # Check budget availability
            if proposal.budget and proposal.approved_amount:
                if proposal.budget.remaining_amount < proposal.approved_amount:
                    messages.error(
                        request,
                        f"Cannot approve: Budget '{proposal.budget.fund_source}' has insufficient funds."
                    )
                    return redirect('administrator_proposals_url')
            else:
                messages.error(request, "Cannot approve: Proposal must have a budget and approved amount.")
                return redirect('administrator_proposals_url')

            # Save proposal (just approve it, don't create project yet)
            proposal.processed_by = request.user
            proposal.save()

            # -------------------------
            # Notifications
            # -------------------------
            # Helper function to get role-based link
            def get_role_link(user):
                if user.role == 'administrator':
                    return reverse('administrator_proposals_url')
                elif user.role == 'staff':
                    return reverse('staff_proposals_url')
                elif user.role == 'proponent':
                    return reverse('proponent_proposals_url')
                else:
                    return reverse('beneficiary_proposals_url')

            # Collect unique receivers (excluding admins - they don't need proposal status notifications)
            receivers_set = set()
            if proposal.submitted_by and proposal.submitted_by.role != 'administrator':
                receivers_set.add(proposal.submitted_by)
            if proposal.proponent and proposal.proponent.role != 'administrator':
                receivers_set.add(proposal.proponent)
            if proposal.beneficiary and proposal.beneficiary.role != 'administrator':
                receivers_set.add(proposal.beneficiary)

            for receiver in receivers_set:
                Notification.objects.create(
                    sender=request.user,
                    receiver=receiver,
                    message=f"Proposal '{proposal.title}' has been approved.",
                    category='proposal',
                    link=get_role_link(receiver)
                )

            # -------------------------
            # Audit Log
            # -------------------------
            AuditLog.objects.create(
                user=request.user,
                action="update",
                model_name="Proposal",
                object_id=str(proposal.pk),
                old_data=old_data,
                new_data={
                    "status": proposal.status,
                    "approved_amount": str(proposal.approved_amount),
                    "project_start": project_start.isoformat() if project_start else None,
                    "project_end": project_end.isoformat() if project_end else None,
                    "processed_by": proposal.processed_by.get_full_name() if proposal.processed_by else None,
                    "beneficiary": proposal.beneficiary.get_full_name() if proposal.beneficiary else None,
                }
            )

            messages.success(request, 'Proposal approved successfully.')

        except ValidationError as e:
            errors = []
            if hasattr(e, 'message_dict'):
                for field, messages_list in e.message_dict.items():
                    if field == '__all__':
                        errors.extend(messages_list)
                    else:
                        for msg in messages_list:
                            errors.append(f"{field.replace('_', ' ').capitalize()}: {msg}")
            elif hasattr(e, 'messages'):
                errors.extend(e.messages)

            formatted_errors = ". ".join(errors) + "." if errors else "Invalid data."
            messages.error(request, f"Failed to approve proposal. {formatted_errors}")

    return redirect('administrator_proposals_url')

@login_required
def administrator_proposals_decline_view(request):
    if request.method == 'POST':
        proposal_id = request.POST.get('proposal_id')
        reason = request.POST.get('decline_reason', '')

        proposal = get_object_or_404(Proposal, pk=proposal_id)

        old_data = {
            "title": proposal.title,
            "description": proposal.description,
            "status": proposal.status,
            "budget": proposal.budget.fund_source if proposal.budget else None,
            "document": proposal.document.url if proposal.document else None,
            "approved_amount": str(proposal.approved_amount) if proposal.approved_amount else None,
            "review_remarks": proposal.review_remarks,
        }

        proposal.status = 'rejected'
        proposal.review_remarks = reason
        proposal.save()

        # -------------------------
        # Notifications
        # -------------------------
        # Notify the submitter and proponent (if different) about the decline
        # Don't notify administrators - they are the ones deciding
        receivers = []
        
        # Add the person who submitted the proposal (if not an admin)
        if proposal.submitted_by and proposal.submitted_by.role != 'administrator':
            receivers.append(proposal.submitted_by)
        
        # Add the proponent (if different from submitter and not an admin)
        if proposal.proponent and proposal.proponent.role != 'administrator' and proposal.proponent not in receivers:
            receivers.append(proposal.proponent)

        for receiver in receivers:
            # Determine the correct link based on the receiver's role
            if receiver.role == 'staff':
                link = reverse('staff_proposals_url')
            elif receiver.role == 'proponent':
                link = reverse('proponent_proposals_url')
            else:
                link = reverse('beneficiary_proposals_url')
            
            Notification.objects.create(
                sender=request.user,
                receiver=receiver,
                message=f"Proposal '{proposal.title}' has been declined.",
                category='proposal',
                link=link
            )

        # -------------------------
        # Audit Log
        # -------------------------
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="Proposal",
            object_id=str(proposal.pk),
            old_data=old_data,
            new_data={"status": proposal.status, "review_remarks": reason}
        )

        messages.success(request, "Proposal declined successfully.")

    return redirect('administrator_proposals_url')
# ---------------------------
# Project List / Dashboard
# ---------------------------
@login_required
def administrator_projects_view(request):
    projects = Project.objects.select_related('proposal', 'budget', 'project_leader').all()
    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])
    proposals = Proposal.objects.filter(status='approved')
    equipment_categories = EquipmentCategory.objects.all()  # For equipment delivery form
    
    # Chart 1: Project Status
    project_status_counts = Project.objects.values('status').annotate(total=Count('id'))
    status_data = {}
    for entry in project_status_counts:
        label = entry['status'] or "Unknown"
        status_data[label] = entry['total']
    project_status_labels = list(status_data.keys())
    project_status_values = list(status_data.values())

    # Chart 2: Top 10 Projects by Funds
    top_projects = Project.objects.order_by('-funds')[:10]
    top_project_labels = []
    top_project_budgets = []
    for p in top_projects:
        title = p.project_title[:20] + '...' if p.project_title and len(p.project_title) > 20 else (p.project_title or "No Title")
        top_project_labels.append(title)
        top_project_budgets.append(float(p.funds) if p.funds else 0.0)

    # Additional aggregates for charts
    # Projects by Municipality
    projects_by_mun = Project.objects.values('mun').annotate(count=Count('id')).order_by('-count')
    mun_labels = [entry['mun'] or 'Unknown' for entry in projects_by_mun]
    mun_values = [entry['count'] for entry in projects_by_mun]

    # Funds by Fund Source
    funds_by_source = Project.objects.values('fund_source').annotate(total=Sum('funds')).order_by('-total')
    fund_source_labels = [entry['fund_source'] or 'Unknown' for entry in funds_by_source]
    fund_source_values = [float(entry['total'] or 0.0) for entry in funds_by_source]

    # Projects by Program
    projects_by_program = Project.objects.values('program').annotate(count=Count('id')).order_by('-count')
    program_labels = [entry['program'] or 'Unknown' for entry in projects_by_program]
    program_values = [entry['count'] for entry in projects_by_program]

    # Projects by Year
    projects_by_year = Project.objects.values('year').annotate(count=Count('id')).order_by('-year')
    year_labels = [str(entry['year'] or 'Unknown') for entry in projects_by_year]
    year_values = [entry['count'] for entry in projects_by_year]

    # ────────────────────────────────────────────
    # SERIALIZE PROJECTS TO JSON FOR ALPINE.JS
    # ────────────────────────────────────────────
    projects_data = []
    for p in projects:
        projects_data.append({
            'id': p.id,
            'no': p.no,
            'project_code': p.project_code or '',
            'year': p.year,
            'project_title': p.project_title or '',
            'agency_grantee': p.agency_grantee or '',
            'program': p.program or '',
            'type_of_project': p.type_of_project or '',
            'status': p.status or '',
            'remarks': p.remarks or '',
            'mun': p.mun or '',
            'province': p.province or 'Biliran',
            'district': p.district or 'Lone District',
            'beneficiary': p.beneficiary or '',
            'beneficiary_address': p.beneficiary_address or '',
            'contact_details': p.contact_details or '',
            'proponent_details': p.proponent_details or '',
            'no_of_beneficiaries': p.no_of_beneficiaries or 0,
            'male': p.male or 0,
            'female': p.female or 0,
            'total_beneficiaries': p.total_beneficiaries or 0,
            'senior_citizen': p.senior_citizen or 0,
            'pwd': p.pwd or 0,
            'fund_source': p.fund_source or '',
            'funds': float(p.funds) if p.funds else 0.0,
            'total_project_cost': float(p.total_project_cost) if p.total_project_cost else 0.0,
            'counterpart_funds': float(p.counterpart_funds) if p.counterpart_funds else 0.0,
            'internally_managed_fund': float(p.internally_managed_fund) if p.internally_managed_fund else 0.0,
            'total_funds_released': float(p.total_funds_released) if p.total_funds_released else 0.0,
            'first_tranche': float(p.first_tranche) if p.first_tranche else 0.0,
            'second_tranche': float(p.second_tranche) if p.second_tranche else 0.0,
            'third_tranche': float(p.third_tranche) if p.third_tranche else 0.0,
            'dost_viii': float(p.dost_viii) if p.dost_viii else 0.0,
            'project_start': p.project_start.isoformat() if p.project_start else '',
            'project_end': p.project_end.isoformat() if p.project_end else '',
            'date_of_release': p.date_of_release.isoformat() if p.date_of_release else '',
            'date_of_completion': p.date_of_completion.isoformat() if p.date_of_completion else '',
            'date_of_donation': p.date_of_donation.isoformat() if p.date_of_donation else '',
            'date_of_inspection_tagging': p.date_of_inspection_tagging.isoformat() if p.date_of_inspection_tagging else '',
            'date_of_liquidation': p.date_of_liquidation.isoformat() if p.date_of_liquidation else '',
            'check_ada_no': p.check_ada_no or '',
            'status_of_liquidation': p.status_of_liquidation or '',
            'amount_liquidated': float(p.amount_liquidated) if p.amount_liquidated else 0.0,
            'original_project_duration': p.original_project_duration or '',
            'extension_date': p.extension_date or '',
            'availed_technologies': p.availed_technologies or '',
            'interventions': p.interventions or '',
            'donation_status': p.donation_status or '',
            'tafr': p.tafr or '',
            'par': p.par or '',
            'list_of_eqpt': p.list_of_eqpt or '',
            'terminal_report': p.terminal_report or '',
            'invoice_receipt': p.invoice_receipt or '',
            'donated': p.donated or '',
            'acknowledgment_receipt_by_grantee': p.acknowledgment_receipt_by_grantee or '',
            'pme_visit': p.pme_visit or '',
            'womens_group': p.womens_group or '',
            'product_photo': p.product_photo.url if p.product_photo else '',
            'supporting_documents': p.supporting_documents.url if p.supporting_documents else '',
            # DOST Equipment Tracking Summary
            'equipment_summary': [
                {
                    'name': eq.budget_allocation.equipment_item.name if eq.budget_allocation and eq.budget_allocation.equipment_item else 'Unknown Equipment',
                    'quantity': eq.delivered_quantity,
                    'unit_cost': float(eq.budget_allocation.equipment_item.estimated_unit_cost or 0) if eq.budget_allocation and eq.budget_allocation.equipment_item else 0.0,
                    'property_tag': eq.property_tag_number or '',
                    'ownership_status': eq.ownership_status or 'dost_owned',
                    'ownership_status_display': eq.get_ownership_status_display() if hasattr(eq, 'get_ownership_status_display') else eq.ownership_status,
                    'delivery_date': eq.delivery_date.isoformat() if eq.delivery_date else '',
                } for eq in p.equipment_deliveries.select_related('budget_allocation__equipment_item').all()
            ] if hasattr(p, 'equipment_deliveries') else [],
        })

    context = {
        'projects': projects,
        'projects_json': json.dumps(projects_data),
        'mun_labels': json.dumps(mun_labels),
        'mun_values': json.dumps(mun_values),
        'fund_source_labels': json.dumps(fund_source_labels),
        'fund_source_values': json.dumps(fund_source_values),
        'program_labels': json.dumps(program_labels),
        'program_values': json.dumps(program_values),
        'year_labels': json.dumps(year_labels),
        'year_values': json.dumps(year_values),
        'budgets': budgets,
        'budgets_json': json.dumps([{
            'id': b.id,
            'fund_source': b.fund_source,
            'fiscal_year': b.fiscal_year,
            'total_amount': float(b.total_amount),
            'remaining_amount': float(b.remaining_amount),
            'status': b.status,
        } for b in budgets]),
        'proposals': proposals,
        'equipment_categories': equipment_categories,
        'project_status_labels': json.dumps(project_status_labels),
        'project_status_values': json.dumps(project_status_values),
        'top_project_labels': json.dumps(top_project_labels),
        'top_project_budgets': json.dumps(top_project_budgets),
    }

    return render(request, 'administrator/projects.html', context)


# ---------------------------
# Project Detail View
# ---------------------------
@login_required
def administrator_projects_detail_view(request, pk):
    project = get_object_or_404(Project, pk=pk)
    
    # Get all signatures for this project
    project_signatures = DigitalSignature.objects.filter(
        content_type='project',
        object_id=str(pk)
    ).select_related('user').order_by('-signed_at')
    
    context = {
        'project': project,
        'proposal': project.proposal,
        'budget': project.budget,
        'project_leader': project.project_leader,
        'project_signatures': project_signatures,
    }
    
    return render(request, 'administrator/project_detail.html', context)


# ---------------------------
# Add Project
# ---------------------------
@login_required
def administrator_projects_add_view(request):
    if request.method == 'POST':
        project_title = request.POST.get('project_title')
        project_description = request.POST.get('project_description')
        funds_input = request.POST.get('funds')
        funds = Decimal(funds_input) if funds_input else Decimal('0.00')
        budget_id = request.POST.get('budget')
        project_leader_id = request.POST.get('project_leader')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        proposal_id = request.POST.get('proposal')
        supporting_documents = request.FILES.get('supporting_documents')

        budget = Budget.objects.get(id=budget_id) if budget_id else None
        proposal = Proposal.objects.get(id=proposal_id) if proposal_id else None
        project_leader = request.user if not project_leader_id else None  # Modify if leader selection implemented

        # ============================================
        # SYNC: Deduct from budget when creating project
        # ============================================
        if budget and funds > Decimal('0.00'):
            if budget.remaining_amount >= funds:
                budget.remaining_amount -= funds
                if budget.remaining_amount == Decimal('0.00'):
                    budget.status = 'fully_allocated'
                elif budget.status in ['available', 'active']:
                    budget.status = 'partially_allocated'
                budget.save(update_fields=['remaining_amount', 'status', 'date_updated'])
            else:
                messages.error(request, f'Insufficient budget! Available: ₱{budget.remaining_amount:,.2f}, Requested: ₱{funds:,.2f}')
                return redirect('administrator_projects_url')
        # ============================================

        project = Project.objects.create(
            project_title=project_title,
            project_description=project_description,
            funds=funds if funds > Decimal('0.00') else None,
            budget=budget,
            proposal=proposal,
            project_leader=project_leader,
            project_start=start_date or None,
            project_end=end_date or None,
            supporting_documents=supporting_documents,
            # Copy location data from proposal
            mun=proposal.municipality if proposal else None,
            province=proposal.province if proposal else None,
            beneficiary_address=proposal.location if proposal else None,
            latitude=proposal.latitude if proposal else None,
            longitude=proposal.longitude if proposal else None
        )

        # -------------------------
        # Audit Log
        # -------------------------
        from .models import AuditLog  # adjust path if needed

        new_data = {
            "project_title": project.project_title,
            "project_description": project.project_description,
            "approved_budget": str(project.approved_budget),
            "budget": project.budget.fund_source if project.budget else None,
            "proposal": project.proposal.title if project.proposal else None,
            "project_leader": project.project_leader.username if project.project_leader else None,
            "start_date": str(project.start_date) if project.start_date else None,
            "end_date": str(project.end_date) if project.end_date else None,
            "supporting_documents": project.supporting_documents.url if project.supporting_documents else None,
        }

        AuditLog.objects.create(
            user=request.user,
            action="create",
            model_name="Project",
            object_id=str(project.pk),
            old_data=None,
            new_data=new_data
        )
        # -------------------------

        messages.success(request, 'Project added successfully.')
        return redirect('administrator_projects_url')

# -------------------------
# ADMINISTRATOR: Task Views
# -------------------------

def administrator_task_list_view(request):
    tasks = Task.objects.select_related('project', 'assigned_to').all()
    projects = Project.objects.all()
    # Filter users to only include DOST staff for task assignment
    users = User.objects.filter(role='dost_staff', status='active')

    context = {
        'tasks': tasks,
        'projects': projects,
        'users': users,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'category_choices': Task.CATEGORY_CHOICES,
    }
    return render(request, 'administrator/tasks.html', context)

# ---------------------------
# Administrator: Projects
# ---------------------------
# Note: The administrator_projects_view is defined earlier in this file (line ~882)

@login_required
def administrator_projects_add_view(request):
    if request.method == 'POST':
        # Helper function to safely get decimal values
        def get_decimal(name):
            val = request.POST.get(name)
            if val == '' or val is None:
                return None
            return Decimal(val)
        
        # Helper for integers
        def get_int(name):
            val = request.POST.get(name)
            if val == '' or val is None:
                return None
            return int(val)
            
        # Helper for dates (empty string returns None)
        def get_date(name):
            val = request.POST.get(name)
            return val if val else None

        proposal_id = request.POST.get('proposal')
        proposal = Proposal.objects.get(id=proposal_id) if proposal_id else None

        try:
            project = Project.objects.create(
                # Identification
                no=get_int('no'),
                project_code=request.POST.get('project_code'),
                year=get_int('year'),
                
                # Basic Info
                project_title=request.POST.get('project_title'),
                # Compatibility field: save description to both if possible, or just use title
                project_description=request.POST.get('project_title'), 
                agency_grantee=request.POST.get('agency_grantee'),
                program=request.POST.get('program'),
                type_of_project=request.POST.get('type_of_project'),
                status=request.POST.get('status'),
                remarks=request.POST.get('remarks'),
                
                # Location
                mun=request.POST.get('mun'),
                province=request.POST.get('province'),
                district=request.POST.get('district'),
                latitude=float(request.POST.get('latitude')) if request.POST.get('latitude') else None,
                longitude=float(request.POST.get('longitude')) if request.POST.get('longitude') else None,
                
                # Beneficiaries
                beneficiary=request.POST.get('beneficiary'),
                beneficiary_address=request.POST.get('beneficiary_address'),
                contact_details=request.POST.get('contact_details'),
                proponent_details=request.POST.get('proponent_details'),
                no_of_beneficiaries=get_int('no_of_beneficiaries'),
                male=get_int('male'),
                female=get_int('female'),
                total_beneficiaries=get_int('total_beneficiaries'),
                senior_citizen=get_int('senior_citizen'),
                pwd=get_int('pwd'),
                
                # Financials
                fund_source=request.POST.get('fund_source'),
                funds=get_decimal('funds'),
                # Compatibility: Map funds to approved_budget
                approved_budget=get_decimal('funds') or Decimal('0.00'),
                
                total_project_cost=get_decimal('total_project_cost'),
                counterpart_funds=get_decimal('counterpart_funds'),
                internally_managed_fund=get_decimal('internally_managed_fund'),
                total_funds_released=get_decimal('total_funds_released'),
                dost_viii=get_decimal('dost_viii'),
                counterpart_fund=get_decimal('counterpart_fund'),
                
                # Tranches
                first_tranche=get_decimal('first_tranche'),
                second_tranche=get_decimal('second_tranche'),
                third_tranche=get_decimal('third_tranche'),
                
                # Dates
                project_start=get_date('project_start'),
                project_end=get_date('project_end'),
                date_of_release=get_date('date_of_release'),
                date_of_completion=get_date('date_of_completion'),
                date_of_donation=get_date('date_of_donation'),
                date_of_inspection_tagging=get_date('date_of_inspection_tagging'),
                date_of_liquidation=get_date('date_of_liquidation'),
                
                # Others
                original_project_duration=request.POST.get('original_project_duration'),
                extension_date=request.POST.get('extension_date'),
                availed_technologies=request.POST.get('availed_technologies'),
                interventions=request.POST.get('interventions'),
                donation_status=request.POST.get('donation_status'),
                
                # Liquidation
                check_ada_no=request.POST.get('check_ada_no'),
                status_of_liquidation=request.POST.get('status_of_liquidation'),
                amount_liquidated=get_decimal('amount_liquidated'),
                
                # Docs Status
                tafr=request.POST.get('tafr'),
                par=request.POST.get('par'),
                list_of_eqpt=request.POST.get('list_of_eqpt'),
                terminal_report=request.POST.get('terminal_report'),
                invoice_receipt=request.POST.get('invoice_receipt'),
                donated=request.POST.get('donated'),
                acknowledgment_receipt_by_grantee=request.POST.get('acknowledgment_receipt_by_grantee'),
                
                # Files
                supporting_documents=request.FILES.get('supporting_documents'),
                product_photo=request.FILES.get('product_photo'),
                
                # Extra
                pme_visit=request.POST.get('pme_visit'),
                womens_group=request.POST.get('womens_group'),
                
                # Proposal link
                proposal=proposal
            )

            # Audit Log
            AuditLog.objects.create(
                user=request.user,
                action="create",
                model_name="Project",
                object_id=str(project.pk),
                old_data=None,
                new_data={
                    "project_title": project.project_title,
                    "funds": str(project.funds) if project.funds else "0",
                    "status": project.status
                }
            )
            messages.success(request, 'Project added successfully.')
        except Exception as e:
            messages.error(request, f"Error adding project: {e}")
            
    return redirect('administrator_projects_url')

@login_required
def administrator_projects_update_view(request, pk):
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        # Helper function to safely get decimal values
        def get_decimal(name):
            val = request.POST.get(name)
            if val == '' or val is None:
                return None
            return Decimal(val)
        
        # Helper for integers
        def get_int(name):
            val = request.POST.get(name)
            if val == '' or val is None:
                return None
            return int(val)
            
        # Helper for dates
        def get_date(name):
            val = request.POST.get(name)
            return val if val else None

        # Capture old data for audit (including financials)
        # Also capture old_funds early for budget sync calculation
        old_funds_for_sync = project.funds or Decimal('0.00')
        old_budget_for_sync = project.budget
        
        old_data = {
            "project_title": project.project_title,
            "status": project.status,
            "fund_source": project.fund_source,
            "funds": str(project.funds) if project.funds else "0.00",
            "total_project_cost": str(project.total_project_cost) if project.total_project_cost else "0.00",
            "counterpart_funds": str(project.counterpart_funds) if project.counterpart_funds else "0.00",
            "total_funds_released": str(project.total_funds_released) if project.total_funds_released else "0.00",
            "budget": project.budget.fund_source if project.budget else None,
        }

        # Update Fields
        project.no = get_int('no')
        project.project_code = request.POST.get('project_code')
        project.year = get_int('year')
        
        project.project_title = request.POST.get('project_title')
        # Update legacy description field too
        project.project_description = request.POST.get('project_title')

        project.agency_grantee = request.POST.get('agency_grantee')
        project.program = request.POST.get('program')
        project.type_of_project = request.POST.get('type_of_project')
        project.status = request.POST.get('status')
        project.remarks = request.POST.get('remarks')
        
        project.mun = request.POST.get('mun')
        project.province = request.POST.get('province')
        project.district = request.POST.get('district')
        
        project.beneficiary = request.POST.get('beneficiary')
        project.beneficiary_address = request.POST.get('beneficiary_address')
        project.contact_details = request.POST.get('contact_details')
        project.proponent_details = request.POST.get('proponent_details')
        project.no_of_beneficiaries = get_int('no_of_beneficiaries')
        project.male = get_int('male')
        project.female = get_int('female')
        project.total_beneficiaries = get_int('total_beneficiaries')
        project.senior_citizen = get_int('senior_citizen')
        project.pwd = get_int('pwd')
        
        # Handle fund_source - may contain "fund_source|fiscal_year" format
        fund_source_raw = request.POST.get('fund_source', '')
        budget_changed = False
        new_budget = None
        
        if '|' in fund_source_raw:
            fund_source_parts = fund_source_raw.split('|')
            project.fund_source = fund_source_parts[0]
            fiscal_year = fund_source_parts[1] if len(fund_source_parts) > 1 else None
            # Find the new budget
            new_budget = Budget.objects.filter(fund_source=fund_source_parts[0], fiscal_year=fiscal_year).first()
            if new_budget and new_budget != old_budget_for_sync:
                budget_changed = True
        else:
            project.fund_source = fund_source_raw
        
        # ============================================
        # SYNC PROJECT FUNDS WITH BUDGET ALLOCATION
        # ============================================
        new_funds = get_decimal('funds') or Decimal('0.00')
        
        # Case 1: Budget changed - refund old, deduct from new
        if budget_changed:
            # Refund the old budget completely
            if old_budget_for_sync and old_funds_for_sync > Decimal('0.00'):
                old_budget_for_sync.remaining_amount += old_funds_for_sync
                if old_budget_for_sync.status in ['exhausted', 'fully_allocated']:
                    old_budget_for_sync.status = 'partially_allocated'
                old_budget_for_sync.save(update_fields=['remaining_amount', 'status', 'date_updated'])
            
            # Deduct from the new budget
            if new_budget and new_funds > Decimal('0.00'):
                if new_budget.remaining_amount >= new_funds:
                    new_budget.remaining_amount -= new_funds
                    if new_budget.remaining_amount == Decimal('0.00'):
                        new_budget.status = 'fully_allocated'
                    elif new_budget.status in ['available', 'active']:
                        new_budget.status = 'partially_allocated'
                    new_budget.save(update_fields=['remaining_amount', 'status', 'date_updated'])
                else:
                    messages.error(request, f'Insufficient budget! Available: ₱{new_budget.remaining_amount:,.2f}, Requested: ₱{new_funds:,.2f}')
                    return redirect('administrator_projects_url')
            
            project.budget = new_budget
        
        # Case 2: Same budget, just funds changed
        elif project.budget and new_funds != old_funds_for_sync:
            funds_difference = new_funds - old_funds_for_sync
            budget = project.budget
            
            if funds_difference > Decimal('0.00'):
                # Increasing funds - deduct more from budget
                if budget.remaining_amount >= funds_difference:
                    budget.remaining_amount -= funds_difference
                    if budget.remaining_amount == Decimal('0.00'):
                        budget.status = 'fully_allocated'
                    budget.save(update_fields=['remaining_amount', 'status', 'date_updated'])
                else:
                    messages.error(request, f'Insufficient budget! Available: ₱{budget.remaining_amount:,.2f}, Requested increase: ₱{funds_difference:,.2f}')
                    return redirect('administrator_projects_url')
            else:
                # Decreasing funds - refund to budget
                refund_amount = abs(funds_difference)
                budget.remaining_amount += refund_amount
                if budget.status in ['exhausted', 'fully_allocated']:
                    budget.status = 'partially_allocated'
                budget.save(update_fields=['remaining_amount', 'status', 'date_updated'])
        
        project.funds = new_funds if new_funds != Decimal('0.00') else None
        # ============================================

        project.total_project_cost = get_decimal('total_project_cost')
        project.counterpart_funds = get_decimal('counterpart_funds')
        project.internally_managed_fund = get_decimal('internally_managed_fund')
        project.total_funds_released = get_decimal('total_funds_released')
        project.dost_viii = get_decimal('dost_viii')
        project.counterpart_fund = get_decimal('counterpart_fund')
        
        project.first_tranche = get_decimal('first_tranche')
        project.second_tranche = get_decimal('second_tranche')
        project.third_tranche = get_decimal('third_tranche')
        
        project.project_start = get_date('project_start')
        project.project_end = get_date('project_end')
        project.date_of_release = get_date('date_of_release')
        project.date_of_completion = get_date('date_of_completion')
        project.date_of_donation = get_date('date_of_donation')
        project.date_of_inspection_tagging = get_date('date_of_inspection_tagging')
        project.date_of_liquidation = get_date('date_of_liquidation')
        
        project.original_project_duration = request.POST.get('original_project_duration')
        project.extension_date = request.POST.get('extension_date')
        project.availed_technologies = request.POST.get('availed_technologies')
        project.interventions = request.POST.get('interventions')
        project.donation_status = request.POST.get('donation_status')
        
        project.check_ada_no = request.POST.get('check_ada_no')
        project.status_of_liquidation = request.POST.get('status_of_liquidation')
        project.amount_liquidated = get_decimal('amount_liquidated')
        
        project.tafr = request.POST.get('tafr')
        project.par = request.POST.get('par')
        project.list_of_eqpt = request.POST.get('list_of_eqpt')
        project.terminal_report = request.POST.get('terminal_report')
        project.invoice_receipt = request.POST.get('invoice_receipt')
        project.donated = request.POST.get('donated')
        project.acknowledgment_receipt_by_grantee = request.POST.get('acknowledgment_receipt_by_grantee')
        project.pme_visit = request.POST.get('pme_visit')
        project.womens_group = request.POST.get('womens_group')

        # Files
        if 'supporting_documents' in request.FILES:
            project.supporting_documents = request.FILES['supporting_documents']
        if 'product_photo' in request.FILES:
            project.product_photo = request.FILES['product_photo']

        project.save()

        # Audit Log - capture all changed financial data
        new_data = {
            "project_title": project.project_title,
            "status": project.status,
            "fund_source": project.fund_source,
            "funds": str(project.funds) if project.funds else "0.00",
            "total_project_cost": str(project.total_project_cost) if project.total_project_cost else "0.00",
            "counterpart_funds": str(project.counterpart_funds) if project.counterpart_funds else "0.00",
            "total_funds_released": str(project.total_funds_released) if project.total_funds_released else "0.00",
            "budget": project.budget.fund_source if project.budget else None,
        }

        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="Project",
            object_id=str(project.pk),
            old_data=old_data,
            new_data=new_data
        )

        messages.success(request, 'Project updated successfully.')
        return redirect('administrator_projects_detail_url', pk=project.pk)

    # GET request - render edit form
    context = {
        'project': project,
    }
    return render(request, 'administrator/project_edit.html', context)


# -------------------------
# Delete Project
# -------------------------
@login_required
def administrator_projects_delete_view(request, pk):
    project = get_object_or_404(Project, pk=pk)

    # Capture old data for audit
    old_data = {
        "project_title": project.project_title,
        "project_code": project.project_code,
        "funds": str(project.funds) if project.funds else None,
        "status": project.status,
        "mun": project.mun,
        "program": project.program,
    }

    try:
        project_title = project.project_title
        
        # ============================================
        # SYNC: Refund budget when deleting project
        # ============================================
        if project.budget and project.funds:
            budget = project.budget
            # remaining_amount is a property, so we need to reduce delivered_equipment_value instead
            budget.delivered_equipment_value = max(Decimal('0.00'), budget.delivered_equipment_value - project.funds)
            if budget.status in ['exhausted', 'fully_allocated']:
                budget.status = 'partially_allocated'
            budget.save(update_fields=['delivered_equipment_value', 'status', 'date_updated'])
        # ============================================
        
        project.delete()

        # Audit Log
        AuditLog.objects.create(
            user=request.user,
            action="delete",
            model_name="Project",
            object_id=str(pk),
            old_data=old_data,
            new_data=None,
        )

        messages.success(request, f"Project '{project_title}' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting project: {e}")

    return redirect('administrator_projects_url')


# -------------------------
# Mass Delete Projects
# -------------------------
@login_required
def administrator_projects_mass_delete_view(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            deleted_count = 0
            for pk in ids:
                try:
                    project = Project.objects.get(pk=pk)
                    old_data = {
                        "project_title": project.project_title,
                        "project_code": project.project_code,
                        "status": project.status,
                    }
                    
                    # ============================================
                    # SYNC: Refund budget when mass deleting
                    # ============================================
                    if project.budget and project.funds:
                        budget = project.budget
                        # remaining_amount is a property, so reduce delivered_equipment_value instead
                        budget.delivered_equipment_value = max(Decimal('0.00'), budget.delivered_equipment_value - project.funds)
                        if budget.status in ['exhausted', 'fully_allocated']:
                            budget.status = 'partially_allocated'
                        budget.save(update_fields=['delivered_equipment_value', 'status', 'date_updated'])
                    # ============================================
                    
                    project.delete()
                    
                    AuditLog.objects.create(
                        user=request.user,
                        action="delete",
                        model_name="Project",
                        object_id=str(pk),
                        old_data=old_data,
                        new_data=None,
                    )
                    deleted_count += 1
                except Project.DoesNotExist:
                    pass
                except Exception as e:
                    messages.error(request, f"Error deleting project ID {pk}: {e}")
            
            if deleted_count > 0:
                messages.success(request, f"Successfully deleted {deleted_count} project(s).")
        else:
            messages.warning(request, "No projects selected for deletion.")
    
    return redirect('administrator_projects_url')


# -------------------------
# DOST Equipment Delivery Recording
# -------------------------
@login_required
def administrator_equipment_delivery_add_view(request):
    """Record equipment delivery with DOST property tag"""
    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        equipment_name = request.POST.get('equipment_name')
        equipment_category_id = request.POST.get('equipment_category')
        quantity = request.POST.get('quantity', 1)
        property_tag_number = request.POST.get('property_tag_number')
        serial_numbers = request.POST.get('serial_numbers')
        delivery_date_str = request.POST.get('delivery_date')
        received_by = request.POST.get('received_by')
        condition_notes = request.POST.get('condition_notes')
        
        try:
            project = Project.objects.get(pk=project_id)
            
            # Parse dates
            from datetime import datetime
            delivery_date = datetime.strptime(delivery_date_str, '%Y-%m-%d').date() if delivery_date_str else timezone.now().date()
            
            # Get budget for this project
            budget = project.budget if hasattr(project, 'budget') else Budget.objects.filter(project=project).first()
            
            if not budget:
                # Create a budget if none exists
                budget = Budget.objects.create(
                    title=f"Budget for {project.project_title}",
                    fiscal_year=project.year or timezone.now().year,
                    fund_source="DOST SETUP",
                    total_amount=project.funds or Decimal('0.00'),
                )
                project.budget = budget
                project.save()
            
            # Get or create equipment item
            from .models import EquipmentCategory, EquipmentItem
            
            # Get category from form or use a default
            if equipment_category_id:
                try:
                    category = EquipmentCategory.objects.get(pk=equipment_category_id)
                except EquipmentCategory.DoesNotExist:
                    category, _ = EquipmentCategory.objects.get_or_create(
                        name='General Equipment',
                        defaults={'description': 'General project equipment'}
                    )
            else:
                category, _ = EquipmentCategory.objects.get_or_create(
                    name='General Equipment',
                    defaults={'description': 'General project equipment'}
                )
            
            # Get or create the equipment item
            equipment_item, _ = EquipmentItem.objects.get_or_create(
                name=equipment_name,
                category=category,
                defaults={'description': f'Equipment for {project.project_title}'}
            )
            
            # Get or create budget allocation - link via budget, not project
            budget_allocation, created = BudgetAllocation.objects.get_or_create(
                budget=budget,
                equipment_item=equipment_item,
                defaults={
                    'allocated_quantity': int(quantity),
                    'allocation_date': delivery_date,
                    'status': 'delivered',
                    'remarks': f'Equipment delivery: {equipment_name}',
                    'allocated_by': request.user,
                }
            )
            
            if not created:
                # Update existing allocation
                budget_allocation.allocated_quantity += int(quantity)
                budget_allocation.save()
            
            # Create the ProjectEquipment delivery record
            equipment_delivery = ProjectEquipment.objects.create(
                budget_allocation=budget_allocation,
                project=project,
                delivered_quantity=int(quantity),
                delivery_date=delivery_date,
                property_tag_number=property_tag_number,
                ownership_status='dost_owned',
                serial_numbers=serial_numbers,
                received_by=received_by,
                condition_notes=condition_notes,
                status='delivered',
                delivered_by=request.user,
            )
            
            # Audit Log
            AuditLog.objects.create(
                user=request.user,
                action="create",
                model_name="ProjectEquipment",
                object_id=str(equipment_delivery.pk),
                old_data=None,
                new_data={
                    "project": project.project_title,
                    "equipment_name": equipment_name,
                    "quantity": quantity,
                    "property_tag_number": property_tag_number,
                    "delivery_date": str(delivery_date),
                    "received_by": received_by,
                },
                reason="DOST equipment delivery recording"
            )
            
            messages.success(request, f'Equipment delivery recorded successfully. Property Tag: {property_tag_number}')
            
        except Project.DoesNotExist:
            messages.error(request, 'Project not found.')
        except Exception as e:
            messages.error(request, f'Error recording equipment delivery: {str(e)}')
    
    return redirect('administrator_projects_url')


# -------------------------
# Delete Proposal
# -------------------------
@login_required
def administrator_proposals_delete_view(request, pk):
    proposal = get_object_or_404(Proposal, pk=pk)

    # Capture old data for audit
    old_data = {
        "title": proposal.title,
        "description": proposal.description,
        "status": proposal.status,
        "proposed_amount": str(proposal.proposed_amount) if proposal.proposed_amount else None,
        "approved_amount": str(proposal.approved_amount) if proposal.approved_amount else None,
        "budget": proposal.budget.fund_source if proposal.budget else None,
    }

    try:
        proposal_title = proposal.title
        proposal.delete()

        # Audit Log
        AuditLog.objects.create(
            user=request.user,
            action="delete",
            model_name="Proposal",
            object_id=str(pk),
            old_data=old_data,
            new_data=None,
        )

        messages.success(request, f"Proposal '{proposal_title}' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting proposal: {e}")

    return redirect('administrator_proposals_url')


# -------------------------
# Mass Delete Proposals
# -------------------------
@login_required
def administrator_proposals_mass_delete_view(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            deleted_count = 0
            for pk in ids:
                try:
                    proposal = Proposal.objects.get(pk=pk)
                    old_data = {
                        "title": proposal.title,
                        "status": proposal.status,
                        "proposed_amount": str(proposal.proposed_amount) if proposal.proposed_amount else None,
                    }
                    proposal.delete()
                    
                    AuditLog.objects.create(
                        user=request.user,
                        action="delete",
                        model_name="Proposal",
                        object_id=str(pk),
                        old_data=old_data,
                        new_data=None,
                    )
                    deleted_count += 1
                except Proposal.DoesNotExist:
                    pass
                except Exception as e:
                    messages.error(request, f"Error deleting proposal ID {pk}: {e}")
            
            if deleted_count > 0:
                messages.success(request, f"Successfully deleted {deleted_count} proposal(s).")
        else:
            messages.warning(request, "No proposals selected for deletion.")
    
    return redirect('administrator_proposals_url')


# -------------------------
# Delete User
# -------------------------
@login_required
def administrator_users_delete_view(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Prevent self-deletion
    if user == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('administrator_users_url')

    # Capture old data for audit
    old_data = {
        "username": user.username,
        "email": user.email,
        "role": user.role,
        "first_name": user.first_name,
        "last_name": user.last_name,
    }

    try:
        user_email = user.email
        user.delete()

        # Audit Log
        AuditLog.objects.create(
            user=request.user,
            action="delete",
            model_name="User",
            object_id=str(user_id),
            old_data=old_data,
            new_data=None,
        )

        messages.success(request, f"User '{user_email}' deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting user: {e}")

    return redirect('administrator_users_url')


# -------------------------
# Mass Delete Users
# -------------------------
@login_required
def administrator_users_mass_delete_view(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            deleted_count = 0
            skipped_count = 0
            for user_id in ids:
                try:
                    user = User.objects.get(pk=user_id)
                    
                    # Skip self-deletion
                    if user == request.user:
                        skipped_count += 1
                        continue
                    
                    old_data = {
                        "username": user.username,
                        "email": user.email,
                        "role": user.role,
                    }
                    user.delete()
                    
                    AuditLog.objects.create(
                        user=request.user,
                        action="delete",
                        model_name="User",
                        object_id=str(user_id),
                        old_data=old_data,
                        new_data=None,
                    )
                    deleted_count += 1
                except User.DoesNotExist:
                    pass
                except Exception as e:
                    messages.error(request, f"Error deleting user ID {user_id}: {e}")
            
            if deleted_count > 0:
                messages.success(request, f"Successfully deleted {deleted_count} user(s).")
            if skipped_count > 0:
                messages.warning(request, f"Skipped {skipped_count} user(s) - cannot delete your own account.")
        else:
            messages.warning(request, "No users selected for deletion.")
    
    return redirect('administrator_users_url')


# -------------------------
# Create Task
# -------------------------
def administrator_task_create_view(request):
    if request.method == 'POST':
        description = request.POST.get('description')
        project_id = request.POST.get('project')
        assigned_to_id = request.POST.get('assigned_to')
        start_date = request.POST.get('start_date')
        due_date = request.POST.get('due_date')
        priority = request.POST.get('priority', 'medium')
        category = request.POST.get('category', 'other')
        progress_percentage = request.POST.get('progress_percentage', 0)
        estimated_hours = request.POST.get('estimated_hours')
        actual_hours = request.POST.get('actual_hours')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        location_name = request.POST.get('location_name')

        project = Project.objects.get(id=project_id)
        assigned_to = User.objects.get(id=assigned_to_id) if assigned_to_id else None

        # Use provided coordinates, or fall back to project coordinates
        task_latitude = latitude if latitude else (project.latitude if project.latitude else None)
        task_longitude = longitude if longitude else (project.longitude if project.longitude else None)
        task_location_name = location_name if location_name else (project.mun + ', ' + project.province if project.mun and project.province else project.beneficiary_address or '')

        task = Task.objects.create(
            title=f"Task for {project.project_title}",
            description=description,
            project=project,
            assigned_to=assigned_to,
            start_date=start_date if start_date else None,
            due_date=due_date,
            priority=priority,
            category=category,
            progress_percentage=int(progress_percentage) if progress_percentage else 0,
            estimated_hours=estimated_hours if estimated_hours else None,
            actual_hours=actual_hours if actual_hours else None,
            latitude=task_latitude,
            longitude=task_longitude,
            location_name=task_location_name,
        )

        # -------------------------
        # Audit Log
        # -------------------------
        from .models import AuditLog  # adjust path if needed

        new_data = {
            "title": task.title,
            "description": task.description,
            "project": task.project.project_title if task.project else None,
            "assigned_to": task.assigned_to.username if task.assigned_to else None,
            "start_date": str(task.start_date) if task.start_date else None,
            "due_date": str(task.due_date) if task.due_date else None,
            "priority": task.priority,
            "category": task.category,
            "status": task.status,
            "progress_percentage": task.progress_percentage,
            "estimated_hours": str(task.estimated_hours) if task.estimated_hours else None,
            "actual_hours": str(task.actual_hours) if task.actual_hours else None,
            "latitude": task.latitude,
            "longitude": task.longitude,
            "location_name": task.location_name,
        }

        AuditLog.objects.create(
            user=request.user,
            action="create",
            model_name="Task",
            object_id=str(task.pk),
            old_data=None,
            new_data=new_data
        )
        # -------------------------

        # -------------------------
        # Notify assigned staff
        # -------------------------
        if assigned_to:
            Notification.objects.create(
                sender=request.user,
                receiver=assigned_to,
                message=f"You have been assigned to task: '{task.title}' for project '{project.project_title}'.",
                category='task',
                link=reverse('staff_task_list_url')
            )

        messages.success(request, "Task created successfully!")
        return redirect('administrator_task_list_url')


# -------------------------
# Edit Task
# -------------------------
def administrator_task_edit_view(request):
    if request.method == 'POST':
        task_id = request.POST.get('id')
        task = get_object_or_404(Task, id=task_id)

        project_id = request.POST.get('project')
        selected_project = Project.objects.filter(id=project_id).first() if project_id else task.project

        # Capture old assigned_to for notification check
        old_assigned_to = task.assigned_to

        # Capture old data
        old_data = {
            "title": task.title,
            "description": task.description,
            "project": task.project.project_title if task.project else None,
            "assigned_to": task.assigned_to.username if task.assigned_to else None,
            "start_date": str(task.start_date) if task.start_date else None,
            "due_date": str(task.due_date) if task.due_date else None,
            "priority": task.priority,
            "category": task.category,
            "status": task.status,
            "progress_percentage": task.progress_percentage,
            "estimated_hours": str(task.estimated_hours) if task.estimated_hours else None,
            "actual_hours": str(task.actual_hours) if task.actual_hours else None,
            "latitude": task.latitude,
            "longitude": task.longitude,
            "location_name": task.location_name,
        }

        # Update fields
        task.title = f"Task for {selected_project.project_title}" if selected_project else task.title
        task.description = request.POST.get('description')
        task.project_id = project_id
        task.assigned_to_id = request.POST.get('assigned_to') or None
        task.start_date = request.POST.get('start_date') or None
        task.due_date = request.POST.get('due_date')
        task.priority = request.POST.get('priority', 'medium')
        task.category = request.POST.get('category', 'other')
        task.status = request.POST.get('status')
        task.progress_percentage = int(request.POST.get('progress_percentage', 0))
        task.estimated_hours = request.POST.get('estimated_hours') or None
        task.actual_hours = request.POST.get('actual_hours') or None
        task.latitude = request.POST.get('latitude')
        task.longitude = request.POST.get('longitude')
        task.location_name = request.POST.get('location_name')
        task.save()

        # -------------------------
        # Notify new assigned staff (if assignment changed)
        # -------------------------
        new_assigned_to = task.assigned_to
        if new_assigned_to and new_assigned_to != old_assigned_to:
            Notification.objects.create(
                sender=request.user,
                receiver=new_assigned_to,
                message=f"You have been assigned to task: '{task.title}' for project '{task.project.project_title if task.project else 'N/A'}'.",
                link=reverse('staff_task_list_url')
            )

        # -------------------------
        # Audit Log
        # -------------------------
        new_data = {
            "title": task.title,
            "description": task.description,
            "project": task.project.project_title if task.project else None,
            "assigned_to": task.assigned_to.username if task.assigned_to else None,
            "start_date": str(task.start_date) if task.start_date else None,
            "due_date": str(task.due_date) if task.due_date else None,
            "priority": task.priority,
            "category": task.category,
            "status": task.status,
            "progress_percentage": task.progress_percentage,
            "estimated_hours": str(task.estimated_hours) if task.estimated_hours else None,
            "actual_hours": str(task.actual_hours) if task.actual_hours else None,
            "latitude": task.latitude,
            "longitude": task.longitude,
            "location_name": task.location_name,
        }

        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="Task",
            object_id=str(task.pk),
            old_data=old_data,
            new_data=new_data
        )
        # -------------------------

        messages.success(request, "Task updated successfully!")
        return redirect('administrator_task_list_url')


@login_required
def administrator_task_delete_view(request, task_id):
    if request.method == 'POST':
        task = get_object_or_404(Task, id=task_id)

        # Capture data for audit log before deletion
        old_data = {
            "title": task.title,
            "description": task.description,
            "project": task.project.project_title if task.project else None,
            "assigned_to": task.assigned_to.username if task.assigned_to else None,
            "due_date": str(task.due_date) if task.due_date else None,
            "status": task.status,
            "latitude": task.latitude,
            "longitude": task.longitude,
            "location_name": task.location_name,
        }

        # Notify assigned staff about task deletion
        if task.assigned_to:
            Notification.objects.create(
                sender=request.user,
                receiver=task.assigned_to,
                message=f"Task '{task.title}' for project '{task.project.project_title if task.project else 'N/A'}' has been deleted.",
                link=reverse('staff_task_list_url')
            )

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="delete",
            model_name="Task",
            object_id=str(task.pk),
            old_data=old_data,
            new_data={}
        )

        # Delete the task
        task.delete()

        messages.success(request, "Task deleted successfully!")
        return redirect('administrator_task_list_url')
    else:
        return redirect('administrator_task_list_url')


def financial_summary_pdf(request):
    # Extract filter parameters
    report_year = request.GET.get('year', str(datetime.now().year))
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    municipality = request.GET.get('municipality')

    # Base queryset for budgets
    budget_queryset = Budget.objects.all()

    # Apply year filter
    if report_year and report_year != 'all':
        try:
            year = int(report_year)
            budget_queryset = budget_queryset.filter(fiscal_year=year)
        except ValueError:
            pass

    # Apply date range filter
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            budget_queryset = budget_queryset.filter(date_allocated__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            budget_queryset = budget_queryset.filter(date_allocated__lte=end)
        except ValueError:
            pass

    # Apply status filter
    if status and status != 'all':
        budget_queryset = budget_queryset.filter(status=status)

    # Note: Municipality filter not applied to budgets since they are provincial-level allocations

    # Fetch filtered budgets
    budgets = budget_queryset.all()

    # Prepare table data
    data = [['Fund Source', 'Fiscal Year', 'Budget (PHP)', 'Spent (PHP)', 'Balance (PHP)', 'Status', 'Date Allocated']]
    
    labels = []
    total_amounts = []
    spent_amounts = []
    remaining_amounts = []

    for budget in budgets:
        spent = budget.transactions.filter(transaction_type='deduction').aggregate(total_spent=Sum('amount'))['total_spent'] or 0
        remaining = budget.total_amount - spent

        data.append([
            budget.fund_source,
            str(budget.fiscal_year),
            f"{budget.total_amount:,.0f}",
            f"{spent:,.0f}",
            f"{remaining:,.0f}",
            budget.status.capitalize(),
            budget.date_allocated.strftime("%b %d, %Y"),
        ])

        # For graph
        labels.append(budget.fund_source)
        total_amounts.append(float(budget.total_amount))
        spent_amounts.append(float(spent))
        remaining_amounts.append(float(remaining))

    # Create HTTP response with PDF content type
    # Generate descriptive title based on filters
    title_parts = ["DOST Financial Summary Report"]
    
    if report_year and report_year != 'all':
        title_parts.append(f"Fiscal Year {report_year}")
    
    if municipality and municipality != 'all':
        title_parts.append(f"Municipality of {municipality}")
    
    if status and status != 'all':
        title_parts.append(f"Status: {status.title()}")
    
    if start_date or end_date:
        date_range = []
        if start_date:
            date_range.append(f"From {start_date}")
        if end_date:
            date_range.append(f"To {end_date}")
        if date_range:
            title_parts.append(" • ".join(date_range))
    
    report_title = " | ".join(title_parts)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="dost_financial_summary_{report_year or datetime.now().year}.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # DOST Header Function
    def add_dost_header(canvas, doc):
        """Add DOST header to every page"""
        canvas.saveState()
        
        # Try to load DOST logo
        try:
            from django.conf import settings
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'dost.png')
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                canvas.drawImage(logo, 30, doc.pagesize[1] - 60, width=50, height=35, mask='auto')
        except:
            pass  # Skip logo if not found
        
        # Center the header text
        page_width = doc.pagesize[0]
        
        # Main title - centered
        canvas.setFont('Helvetica-Bold', 12)
        title_text = 'Department of Science and Technology'
        title_width = canvas.stringWidth(title_text, 'Helvetica-Bold', 12)
        title_x = (page_width - title_width) / 2
        canvas.drawString(title_x, doc.pagesize[1] - 40, title_text)
        
        # Subtitle - centered
        canvas.setFont('Helvetica', 9)
        subtitle_text = 'Provincial Science and Technology Center - Biliran'
        subtitle_width = canvas.stringWidth(subtitle_text, 'Helvetica', 9)
        subtitle_x = (page_width - subtitle_width) / 2
        canvas.drawString(subtitle_x, doc.pagesize[1] - 53, subtitle_text)
        
        # Report generation date - right aligned
        canvas.setFont('Helvetica', 7)
        current_time = datetime.now().strftime('%B %d, %Y %I:%M %p')
        date_text = f'Generated: {current_time}'
        date_width = canvas.stringWidth(date_text, 'Helvetica', 7)
        date_x = page_width - date_width - 30  # 30px margin from right
        canvas.drawString(date_x, doc.pagesize[1] - 40, date_text)
        
        canvas.restoreState()

    # Create custom page template with header
    frame = Frame(30, 30, doc.pagesize[0]-60, doc.pagesize[1]-80, id='normal')
    template = PageTemplate(id='dost_template', frames=frame, onPage=add_dost_header)
    doc.addPageTemplates([template])

    # Add title with filter information
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=10,
        alignment=1  # Center
    )
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 12))

    # ---------------------------
    # Generate Graph with Matplotlib
    # ---------------------------
    fig, ax = plt.subplots(figsize=(8.6, 4.4))
    fig.patch.set_facecolor('#F8FAFC')
    ax.set_facecolor('#FFFFFF')
    bar_width = 0.24
    index = range(len(labels))

    bars_total = ax.bar(index, total_amounts, bar_width, label='Total Budget', color='#2563EB', edgecolor='white', linewidth=0.8)
    bars_spent = ax.bar([i + bar_width for i in index], spent_amounts, bar_width, label='Spent', color='#EF4444', edgecolor='white', linewidth=0.8)
    bars_remaining = ax.bar([i + bar_width * 2 for i in index], remaining_amounts, bar_width, label='Remaining', color='#10B981', edgecolor='white', linewidth=0.8)

    ax.set_xlabel('Fund Source', fontsize=10, fontweight='bold')
    ax.set_ylabel('Amount (PHP)', fontsize=10, fontweight='bold')
    ax.set_title('Financial Overview per Budget', fontsize=13, fontweight='bold', pad=12)
    ax.set_xticks([i + bar_width for i in index])
    ax.set_xticklabels(labels, rotation=35, ha='right', fontsize=9)
    ax.tick_params(axis='y', labelsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.25)
    for spine in ['top', 'right', 'left']:
        ax.spines[spine].set_visible(False)
    ax.legend(frameon=False, ncol=3, loc='upper left', fontsize=9)

    # Format y-axis as currency without decimals
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'PHP {int(x):,}'))

    def _annotate_currency_bars(bars):
        ymax = max(total_amounts + spent_amounts + remaining_amounts) if (total_amounts + spent_amounts + remaining_amounts) else 0
        pad = ymax * 0.012 if ymax else 0
        for bar in bars:
            h = bar.get_height()
            if h <= 0:
                continue
            ax.text(bar.get_x() + bar.get_width() / 2, h + pad, f'{int(h):,}', ha='center', va='bottom', fontsize=7, color='#334155')

    _annotate_currency_bars(bars_total)
    _annotate_currency_bars(bars_spent)
    _annotate_currency_bars(bars_remaining)

    # Save plot to BytesIO buffer
    buf = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format='PNG', dpi=220, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)

    # Add graph to PDF
    img = Image(buf, width=450, height=250)
    elements.append(img)
    elements.append(Spacer(1, 12))

    # ---------------------------
    # Create Table
    # ---------------------------
    table = Table(data, repeatRows=1, colWidths=[120, 60, 80, 80, 80, 70, 80])
    table.setStyle(TableStyle([
        # Header row background
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        # Grid only, no background for rows
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

def proposal_status_pdf(request):
    # Extract filter parameters
    report_year = request.GET.get('year', str(datetime.now().year))
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    municipality = request.GET.get('municipality')

    # Base queryset for proposals
    proposal_queryset = Proposal.objects.all()

    # Apply year filter
    if report_year and report_year != 'all':
        try:
            year = int(report_year)
            proposal_queryset = proposal_queryset.filter(submission_date__year=year)
        except ValueError:
            pass

    # Apply date range filter
    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            proposal_queryset = proposal_queryset.filter(submission_date__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            proposal_queryset = proposal_queryset.filter(submission_date__lte=end)
        except ValueError:
            pass

    # Apply status filter
    if status and status != 'all':
        proposal_queryset = proposal_queryset.filter(status=status)

    # Apply municipality filter (proposals have municipality field)
    if municipality and municipality != 'all':
        proposal_queryset = proposal_queryset.filter(municipality__iexact=municipality)

    # Fetch filtered proposals
    proposals = proposal_queryset.all().order_by('-submission_date')

    # ---------------------------
    # Prepare Table Data
    # ---------------------------
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    wrap_style = ParagraphStyle(
        'wrap', 
        parent=normal_style, 
        fontSize=8, 
        leading=10
    )

    table_data = [['Title', 'Submitted By', 'Proponent', 'Beneficiary', 'Submission Date', 'Status', 
                   'Proposed (PHP)', 'Approved (PHP)', 'Budget Source']]
    
    # Count statuses for chart
    status_counts = {
        'Pending': 0,
        'Approved': 0,
        'Declined': 0,  # merge Rejected + Needs Revision
    }

    for proposal in proposals:
        status_display = proposal.get_status_display()
        # Merge 'Rejected' and 'Needs Revision' into 'Declined'
        if status_display in ['Rejected', 'Needs Revision']:
            chart_status = 'Declined'
        elif status_display == 'For Review':
            continue  # skip 'For Review'
        else:
            chart_status = status_display

        # Increment chart counts
        status_counts[chart_status] += 1

        submitted_by_name = proposal.submitted_by.get_full_name() if proposal.submitted_by else 'N/A'
        proponent_name = proposal.processed_by.get_full_name() if proposal.processed_by else 'N/A'
        beneficiary_name = proposal.beneficiary.get_full_name() if proposal.beneficiary else 'N/A'
        budget_source = proposal.budget.fund_source if proposal.budget else 'N/A'

        # Table row uses merged status
        table_status = chart_status

        table_data.append([
            Paragraph(proposal.title, wrap_style),
            Paragraph(submitted_by_name, wrap_style),
            Paragraph(proponent_name, wrap_style),
            Paragraph(beneficiary_name, wrap_style),
            proposal.submission_date.strftime("%b %d, %Y"),
            table_status,
            f"{proposal.proposed_amount:,.0f}",
            f"{proposal.approved_amount:,.0f}" if proposal.approved_amount is not None else '-',
            Paragraph(budget_source, wrap_style),
        ])

    # ---------------------------
    # Generate Bar Chart
    # ---------------------------
    labels = list(status_counts.keys())
    counts = list(status_counts.values())
    colors_list = ['gold', 'lightgreen', 'red']  # Pending, Approved, Declined

    fig, ax = plt.subplots(figsize=(6.2, 5.4))
    fig.patch.set_facecolor('#F8FAFC')
    ax.set_facecolor('#FFFFFF')
    bars = ax.bar(labels, counts, color=['#F59E0B', '#10B981', '#EF4444'], edgecolor='white', linewidth=0.9)
    ax.set_ylabel('Number of Proposals', fontsize=10, fontweight='bold')
    ax.set_title('Proposal Status Overview', fontsize=13, fontweight='bold', pad=12)
    ax.tick_params(axis='both', labelsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.25)
    for spine in ['top', 'right', 'left']:
        ax.spines[spine].set_visible(False)

    # Make y-axis integer only
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))

    # Add value labels on top of bars
    for bar in bars:
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width()/2,
            height + 0.1,
            f'{int(height)}',
            ha='center',
            va='bottom',
            fontsize=9,
            color='#334155'
        )

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='PNG', dpi=220, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)

    # ---------------------------
    # Create PDF in Landscape with DOST Header
    # ---------------------------
    # Generate descriptive title based on filters
    title_parts = ["DOST Proposal Status Report"]
    
    if report_year and report_year != 'all':
        title_parts.append(f"Fiscal Year {report_year}")
    
    if municipality and municipality != 'all':
        title_parts.append(f"Municipality of {municipality}")
    
    if status and status != 'all':
        title_parts.append(f"Status: {status.title()}")
    
    if start_date or end_date:
        date_range = []
        if start_date:
            date_range.append(f"From {start_date}")
        if end_date:
            date_range.append(f"To {end_date}")
        if date_range:
            title_parts.append(" • ".join(date_range))
    
    report_title = " | ".join(title_parts)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="dost_proposal_status_{report_year or datetime.now().year}.pdf"'

    # Create custom page template with header
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter)
    )
    elements = []

    # DOST Header Function
    def add_dost_header(canvas, doc):
        """Add DOST header to every page"""
        canvas.saveState()
        
        # Try to load DOST logo
        try:
            from django.conf import settings
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'dost.png')
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                canvas.drawImage(logo, 30, doc.pagesize[1] - 60, width=50, height=35, mask='auto')
        except:
            pass  # Skip logo if not found
        
        # Center the header text
        page_width = doc.pagesize[0]
        
        # Main title - centered
        canvas.setFont('Helvetica-Bold', 12)
        title_text = 'Department of Science and Technology'
        title_width = canvas.stringWidth(title_text, 'Helvetica-Bold', 12)
        title_x = (page_width - title_width) / 2
        canvas.drawString(title_x, doc.pagesize[1] - 40, title_text)
        
        # Subtitle - centered
        canvas.setFont('Helvetica', 9)
        subtitle_text = 'Provincial Science and Technology Center - Biliran'
        subtitle_width = canvas.stringWidth(subtitle_text, 'Helvetica', 9)
        subtitle_x = (page_width - subtitle_width) / 2
        canvas.drawString(subtitle_x, doc.pagesize[1] - 53, subtitle_text)
        
        # Report generation date - right aligned
        canvas.setFont('Helvetica', 7)
        current_time = datetime.now().strftime('%B %d, %Y %I:%M %p')
        date_text = f'Generated: {current_time}'
        date_width = canvas.stringWidth(date_text, 'Helvetica', 7)
        date_x = page_width - date_width - 30  # 30px margin from right
        canvas.drawString(date_x, doc.pagesize[1] - 40, date_text)
        
        canvas.restoreState()

    # Create custom page template with header
    frame = Frame(30, 30, doc.pagesize[0]-60, doc.pagesize[1]-80, id='normal')
    template = PageTemplate(id='dost_template', frames=frame, onPage=add_dost_header)
    doc.addPageTemplates([template])

    # Add title with filter information
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=10,
        alignment=1  # Center
    )
    elements.append(Paragraph(report_title, title_style))
    elements.append(Spacer(1, 20))

    # Add bar chart
    elements.append(Image(buf, width=400, height=300))
    elements.append(Spacer(1, 20))

    # ---------------------------
    # Auto-adjust table
    # ---------------------------
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response

def approved_projects_pdf(request, report_year=None):
    import io
    from django.http import HttpResponse
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import landscape, letter, A4, portrait
    from reportlab.lib import colors
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator

    # ---------------------------
    # Filter projects and order by project_code
    # ---------------------------
    projects = Project.objects.all().order_by('project_code')
    if report_year:
        projects = projects.filter(approval_date__year=report_year)

    # ---------------------------
    # Bar Chart (Approved Budgets Only)
    # ---------------------------
    labels = [p.project_title for p in projects]
    approved_amounts = [float(p.approved_budget) for p in projects]

    num_projects = max(len(labels), 1)
    fig_width = 16  # full page width
    fig_height = 6
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    fig.patch.set_facecolor('#F8FAFC')
    ax.set_facecolor('#FFFFFF')
    bars = ax.bar(range(num_projects), approved_amounts, color='#2563EB', edgecolor='white', linewidth=0.8)
    ax.set_title(f'Approved Project Budgets – Fiscal Year {report_year or "All"}', fontsize=13, fontweight='bold', pad=12)
    ax.set_xlabel('Project', fontsize=10, fontweight='bold')
    ax.set_ylabel('Approved Budget (PHP)', fontsize=10, fontweight='bold')
    ax.set_xticks(range(num_projects))
    ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
    ax.tick_params(axis='y', labelsize=9)
    ax.grid(axis='y', linestyle='--', alpha=0.25)
    for spine in ['top', 'right', 'left']:
        ax.spines[spine].set_visible(False)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x):,}'))
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))

    peak = max(approved_amounts) if approved_amounts else 0
    offset = peak * 0.01 if peak else 0
    for bar in bars:
        h = bar.get_height()
        if h <= 0:
            continue
        ax.text(bar.get_x() + bar.get_width() / 2, h + offset, f'{int(h):,}', ha='center', va='bottom', fontsize=7, color='#334155')

    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='PNG', dpi=150)
    plt.close(fig)
    buf.seek(0)

    # ---------------------------
    # Table Data
    # ---------------------------
    styles = getSampleStyleSheet()
    wrap_style = ParagraphStyle('wrap', parent=styles['Normal'], fontSize=8, leading=10)

    table_data = [
        ['Project Code', 'Project Title', 'PSTO Staff', 'Proponent', 'Beneficiary', 'Budget (PHP)', 'Status', 'Start Date', 'End Date', 'Tasks']
    ]

    for project in projects:
        project_leader = project.project_leader.full_name() if project.project_leader else 'N/A'
        processed_by = project.proposal.processed_by.full_name() if project.proposal and project.proposal.processed_by else 'N/A'
        beneficiary = project.proposal.beneficiary.full_name() if project.proposal and project.proposal.beneficiary else 'N/A'

        # Numbered tasks with extra line breaks
        tasks_list = project.tasks.all()
        tasks_str = ""
        for i, t in enumerate(tasks_list, start=1):
            due = t.due_date.strftime('%b %d, %Y') if t.due_date else '-'
            tasks_str += f"{i}. {t.title} ({t.get_status_display()}) - Due: {due}<br/><br/>"

        table_data.append([
            project.project_code or '-',
            Paragraph(project.project_title, wrap_style),
            Paragraph(project_leader, wrap_style),
            Paragraph(processed_by, wrap_style),
            Paragraph(beneficiary, wrap_style),
            f"{project.approved_budget:,.0f}",
            project.get_status_display(),
            project.start_date.strftime('%b %d, %Y') if project.start_date else '-',
            project.end_date.strftime('%b %d, %Y') if project.end_date else '-',
            Paragraph(tasks_str.strip().replace('<br/>', '<br />'), wrap_style)
        ])

    # ---------------------------
    # Build PDF
    # ---------------------------
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="approved_projects_{report_year or "all"}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15
    )

    elements = [
        # Page 1: Bar Chart
        Paragraph(f"Approved Project Budgets – Fiscal Year {report_year or 'All'}", styles['Title']),
        Spacer(1, 20),
        Image(buf, width=doc.pagesize[0]-30, height=400),  # full width
        PageBreak(),  # start table on a new page

        # Page 2: Table
        Paragraph(f"Approved Project Budgets Table – Fiscal Year {report_year or 'All'}", styles['Title']),
        Spacer(1, 20),
        Table(table_data, repeatRows=1, style=TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
    ]

    doc.build(elements)
    return response


@login_required
def export_full_report_pdf(request):
    """
    Comprehensive PDF export containing all charts, summary data, and complete project/proposal lists.
    Supports selective section inclusion via query parameters.
    """
    import io
    from django.http import HttpResponse
    from reportlab.platypus.doctemplate import BaseDocTemplate
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, PageTemplate, Frame
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import landscape, letter, A4, portrait
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.utils import ImageReader
    import os
    import matplotlib.pyplot as plt
    from matplotlib.ticker import MaxNLocator
    from decimal import Decimal
    from django.db.models import Sum, Count, Q
    from datetime import datetime
    from django.utils import timezone

    # Get filter parameters from request
    selected_year = request.GET.get('year')
    selected_municipality = request.GET.get('municipality')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_status = request.GET.get('status')
    
    # Get sorting parameters
    sort_projects = request.GET.get('sort_projects', 'municipality')
    sort_proposals = request.GET.get('sort_proposals', 'municipality')
    
    # Get section inclusion parameters (default to True if not specified)
    include_summary = request.GET.get('include_summary', '1') == '1'
    include_equipment = request.GET.get('include_equipment', '1') == '1'
    include_projects = request.GET.get('include_projects', '1') == '1'
    include_proposals = request.GET.get('include_proposals', '1') == '1'
    include_charts = request.GET.get('include_charts', '1') == '1'
    include_signatory = request.GET.get('include_signatory', '1') == '1'
    include_esignature = request.GET.get('include_esignature', '0') == '1'  # Default to False
    page_size = (request.GET.get('page_size') or 'letter').lower()
    page_orientation = (request.GET.get('page_orientation') or 'landscape').lower()
    
    # Only set current year as default if explicitly requested, otherwise don't filter by year
    current_year = datetime.now().year
    # Don't set default year - only filter if user explicitly selects a year

    # Base querysets for filtering
    project_queryset = Project.objects.all()
    proposal_queryset = Proposal.objects.all()
    budget_queryset = Budget.objects.all()
    task_queryset = Task.objects.all()

    # Apply filters
    if selected_year:
        try:
            year_int = int(selected_year)
            project_queryset = project_queryset.filter(year=year_int)
            proposal_queryset = proposal_queryset.filter(submission_date__year=year_int)
            budget_queryset = budget_queryset.filter(fiscal_year=year_int)
        except ValueError:
            pass

    if selected_municipality and selected_municipality != '':
        project_queryset = project_queryset.filter(mun__iexact=selected_municipality)
        proposal_queryset = proposal_queryset.filter(municipality__iexact=selected_municipality)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__gte=start) | 
                Q(date_of_completion__gte=start) |
                Q(approval_date__date__gte=start)
            )
            proposal_queryset = proposal_queryset.filter(submission_date__date__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__lte=end) | 
                Q(date_of_completion__lte=end) |
                Q(approval_date__date__lte=end)
            )
            proposal_queryset = proposal_queryset.filter(submission_date__date__lte=end)
        except ValueError:
            pass

    if selected_status and selected_status != '':
        project_queryset = project_queryset.filter(status__iexact=selected_status)

    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=20
    )
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        spaceBefore=15
    )
    # Page layout
    pagesize_map = {
        'letter': letter,
        'a4': A4,
    }
    base_pagesize = pagesize_map.get(page_size, letter)
    pagesize = landscape(base_pagesize) if page_orientation == 'landscape' else portrait(base_pagesize)

    page_width, page_height = pagesize
    margin = 24
    header_height = 60
    frame = Frame(
        margin,
        margin,
        page_width - (margin * 2),
        page_height - header_height - (margin * 2),
        id='normal'
    )
    content_width = frame._width
    content_height = frame._height

    table_font_size = 6 if page_orientation == 'landscape' else 7
    header_font_size = table_font_size

    # Style for wrapping text in table cells
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=table_font_size,
        leading=table_font_size + 2,
    )

    def scale_widths(widths, target=content_width):
        total = sum(widths) or 1
        scale = target / total
        return [max(12, w * scale) for w in widths]

    chart_cell_width = content_width / 2
    chart_image_width = chart_cell_width - 8
    chart_small_height = min(240, content_height * 0.34)
    chart_wide_height = min(260, content_height * 0.38)

    def fig_size_for(width_pt, height_pt):
        return (
            max(3.2, width_pt / 72),
            max(2.4, height_pt / 72),
        )

    small_figsize = fig_size_for(chart_image_width, chart_small_height)
    wide_figsize = fig_size_for(content_width, chart_wide_height)

    # -----------------------------
    # Gather All Data (using filtered querysets)
    # -----------------------------
    
    # Financial Summary (calculated from budget allocations)
    total_budget = budget_queryset.aggregate(total=Sum('total_equipment_value'))['total'] or Decimal('0.00')
    total_spent = budget_queryset.aggregate(total=Sum('delivered_equipment_value'))['total'] or Decimal('0.00')
    total_remaining = total_budget - total_spent
    utilization_rate = round((float(total_spent) / float(total_budget) * 100), 2) if total_budget > 0 else 0

    # Proposal Status
    proposal_counts_qs = proposal_queryset.values('status').annotate(total=Count('id'))
    proposal_status_counts = {'pending': 0, 'approved': 0, 'rejected': 0, 'for_review': 0, 'needs_revision': 0}
    for p in proposal_counts_qs:
        if p['status'] in proposal_status_counts:
            proposal_status_counts[p['status']] = p['total']
    total_proposals = sum(proposal_status_counts.values())

    # Project Status
    project_counts_qs = project_queryset.values('status').annotate(total=Count('id'))
    project_status_counts = {'new': 0, 'ongoing': 0, 'completed': 0, 'terminated': 0}
    for p in project_counts_qs:
        status = p['status'] or 'new'
        # Normalize status to match the buckets
        if status.lower() in ['new', 'proposal', 'proposed']:
            project_status_counts['new'] += p['total']
        elif status.lower() in ['ongoing', 'on-going', 'in_progress', 'in progress', 'inprogress']:
            project_status_counts['ongoing'] += p['total']
        elif status.lower() in ['completed', 'complete']:
            project_status_counts['completed'] += p['total']
        elif status.lower() in ['terminated', 'cancelled', 'canceled']:
            project_status_counts['terminated'] += p['total']
        else:
            project_status_counts['new'] += p['total']  # Default to new
    total_projects = sum(project_status_counts.values())
    completion_rate = round((project_status_counts['completed'] / total_projects * 100), 1) if total_projects > 0 else 0

    # User Role Distribution (users are not filtered)
    user_counts_qs = User.objects.values('role').annotate(total=Count('id'))
    user_role_counts = {'admin': 0, 'dost_staff': 0, 'proponent': 0, 'beneficiary': 0}
    for u in user_counts_qs:
        if u['role'] in user_role_counts:
            user_role_counts[u['role']] = u['total']
    total_users = sum(user_role_counts.values())

    # Projects by Municipality (field is 'mun' in model)
    municipality_qs = project_queryset.values('mun').annotate(count=Count('id')).order_by('mun')
    municipality_counts = {m['mun']: m['count'] for m in municipality_qs if m['mun']}
    top_municipality_name = '-'
    top_municipality_count = 0
    if municipality_counts:
        top_municipality_name, top_municipality_count = max(municipality_counts.items(), key=lambda item: item[1])

    # Proposals by Municipality
    proposal_municipality_qs = proposal_queryset.values('municipality').annotate(count=Count('id')).order_by('-count')
    proposal_municipality_counts = {m['municipality'] or 'Unspecified': m['count'] for m in proposal_municipality_qs}
    top_proposal_municipality_name = '-'
    top_proposal_municipality_count = 0
    if proposal_municipality_counts:
        top_proposal_municipality_name, top_proposal_municipality_count = max(
            proposal_municipality_counts.items(), key=lambda item: item[1]
        )

    # Projects by Program and Type
    program_qs = project_queryset.values('program').annotate(count=Count('id')).order_by('-count')
    program_counts = {p['program'] or 'Unspecified': p['count'] for p in program_qs}
    type_qs = project_queryset.values('type_of_project').annotate(count=Count('id')).order_by('-count')
    type_counts = {t['type_of_project'] or 'Unspecified': t['count'] for t in type_qs}

    top_program_name = '-'
    top_program_count = 0
    if program_counts:
        top_program_name, top_program_count = max(program_counts.items(), key=lambda item: item[1])

    top_type_name = '-'
    top_type_count = 0
    if type_counts:
        top_type_name, top_type_count = max(type_counts.items(), key=lambda item: item[1])

    # Top Proponents by Proposal Count
    top_proponent_labels = []
    top_proponent_values = []
    top_proponent_name = '-'
    top_proponent_count = 0
    top_prop_qs = proposal_queryset.exclude(submitted_by__isnull=True).values(
        'submitted_by__first_name', 'submitted_by__last_name', 'submitted_by__username'
    ).annotate(total=Count('id')).order_by('-total')[:10]
    for row in top_prop_qs:
        full_name = f"{row.get('submitted_by__first_name') or ''} {row.get('submitted_by__last_name') or ''}".strip()
        if not full_name:
            full_name = row.get('submitted_by__username') or 'Unknown'
        top_proponent_labels.append(full_name)
        top_proponent_values.append(row['total'])
    if top_proponent_labels:
        top_proponent_name = top_proponent_labels[0]
        top_proponent_count = top_proponent_values[0]

    # Task Summary (filter tasks by projects)
    project_ids = list(project_queryset.values_list('id', flat=True))
    total_tasks = task_queryset.filter(project__in=project_ids).count() if project_ids else 0
    tasks_pending = task_queryset.filter(project__in=project_ids, status='pending').count() if project_ids else 0
    tasks_in_progress = task_queryset.filter(project__in=project_ids, status='in_progress').count() if project_ids else 0
    tasks_completed = task_queryset.filter(project__in=project_ids, status='completed').count() if project_ids else 0
    tasks_delayed = task_queryset.filter(project__in=project_ids, status='delayed').count() if project_ids else 0

    # Allocation status counts (equipment-based assistance)
    allocations_pending = BudgetAllocation.objects.filter(status='allocated').count()
    allocations_delivered = BudgetAllocation.objects.filter(status='delivered').count()

    # ALL Projects for complete list (filtered and sorted)
    # Define sorting mappings for projects
    project_sort_mapping = {
        'title': 'project_title',
        'title_desc': '-project_title',
        'municipality': 'mun',
        'municipality_desc': '-mun',
        'status': 'status',
        'funds': '-funds',
        'funds_asc': 'funds',
        'date': '-project_start',
        'date_asc': 'project_start',
        'proponent': 'proposal__proponent__last_name',
    }
    project_order = project_sort_mapping.get(sort_projects, 'mun')
    all_projects = project_queryset.all().order_by(project_order, 'project_title')
    
    # ALL Proposals for complete list (filtered and sorted)
    # Define sorting mappings for proposals
    proposal_sort_mapping = {
        'title': 'title',
        'title_desc': '-title',
        'municipality': 'municipality',
        'municipality_desc': '-municipality',
        'status': 'status',
        'amount': '-proposed_amount',
        'amount_asc': 'proposed_amount',
        'date': '-submission_date',
        'date_asc': 'submission_date',
        'proponent': 'proponent__last_name',
    }
    proposal_order = proposal_sort_mapping.get(sort_proposals, 'municipality')
    all_proposals = proposal_queryset.all().order_by(proposal_order, 'title')

    # -----------------------------
    # Project & Proponent Insights (for PDF)
    # -----------------------------

    # -----------------------------
    # Create Charts (Compact sizes with high DPI for quality)
    # -----------------------------
    chart_images = []

    def style_axes(ax, title, ylabel=None, xlabel=None, rotate_x=False):
        ax.set_facecolor('#FFFFFF')
        if ylabel:
            ax.set_ylabel(ylabel, fontsize=10, fontweight='bold')
        if xlabel:
            ax.set_xlabel(xlabel, fontsize=10, fontweight='bold')
        ax.set_title(title, fontsize=12, fontweight='bold', pad=10)
        ax.tick_params(axis='both', labelsize=9)
        ax.grid(axis='y', linestyle='--', alpha=0.25)
        if rotate_x:
            for tick in ax.get_xticklabels():
                tick.set_rotation(35)
                tick.set_ha('right')
        for spine in ['top', 'right', 'left']:
            ax.spines[spine].set_visible(False)

    def annotate_bars(ax, bars, horizontal=False):
        if horizontal:
            peak = max((bar.get_width() for bar in bars), default=0)
            pad = peak * 0.015 if peak else 0
            for bar in bars:
                width = bar.get_width()
                if width <= 0:
                    continue
                ax.text(width + pad, bar.get_y() + (bar.get_height() / 2), f'{int(width)}', va='center', fontsize=8, color='#334155')
            return

        peak = max((bar.get_height() for bar in bars), default=0)
        pad = peak * 0.015 if peak else 0
        for bar in bars:
            height = bar.get_height()
            if height <= 0:
                continue
            ax.text(bar.get_x() + (bar.get_width() / 2), height + pad, f'{int(height)}', ha='center', va='bottom', fontsize=8, color='#334155')
    
    # 1. Projects by Program Chart
    fig1, ax1 = plt.subplots(figsize=small_figsize)
    fig1.patch.set_facecolor('#F8FAFC')
    program_labels = list(program_counts.keys())
    program_values = list(program_counts.values())
    bars1 = ax1.bar(program_labels, program_values, color='#0EA5A4', edgecolor='white', linewidth=0.8)
    style_axes(ax1, 'Projects by Program', ylabel='Projects', rotate_x=True)
    ax1.yaxis.set_major_locator(MaxNLocator(integer=True))
    annotate_bars(ax1, bars1)
    plt.tight_layout()
    buf1 = io.BytesIO()
    plt.savefig(buf1, format='PNG', dpi=200, bbox_inches='tight')
    plt.close(fig1)
    buf1.seek(0)
    chart_images.append(('Projects by Program', buf1))

    # 2. Proposal Status Chart
    fig2, ax2 = plt.subplots(figsize=small_figsize)
    fig2.patch.set_facecolor('#F8FAFC')
    labels2 = ['Pending', 'Approved', 'Declined']
    values2 = [proposal_status_counts['pending'], proposal_status_counts['approved'], proposal_status_counts['rejected']]
    colors2 = ['#F59E0B', '#10B981', '#EF4444']
    bars2 = ax2.bar(labels2, values2, color=colors2, edgecolor='white', linewidth=0.8)
    style_axes(ax2, 'Proposal Status', ylabel='Count')
    ax2.yaxis.set_major_locator(MaxNLocator(integer=True))
    annotate_bars(ax2, bars2)
    plt.tight_layout()
    buf2 = io.BytesIO()
    plt.savefig(buf2, format='PNG', dpi=200, bbox_inches='tight')
    plt.close(fig2)
    buf2.seek(0)
    chart_images.append(('Proposal Status', buf2))

    # 3. Project Status Chart (Bar chart for clarity)
    fig3, ax3 = plt.subplots(figsize=small_figsize)
    fig3.patch.set_facecolor('#F8FAFC')
    labels3 = ['New', 'Ongoing', 'Completed', 'Terminated']
    values3 = [project_status_counts['new'], project_status_counts['ongoing'], 
               project_status_counts['completed'], project_status_counts['terminated']]
    colors3 = ['#9CA3AF', '#3B82F6', '#22C55E', '#EF4444']
    bars3 = ax3.bar(labels3, values3, color=colors3, edgecolor='white', linewidth=0.8)
    style_axes(ax3, 'Project Status Distribution', ylabel='Count')
    ax3.yaxis.set_major_locator(MaxNLocator(integer=True))
    annotate_bars(ax3, bars3)
    plt.tight_layout()
    buf3 = io.BytesIO()
    plt.savefig(buf3, format='PNG', dpi=200, bbox_inches='tight')
    plt.close(fig3)
    buf3.seek(0)
    chart_images.append(('Project Status', buf3))

    # 4. User Role Distribution Chart (use legend instead of labels on pie)
    fig4, ax4 = plt.subplots(figsize=small_figsize)
    fig4.patch.set_facecolor('#F8FAFC')
    labels4 = ['Admin', 'Staff', 'Proponent', 'Beneficiary']
    values4 = [user_role_counts['admin'], user_role_counts['dost_staff'], 
               user_role_counts['proponent'], user_role_counts['beneficiary']]
    colors4 = ['#8B5CF6', '#EC4899', '#0EA5E9', '#F59E0B']
    wedges4, texts4, autotexts4 = ax4.pie(
        values4,
        colors=colors4,
        autopct='%1.1f%%',
        startangle=90,
        textprops={'fontsize': 9, 'color': '#334155'},
        pctdistance=0.75,
        wedgeprops={'linewidth': 1, 'edgecolor': 'white'}
    )
    for t in autotexts4:
        t.set_color('white')
        t.set_fontsize(8)
    ax4.legend(wedges4, labels4, loc='center left', bbox_to_anchor=(1, 0.5), fontsize=9)
    ax4.set_title('User Roles', fontsize=12, fontweight='bold')
    plt.tight_layout()
    buf4 = io.BytesIO()
    plt.savefig(buf4, format='PNG', dpi=200, bbox_inches='tight')
    plt.close(fig4)
    buf4.seek(0)
    chart_images.append(('User Roles', buf4))

    # 5. Projects by Municipality Chart
    if municipality_counts:
        fig5, ax5 = plt.subplots(figsize=wide_figsize)
        fig5.patch.set_facecolor('#F8FAFC')
        muni_labels = list(municipality_counts.keys())
        muni_values = list(municipality_counts.values())
        colors5 = ['#3B82F6', '#22C55E', '#F59E0B', '#EF4444', '#8B5CF6', '#EC4899', '#0EA5E9', '#9CA3AF']
        bars5 = ax5.bar(muni_labels, muni_values, color=colors5[:len(muni_labels)], edgecolor='white', linewidth=0.8)
        style_axes(ax5, 'Projects by Municipality', ylabel='Projects', xlabel='Municipality', rotate_x=True)
        ax5.yaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_bars(ax5, bars5)
        plt.tight_layout()
        buf5 = io.BytesIO()
        plt.savefig(buf5, format='PNG', dpi=200, bbox_inches='tight')
        plt.close(fig5)
        buf5.seek(0)
        chart_images.append(('Municipality', buf5))

    # 6. Proposals by Municipality Chart
    if proposal_municipality_counts:
        fig6, ax6 = plt.subplots(figsize=wide_figsize)
        fig6.patch.set_facecolor('#F8FAFC')
        prop_muni_labels = list(proposal_municipality_counts.keys())
        prop_muni_values = list(proposal_municipality_counts.values())
        bars6 = ax6.bar(prop_muni_labels, prop_muni_values, color='#0EA5E9', edgecolor='white', linewidth=0.8)
        style_axes(ax6, 'Proposals by Municipality', ylabel='Proposals', xlabel='Municipality', rotate_x=True)
        ax6.yaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_bars(ax6, bars6)
        plt.tight_layout()
        buf6 = io.BytesIO()
        plt.savefig(buf6, format='PNG', dpi=200, bbox_inches='tight')
        plt.close(fig6)
        buf6.seek(0)
        chart_images.append(('Proposals by Municipality', buf6))

    # 7. Top Proponents Chart
    if top_proponent_labels:
        fig7, ax7 = plt.subplots(figsize=wide_figsize)
        fig7.patch.set_facecolor('#F8FAFC')
        bars7 = ax7.barh(top_proponent_labels, top_proponent_values, color='#2563EB', edgecolor='white', linewidth=0.8)
        style_axes(ax7, 'Top Proponents by Proposal Count', xlabel='Proposals')
        ax7.xaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_bars(ax7, bars7, horizontal=True)
        plt.tight_layout()
        buf7 = io.BytesIO()
        plt.savefig(buf7, format='PNG', dpi=200, bbox_inches='tight')
        plt.close(fig7)
        buf7.seek(0)
        chart_images.append(('Top Proponents', buf7))

    # 8. Projects by Type Chart
    if type_counts:
        fig8, ax8 = plt.subplots(figsize=wide_figsize)
        fig8.patch.set_facecolor('#F8FAFC')
        type_labels = list(type_counts.keys())
        type_values = list(type_counts.values())
        bars8 = ax8.bar(type_labels, type_values, color='#F59E0B', edgecolor='white', linewidth=0.8)
        style_axes(ax8, 'Projects by Type', ylabel='Projects', rotate_x=True)
        ax8.yaxis.set_major_locator(MaxNLocator(integer=True))
        annotate_bars(ax8, bars8)
        plt.tight_layout()
        buf8 = io.BytesIO()
        plt.savefig(buf8, format='PNG', dpi=200, bbox_inches='tight')
        plt.close(fig8)
        buf8.seek(0)
        chart_images.append(('Projects by Type', buf8))

    # -----------------------------
    # DOST Header Function for Every Page
    # -----------------------------
    def add_dost_header(canvas, doc):
        """Add DOST header to every page"""
        canvas.saveState()

        # Metadata used by built-in PDF viewers.
        canvas.setTitle(pdf_title)
        canvas.setAuthor('DOST Biliran')
        canvas.setSubject('Comprehensive Report')
        
        page_width = doc.pagesize[0]
        page_height = doc.pagesize[1]
        
        # Header background
        canvas.setFillColor(colors.HexColor('#F8FAFC'))
        canvas.rect(0, page_height - header_height, page_width, header_height, fill=1)
        canvas.setFillColor(colors.black)
        
        # Try to load DOST logo
        logo_width = 60
        logo_height = 45
        logo_x = 40
        logo_y = page_height - header_height + (header_height - logo_height) / 2
        
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'dost.png')
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                canvas.drawImage(logo, logo_x, logo_y, width=logo_width, height=logo_height, mask='auto')
        except Exception as e:
            # Draw a placeholder if logo fails to load
            canvas.setFillColor(colors.HexColor('#2563EB'))
            canvas.rect(logo_x, logo_y, logo_width, logo_height, fill=1)
            canvas.setFillColor(colors.white)
            canvas.setFont('Helvetica-Bold', 8)
            canvas.drawCentredString(logo_x + logo_width/2, logo_y + logo_height/2 - 3, 'DOST')
            canvas.setFillColor(colors.black)
        
        # Center the header text (accounting for logo on left)
        text_start_x = logo_x + logo_width + 20
        
        # Main title
        canvas.setFont('Helvetica-Bold', 14)
        title_text = 'Department of Science and Technology'
        canvas.drawString(text_start_x, page_height - (header_height * 0.55), title_text)
        
        # Subtitle
        canvas.setFont('Helvetica', 10)
        subtitle_text = 'Provincial Science and Technology Center - Biliran'
        canvas.drawString(text_start_x, page_height - (header_height * 0.78), subtitle_text)
        
        # Report generation date - right aligned
        canvas.setFont('Helvetica', 8)
        current_time = datetime.now().strftime('%B %d, %Y %I:%M %p')
        date_text = f'Generated: {current_time}'
        date_width = canvas.stringWidth(date_text, 'Helvetica', 8)
        date_x = page_width - date_width - 30
        canvas.drawString(date_x, page_height - (header_height * 0.55), date_text)
        
        # Thin line separator
        canvas.setStrokeColor(colors.HexColor('#E5E7EB'))
        canvas.setLineWidth(0.5)
        canvas.line(30, page_height - header_height, page_width - 30, page_height - header_height)
        
        canvas.restoreState()

    # -----------------------------
    # Build PDF with Custom Template
    # -----------------------------
    
    # Generate descriptive title based on filters
    title_parts = ["DOST Biliran Comprehensive Report"]
    
    if selected_year and selected_year != 'all':
        title_parts.append(f"Fiscal Year {selected_year}")
    
    if selected_municipality and selected_municipality != 'all':
        title_parts.append(f"Municipality of {selected_municipality}")
    
    if selected_status and selected_status != 'all':
        title_parts.append(f"Status: {selected_status.title()}")
    
    if start_date or end_date:
        date_range = []
        if start_date:
            date_range.append(f"From {start_date}")
        if end_date:
            date_range.append(f"To {end_date}")
        if date_range:
            title_parts.append(" • ".join(date_range))
    
    report_title = " | ".join(title_parts)
    
    raw_filename = (request.GET.get('filename') or '').strip()
    fallback_filename = f"dost_report_{selected_year or 'all_years'}"
    if raw_filename:
        sanitized = ''.join(ch for ch in raw_filename if ch.isalnum() or ch in (' ', '-', '_', '.')).strip(' .')
        base_filename = sanitized or fallback_filename
    else:
        base_filename = fallback_filename
    if not base_filename.lower().endswith('.pdf'):
        base_filename = f"{base_filename}.pdf"
    pdf_title = os.path.splitext(base_filename)[0]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{base_filename}"'

    # Create custom page template with header
    template = PageTemplate(id='dost_template', frames=frame, onPage=add_dost_header)
    
    doc = BaseDocTemplate(response, pagesize=pagesize)
    doc.addPageTemplates([template])
    # Force the template to be used for all pages
    doc._pageTemplate = template
    elements = []

    # ========================================
    # PAGE 1: Report Title & Summary
    # ========================================
    
    # Report Title with filter information
    elements.append(Paragraph(report_title, title_style))
    elements.append(Paragraph(f"Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", 
                              ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=9, alignment=1, textColor=colors.gray)))
    elements.append(Spacer(1, 15))

    # Summary & Financial Tables Side by Side (only if summary is included)
    if include_summary:
        summary_data = [
            ['Metric', 'Value'],
            ['Total Projects', str(total_projects)],
            ['Ongoing Projects', str(project_status_counts['ongoing'])],
            ['Completed Projects', str(project_status_counts['completed'])],
            ['Total Proposals', str(total_proposals)],
            ['Total Users', str(total_users)],
            ['Total Tasks', str(total_tasks)],
            ['Completion Rate', f'{completion_rate}%'],
        ]
        summary_table = Table(summary_data, colWidths=scale_widths([140, 90], target=content_width / 2))
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        highlights_data = [
            ['Highlight', 'Value'],
            ['Top Area (Projects)', f"{top_municipality_name} ({top_municipality_count})"],
            ['Top Area (Proposals)', f"{top_proposal_municipality_name} ({top_proposal_municipality_count})"],
            ['Top Proponent', f"{top_proponent_name} ({top_proponent_count})"],
            ['Top Program', f"{top_program_name} ({top_program_count})"],
            ['Top Project Type', f"{top_type_name} ({top_type_count})"],
        ]
        highlights_table = Table(highlights_data, colWidths=scale_widths([150, 170], target=content_width / 2))
        highlights_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECFEFF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#A5F3FC')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        combined_tables = Table([[summary_table, highlights_table]], colWidths=[content_width / 2, content_width / 2])
        combined_tables.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(combined_tables)
        elements.append(Spacer(1, 20))

    # ========================================
    # COMPLETE PROJECTS LIST (ALL DATA)
    # ========================================
    if include_projects:
        elements.append(Paragraph("Complete Projects List", title_style))
        elements.append(Paragraph(f"Total Projects: {all_projects.count()}", styles['Normal']))
        elements.append(Spacer(1, 15))

        def fmt_date(value):
            return value.strftime('%m/%d/%Y') if value else 'N/A'

        def short_text(value, limit=40):
            if not value:
                return 'N/A'
            value = str(value)
            return value if len(value) <= limit else value[:max(0, limit - 3)] + '...'

        # Section 1: Identification & Basic Info
        section1_header = ['#', 'Code', 'Year', 'Project Title', 'Agency', 'Program', 'Type', 'Status', 'Remarks']
        section1_data = [section1_header]
        for idx, proj in enumerate(all_projects, 1):
            section1_data.append([
                str(idx),
                proj.project_code or 'N/A',
                str(proj.year or 'N/A'),
                Paragraph(short_text(proj.project_title, 60), cell_style),
                Paragraph(short_text(proj.agency_grantee, 40), cell_style),
                Paragraph(short_text(proj.program, 30), cell_style),
                Paragraph(short_text(proj.type_of_project, 25), cell_style),
                (proj.status or 'N/A').title(),
                Paragraph(short_text(proj.remarks, 25), cell_style),
            ])

        section1_table = Table(
            section1_data,
            colWidths=scale_widths([22, 45, 35, 130, 90, 70, 60, 55, 60]),
            repeatRows=1
        )
        section1_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EFF6FF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#DBEAFE')) for i in range(2, len(section1_data), 2)],
        ]))
        elements.append(section1_table)
        elements.append(Spacer(1, 16))

        # Section 2: Location & Beneficiaries
        section2_header = ['#', 'Project', 'Municipality', 'Province', 'District', 'Beneficiary', 'Beneficiary Address', 'Contact', 'Proponent', 'No. Ben', 'Male', 'Female', 'Total', 'Senior', 'PWD']
        section2_data = [section2_header]
        for idx, proj in enumerate(all_projects, 1):
            section2_data.append([
                str(idx),
                Paragraph(short_text(proj.project_title, 40), cell_style),
                proj.mun or 'N/A',
                proj.province or 'N/A',
                proj.district or 'N/A',
                Paragraph(short_text(proj.beneficiary, 30), cell_style),
                Paragraph(short_text(proj.beneficiary_address, 35), cell_style),
                Paragraph(short_text(proj.contact_details, 20), cell_style),
                Paragraph(short_text(proj.proponent_details, 25), cell_style),
                str(proj.no_of_beneficiaries or 0),
                str(proj.male or 0),
                str(proj.female or 0),
                str(proj.total_beneficiaries or 0),
                str(proj.senior_citizen or 0),
                str(proj.pwd or 0),
            ])

        section2_table = Table(
            section2_data,
            colWidths=scale_widths([22, 90, 55, 50, 45, 70, 85, 55, 70, 35, 30, 35, 35, 35, 30]),
            repeatRows=1
        )
        section2_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (9, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECFEFF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#99F6E4')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#CCFBF1')) for i in range(2, len(section2_data), 2)],
        ]))
        elements.append(section2_table)
        elements.append(Spacer(1, 16))

        # Section 3: Financials & Tranches
        section3_header = ['#', 'Project', 'Fund Source', 'Approved Budget', 'Total Cost', 'Counterpart', 'Internal Fund', 'Released', '1st', '2nd', '3rd']
        section3_data = [section3_header]
        for idx, proj in enumerate(all_projects, 1):
            section3_data.append([
                str(idx),
                Paragraph(short_text(proj.project_title, 40), cell_style),
                Paragraph(short_text(proj.fund_source, 25), cell_style),
                f'PHP {int(proj.funds or 0):,}',
                f'PHP {int(proj.total_project_cost or 0):,}',
                f'PHP {int(proj.counterpart_funds or 0):,}',
                f'PHP {int(proj.internally_managed_fund or 0):,}',
                f'PHP {int(proj.total_funds_released or 0):,}',
                f'PHP {int(proj.first_tranche or 0):,}',
                f'PHP {int(proj.second_tranche or 0):,}',
                f'PHP {int(proj.third_tranche or 0):,}',
            ])

        section3_table = Table(
            section3_data,
            colWidths=scale_widths([22, 100, 70, 80, 80, 70, 70, 70, 55, 55, 55]),
            repeatRows=1
        )
        section3_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EFF6FF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#DBEAFE')) for i in range(2, len(section3_data), 2)],
        ]))
        elements.append(section3_table)
        elements.append(Spacer(1, 16))

        # Section 4: Dates & Liquidation
        section4_header = ['#', 'Project', 'Start', 'End', 'Release', 'Completion', 'Duration', 'Extension', 'Check/ADA', 'Liq Status', 'Liq Date', 'Amount Liq']
        section4_data = [section4_header]
        for idx, proj in enumerate(all_projects, 1):
            section4_data.append([
                str(idx),
                Paragraph(short_text(proj.project_title, 40), cell_style),
                fmt_date(proj.project_start),
                fmt_date(proj.project_end),
                fmt_date(proj.date_of_release),
                fmt_date(proj.date_of_completion),
                short_text(proj.original_project_duration, 15),
                short_text(proj.extension_date, 15),
                short_text(proj.check_ada_no, 15),
                short_text(proj.status_of_liquidation, 15),
                fmt_date(proj.date_of_liquidation),
                f'PHP {int(proj.amount_liquidated or 0):,}',
            ])

        section4_table = Table(
            section4_data,
            colWidths=scale_widths([22, 100, 50, 50, 55, 60, 45, 45, 60, 55, 55, 70]),
            repeatRows=1
        )
        section4_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (6, -1), 'CENTER'),
            ('ALIGN', (11, 0), (11, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F3FF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#C4B5FD')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#EDE9FE')) for i in range(2, len(section4_data), 2)],
        ]))
        elements.append(section4_table)
        elements.append(Spacer(1, 16))

        # Section 5A: Tech & Documents
        section5a_header = ['#', 'Project', 'Tech Availed', 'Interventions', 'TAFR', 'PAR', 'Terminal Report', 'Invoice/Receipt', 'Equipment List']
        section5a_data = [section5a_header]
        for idx, proj in enumerate(all_projects, 1):
            section5a_data.append([
                str(idx),
                Paragraph(short_text(proj.project_title, 40), cell_style),
                Paragraph(short_text(proj.availed_technologies, 25), cell_style),
                Paragraph(short_text(proj.interventions, 25), cell_style),
                short_text(proj.tafr, 10),
                short_text(proj.par, 10),
                short_text(proj.terminal_report, 15),
                short_text(proj.invoice_receipt, 15),
                Paragraph(short_text(proj.list_of_eqpt, 25), cell_style),
            ])

        section5a_table = Table(
            section5a_data,
            colWidths=scale_widths([22, 90, 70, 70, 35, 35, 55, 60, 70]),
            repeatRows=1
        )
        section5a_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (7, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECFEFF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#99F6E4')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#CCFBF1')) for i in range(2, len(section5a_data), 2)],
        ]))
        elements.append(section5a_table)
        elements.append(Spacer(1, 16))

        # Section 5B: Donation & Other
        section5b_header = ['#', 'Project', 'Donated', 'Donation Date', 'Donation Status', 'PME Visit', 'Women Group', 'Inspection Date', 'Receipt by Grantee', 'Photo']
        section5b_data = [section5b_header]
        for idx, proj in enumerate(all_projects, 1):
            section5b_data.append([
                str(idx),
                Paragraph(short_text(proj.project_title, 40), cell_style),
                short_text(proj.donated, 10),
                fmt_date(proj.date_of_donation),
                short_text(proj.donation_status, 15),
                short_text(proj.pme_visit, 10),
                short_text(proj.womens_group, 10),
                fmt_date(proj.date_of_inspection_tagging),
                short_text(proj.acknowledgment_receipt_by_grantee, 15),
                'Yes' if proj.product_photo else 'No',
            ])

        section5b_table = Table(
            section5b_data,
            colWidths=scale_widths([22, 90, 40, 55, 60, 40, 40, 55, 60, 30]),
            repeatRows=1
        )
        section5b_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECFEFF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#99F6E4')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#CCFBF1')) for i in range(2, len(section5b_data), 2)],
        ]))
        elements.append(section5b_table)
        elements.append(Spacer(1, 20))

    # ========================================
    # COMPLETE PROPOSALS LIST (ALL DATA)
    # ========================================
    if include_proposals:
        elements.append(Paragraph("Complete Proposals List", title_style))
        elements.append(Paragraph(f"Total Proposals: {all_proposals.count()}", styles['Normal']))
        elements.append(Spacer(1, 15))
        
        # Proposals table header
        proposals_header = ['#', 'Title', 'Submitted By', 'Status', 'Proposed (PHP)', 'Approved (PHP)', 'Location', 'Submission Date']
        proposals_data = [proposals_header]
        
        for idx, prop in enumerate(all_proposals, 1):
            # Truncate long titles
            title = prop.title or 'Untitled'
            if len(title) > 60:
                title = title[:57] + '...'
            
            submitted_by = prop.submitted_by.full_name() if prop.submitted_by else 'N/A'
            if len(submitted_by) > 25:
                submitted_by = submitted_by[:22] + '...'
            
            location = prop.location or prop.municipality or 'N/A'
            if len(location) > 20:
                location = location[:17] + '...'
            
            # Format status nicely
            status_display = prop.get_status_display() if hasattr(prop, 'get_status_display') else (prop.status or 'N/A').replace('_', ' ').title()
            
            proposals_data.append([
                str(idx),
                Paragraph(title, cell_style),
                Paragraph(submitted_by, cell_style),
                status_display,
                f'PHP {int(prop.proposed_amount):,}' if prop.proposed_amount else 'PHP 0',
                f'PHP {int(prop.approved_amount):,}' if prop.approved_amount else '-',
                location,
                prop.submission_date.strftime('%m/%d/%Y') if prop.submission_date else 'N/A',
            ])
        
        # Column widths for landscape
        proposals_table = Table(
            proposals_data,
            colWidths=scale_widths([25, 180, 100, 70, 90, 90, 80, 75]),
            repeatRows=1
        )
        proposals_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#065F46')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (5, -1), 'RIGHT'),
            ('ALIGN', (7, 0), (7, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECFDF5')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#A7F3D0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            # Alternate row colors
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#D1FAE5')) for i in range(2, len(proposals_data), 2)],
        ]))
        elements.append(proposals_table)
        elements.append(Spacer(1, 20))

    # ========================================
    # PROJECT & PROPONENT INSIGHTS
    # ========================================
    if include_equipment:
        elements.append(Paragraph("Project & Proponent Insights", title_style))
        elements.append(Paragraph("Highlights on top proponents, programs, project types, and municipalities.", styles['Normal']))
        elements.append(Spacer(1, 15))

        # Top proponents table
        proponents_header = ['#', 'Proponent', 'Proposals']
        proponents_data = [proponents_header]
        for idx, (name, count) in enumerate(zip(top_proponent_labels, top_proponent_values), 1):
            proponents_data.append([str(idx), Paragraph(name, cell_style), str(count)])

        proponents_table = Table(proponents_data, colWidths=scale_widths([25, 200, 80]), repeatRows=1)
        proponents_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EFF6FF')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#DBEAFE')) for i in range(2, len(proponents_data), 2)],
        ]))

        # Projects by program table
        program_header = ['Program', 'Projects']
        program_data = [program_header]
        for name, count in list(program_counts.items())[:10]:
            program_data.append([Paragraph(name, cell_style), str(count)])
        program_table = Table(program_data, colWidths=scale_widths([220, 85], target=content_width / 2), repeatRows=1)
        program_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16A34A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECFDF5')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#A7F3D0')),
        ]))

        # Projects by type table
        type_header = ['Project Type', 'Projects']
        type_data = [type_header]
        for name, count in list(type_counts.items())[:10]:
            type_data.append([Paragraph(name, cell_style), str(count)])
        type_table = Table(type_data, colWidths=scale_widths([220, 85], target=content_width / 2), repeatRows=1)
        type_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), table_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FFFBEB')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#FDE68A')),
        ]))

        elements.append(proponents_table)
        elements.append(Spacer(1, 12))
        elements.append(Table([[program_table, type_table]], colWidths=[content_width / 2, content_width / 2]))

    # CHARTS - fixed pagination (4 charts per page)
    if include_charts and chart_images:
        chart_caption_style = ParagraphStyle(
            'ChartCaptionStyle',
            parent=styles['Normal'],
            fontSize=8,
            leading=9,
            alignment=1,
            textColor=colors.HexColor('#334155')
        )

        charts_per_page = 4
        charts_per_row = 2
        rows_per_page = 2
        gutter = 10
        total_chart_pages = (len(chart_images) + charts_per_page - 1) // charts_per_page

        for page_index in range(total_chart_pages):
            if page_index == 0:
                elements.append(PageBreak())
            else:
                elements.append(PageBreak())

            title_suffix = f" (Page {page_index + 1} of {total_chart_pages})" if total_chart_pages > 1 else ""
            elements.append(Paragraph(f"Charts & Graphs Overview{title_suffix}", section_style))
            elements.append(Spacer(1, 6))

            page_charts = chart_images[page_index * charts_per_page:(page_index + 1) * charts_per_page]

            card_width = (content_width - gutter) / charts_per_row
            available_grid_height = max(240, content_height - 46)
            card_height = (available_grid_height - gutter) / rows_per_page

            # Keep a stable aspect ratio for non-distorted chart cards.
            target_ratio = 1.62  # width / height
            image_max_width = card_width - 14
            image_max_height = max(80, card_height - 22)
            image_height = min(image_max_height, image_max_width / target_ratio)
            image_width = min(image_max_width, image_height * target_ratio)

            grid_data = []
            idx = 0
            for _ in range(rows_per_page):
                row_cards = []
                for _ in range(charts_per_row):
                    if idx < len(page_charts):
                        chart_title, chart_buffer = page_charts[idx]
                        chart_card = Table(
                            [[Paragraph(chart_title, chart_caption_style)], [Image(chart_buffer, width=image_width, height=image_height)]],
                            colWidths=[card_width - 8],
                            rowHeights=[10, image_height + 4]
                        )
                        chart_card.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFFFF')),
                            ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#CBD5E1')),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('TOPPADDING', (0, 0), (-1, 0), 2),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 1),
                            ('LEFTPADDING', (0, 0), (-1, -1), 2),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ]))
                        row_cards.append(chart_card)
                        idx += 1
                    else:
                        row_cards.append(Spacer(1, 1))
                grid_data.append(row_cards)

            charts_grid = Table(
                grid_data,
                colWidths=[card_width, card_width],
                rowHeights=[card_height, card_height]
            )
            charts_grid.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(charts_grid)

    # ========================================
    # SIGNATORY SECTION (Lower Right)
    # ========================================
    if include_signatory:
        elements.append(Spacer(1, 20))

        # Get current user's full name
        user_full_name = request.user.get_full_name() if request.user.get_full_name() else request.user.username

        # Check if user has a digital signature and if esignature is requested
        user_signature = None
        if include_esignature:
            user_signature = DigitalSignature.objects.filter(user=request.user).order_by('-signed_at').first()
        
        # Build signature section with or without image
        signature_elements = []
        
        if user_signature and include_esignature:
            # Try to get signature image
            signature_img = None
            try:
                if user_signature.signature_image and hasattr(user_signature.signature_image, 'path'):
                    # Use uploaded image file
                    if os.path.exists(user_signature.signature_image.path):
                        signature_img = Image(user_signature.signature_image.path, width=150, height=50)
                elif user_signature.signature_data:
                    # Use base64 signature data (drawn signature)
                    import base64
                    sig_data = user_signature.signature_data
                    if sig_data.startswith('data:image'):
                        sig_data = sig_data.split(',')[1]
                    sig_bytes = base64.b64decode(sig_data)
                    sig_io = io.BytesIO(sig_bytes)
                    signature_img = Image(sig_io, width=150, height=50)
            except Exception as e:
                print(f"Error loading signature: {e}")
                signature_img = None
            
            if signature_img:
                signature_data = [
                    ['Prepared by:', '', ''],
                    ['', signature_img, ''],
                    ['', user_full_name, ''],
                    ['', datetime.now().strftime('%B %d, %Y'), ''],
                ]
            else:
                signature_data = [
                    ['Prepared by:', '', ''],
                    ['', user_full_name, ''],
                    ['', datetime.now().strftime('%B %d, %Y'), ''],
                ]
        else:
            # No signature available
            signature_data = [
                ['Prepared by:', '', ''],
                ['', user_full_name, ''],
                ['', datetime.now().strftime('%B %d, %Y'), ''],
            ]

        signature_table = Table(signature_data, colWidths=scale_widths([100, 200, 100], target=content_width / 2))
        signature_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Center signature and name
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),  # Very reduced padding
            ('TOPPADDING', (0, 0), (-1, -1), 1),     # Very reduced padding
            # Add a line above the name row
            ('LINEABOVE', (1, -2), (1, -2), 1, colors.black),
        ]))

        # Create a table that spans full width with signatory on the right
        signatory_content = Table([['', signature_table]], colWidths=[content_width / 2, content_width / 2])
        signatory_content.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),  # Align the signature table to the right
            ('VALIGN', (1, 0), (1, 0), 'BOTTOM'),
        ]))
        elements.append(signatory_content)

    # Build and return PDF
    doc.build(elements)
    return response


@login_required
def administrator_reports_view(request):
    from decimal import Decimal
    from django.db.models import Sum, Count, Q
    import json
    print("=== DEBUG: administrator_reports_view called ===")  # DEBUG

    # Get filter parameters
    selected_year = request.GET.get('year')
    selected_municipality = request.GET.get('municipality')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_status = request.GET.get('status')
    
    # Don't set default year - only filter if user explicitly selects a year
    current_year = datetime.now().year

    # Base querysets for filtering
    project_queryset = Project.objects.all()
    proposal_queryset = Proposal.objects.all()
    budget_queryset = Budget.objects.all()
    task_queryset = Task.objects.all()

    # Apply filters
    if selected_year:
        try:
            year_int = int(selected_year)
            project_queryset = project_queryset.filter(year=year_int)
            proposal_queryset = proposal_queryset.filter(submission_date__year=year_int)
            budget_queryset = budget_queryset.filter(fiscal_year=year_int)
        except ValueError:
            pass

    if selected_municipality and selected_municipality != '':
        project_queryset = project_queryset.filter(mun__iexact=selected_municipality)
        proposal_queryset = proposal_queryset.filter(municipality__iexact=selected_municipality)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__gte=start) | 
                Q(date_of_completion__gte=start) |
                Q(approval_date__date__gte=start)
            )
            proposal_queryset = proposal_queryset.filter(submission_date__date__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__lte=end) | 
                Q(date_of_completion__lte=end) |
                Q(approval_date__date__lte=end)
            )
            proposal_queryset = proposal_queryset.filter(submission_date__date__lte=end)
        except ValueError:
            pass

    if selected_status and selected_status != '':
        project_queryset = project_queryset.filter(status__iexact=selected_status)
        proposal_queryset = proposal_queryset.filter(status__iexact=selected_status)

    # -----------------------------
    # Total Budget (calculated from budget allocations)
    # -----------------------------
    total_budget = budget_queryset.aggregate(total=Sum('total_equipment_value'))['total'] or Decimal('0.00')
    total_delivered = budget_queryset.aggregate(total=Sum('delivered_equipment_value'))['total'] or Decimal('0.00')
    total_spent = total_delivered
    total_remaining = total_budget - total_spent

    total_budget_int = int(total_budget)
    total_spent_int = int(total_spent)
    total_remaining_int = int(total_remaining)
    
    # Utilization rate
    utilization_rate = round((float(total_spent) / float(total_budget) * 100), 2) if total_budget > 0 else 0

    # -----------------------------
    # Proposal Status Counts
    # -----------------------------
    proposal_counts_qs = proposal_queryset.values('status').annotate(total=Count('id'))
    proposal_status_counts = {'pending': 0, 'approved': 0, 'rejected': 0, 'for_review': 0, 'needs_revision': 0}
    for p in proposal_counts_qs:
        if p['status'] in proposal_status_counts:
            proposal_status_counts[p['status']] = p['total']
    
    total_proposals = sum(proposal_status_counts.values())

    # -----------------------------
    # Project Status Counts
    # -----------------------------
    project_status_counts = {'New': 0, 'Ongoing': 0, 'Completed': 0, 'Terminated': 0}
    project_counts_qs = project_queryset.values('status').annotate(total=Count('id'))
    for p in project_counts_qs:
        status = p['status'] or 'New'
        # Normalize status
        if status.lower() in ['new', 'proposal']:
            project_status_counts['New'] += p['total']
        elif status.lower() in ['ongoing', 'on-going', 'in_progress']:
            project_status_counts['Ongoing'] += p['total']
        elif status.lower() in ['completed', 'complete']:
            project_status_counts['Completed'] += p['total']
        elif status.lower() in ['terminated', 'cancelled']:
            project_status_counts['Terminated'] += p['total']
        else:
            project_status_counts['New'] += p['total']
    
    total_projects = sum(project_status_counts.values())

    projects_new = project_status_counts.get('New', 0)
    projects_ongoing = project_status_counts.get('Ongoing', 0)
    projects_completed = project_status_counts.get('Completed', 0)
    projects_terminated = project_status_counts.get('Terminated', 0)

    # -----------------------------
    # Project Completion Rate
    # -----------------------------
    completed_projects = project_status_counts.get('Completed', 0)
    project_completion_rate = round((completed_projects / total_projects * 100), 2) if total_projects > 0 else 0

    # -----------------------------
    # Task Status Counts (filtered to projects in scope)
    # -----------------------------
    project_ids = list(project_queryset.values_list('id', flat=True))
    scoped_tasks = task_queryset.filter(project__in=project_ids) if project_ids else task_queryset.none()
    total_tasks = scoped_tasks.count()
    tasks_pending = scoped_tasks.filter(status='pending').count()
    tasks_in_progress = scoped_tasks.filter(status='in_progress').count()
    tasks_completed = scoped_tasks.filter(status='completed').count()
    tasks_delayed = scoped_tasks.filter(status='delayed').count()
    task_status_counts = {
        'pending': tasks_pending,
        'in_progress': tasks_in_progress,
        'completed': tasks_completed,
        'delayed': tasks_delayed,
    }

    # -----------------------------
    # User Role Counts
    # -----------------------------
    user_role_counts = {}
    user_counts_qs = User.objects.values('role').annotate(total=Count('id'))
    for u in user_counts_qs:
        role_label = u['role'].replace('_', ' ').title()
        user_role_counts[role_label] = u['total']
    
    total_users = sum(user_role_counts.values())

    # -----------------------------
    # Projects by Municipality
    # -----------------------------
    municipality_counts = {}
    mun_qs = project_queryset.values('mun').annotate(total=Count('id'))
    for m in mun_qs:
        mun_name = m['mun'] or 'Unspecified'
        municipality_counts[mun_name] = m['total']

    top_municipality_name = '-'
    top_municipality_count = 0
    if municipality_counts:
        top_municipality_name, top_municipality_count = max(
            municipality_counts.items(), key=lambda item: item[1]
        )

    # -----------------------------
    # Top Projects by Beneficiaries (Top 10)
    # -----------------------------
    projects = project_queryset.all().order_by('-total_beneficiaries')[:10]
    scatter_labels = [p.project_title[:30] + '...' if len(p.project_title or '') > 30 else (p.project_title or 'Untitled') for p in projects]
    scatter_data = [int(p.total_beneficiaries or 0) for p in projects]

    # -----------------------------
    # Proposals by Municipality
    # -----------------------------
    proposal_municipality_counts = {}
    proposal_mun_qs = proposal_queryset.values('municipality').annotate(total=Count('id'))
    for m in proposal_mun_qs:
        mun_name = m['municipality'] or 'Unspecified'
        proposal_municipality_counts[mun_name] = m['total']

    # -----------------------------
    # Projects by Program and Type
    # -----------------------------
    program_counts = {}
    program_qs = project_queryset.values('program').annotate(total=Count('id')).order_by('-total')
    for row in program_qs:
        program_name = row['program'] or 'Unspecified'
        program_counts[program_name] = row['total']

    type_counts = {}
    type_qs = project_queryset.values('type_of_project').annotate(total=Count('id')).order_by('-total')
    for row in type_qs:
        type_name = row['type_of_project'] or 'Unspecified'
        type_counts[type_name] = row['total']

    # -----------------------------
    # Top Proponents by Proposal Count
    # -----------------------------
    top_proponent_labels = []
    top_proponent_values = []
    top_prop_qs = proposal_queryset.exclude(submitted_by__isnull=True).values(
        'submitted_by__first_name', 'submitted_by__last_name', 'submitted_by__username'
    ).annotate(total=Count('id')).order_by('-total')[:10]
    for row in top_prop_qs:
        full_name = f"{row.get('submitted_by__first_name') or ''} {row.get('submitted_by__last_name') or ''}".strip()
        if not full_name:
            full_name = row.get('submitted_by__username') or 'Unknown'
        top_proponent_labels.append(full_name)
        top_proponent_values.append(row['total'])

    # -----------------------------
    # Equipment Allocation Summary
    # -----------------------------
    # Summary counts and aggregates based on BudgetAllocation (equipment-based budgeting)
    total_allocations = BudgetAllocation.objects.count()
    total_allocated_items = BudgetAllocation.objects.aggregate(total=Sum('allocated_quantity'))['total'] or 0
    total_delivered_items = BudgetAllocation.objects.aggregate(total=Sum('delivered_quantity'))['total'] or 0
    allocations_pending = BudgetAllocation.objects.filter(status='allocated').count()
    allocations_delivered = BudgetAllocation.objects.filter(status='delivered').count()
    allocation_status_counts = {
        'pending': allocations_pending,
        'delivered': allocations_delivered,
    }

    # Top Equipment by Allocated Quantity (for charts)
    top_equipment_qs = BudgetAllocation.objects.values('equipment_item__name').annotate(total_allocated=Sum('allocated_quantity')).order_by('-total_allocated')[:10]
    top_equipment_labels = [e['equipment_item__name'] for e in top_equipment_qs]
    top_equipment_values = [e['total_allocated'] for e in top_equipment_qs]
    # Total budgets count for summary cards
    total_budgets = Budget.objects.count()

    # -----------------------------
    # Monthly Project Trend (All Years or Selected Year)
    # -----------------------------
    monthly_projects = [0] * 12
    if selected_year:
        # Filter by selected year
        filter_year = int(selected_year)
        for p in project_queryset.filter(project_start__year=filter_year):
            if p.project_start:
                monthly_projects[p.project_start.month - 1] += 1
    else:
        # Show all years combined
        for p in project_queryset.all():
            if p.project_start:
                monthly_projects[p.project_start.month - 1] += 1

    # Get filter options for the form
    available_years = sorted(list(Project.objects.values_list('year', flat=True).distinct().exclude(year__isnull=True)))
    available_municipalities = sorted(list(Project.objects.values_list('mun', flat=True).distinct().exclude(mun__isnull=True).exclude(mun='')))
    available_statuses = sorted(list(Project.objects.values_list('status', flat=True).distinct().exclude(status__isnull=True).exclude(status='')))

    # -----------------------------
    # DOST Tranche Summary
    # -----------------------------
    tranche_queryset = TrancheRelease.objects.all()
    if selected_year:
        try:
            year_int = int(selected_year)
            tranche_queryset = tranche_queryset.filter(release_date__year=year_int)
        except ValueError:
            pass
    
    # Tranche statistics - using correct field names
    tranche_status_counts = {
        'pending': tranche_queryset.filter(liquidation_status='pending').count(),
        'released': tranche_queryset.filter(is_released=True).count(),
        'partially_liquidated': tranche_queryset.filter(liquidation_status='partial').count(),
        'fully_liquidated': tranche_queryset.filter(liquidation_status='approved').count(),
    }
    total_tranches = sum(tranche_status_counts.values())
    
    # Total tranche amounts - using correct field name (liquidation_amount)
    total_tranche_amount = tranche_queryset.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_liquidated_amount = tranche_queryset.aggregate(total=Sum('liquidation_amount'))['total'] or Decimal('0.00')
    liquidation_rate = round((float(total_liquidated_amount) / float(total_tranche_amount) * 100), 2) if total_tranche_amount > 0 else 0
    
    # Tranches by project (top 10)
    tranche_by_project = {}
    for tranche in tranche_queryset.select_related('project')[:50]:
        project_title = tranche.project.project_title[:25] + '...' if len(tranche.project.project_title or '') > 25 else (tranche.project.project_title or 'Untitled')
        if project_title not in tranche_by_project:
            tranche_by_project[project_title] = {'released': 0, 'liquidated': 0}
        tranche_by_project[project_title]['released'] += float(tranche.amount or 0)
        tranche_by_project[project_title]['liquidated'] += float(tranche.liquidation_amount or 0)
    
    # Get top 10 projects by tranche amount
    sorted_tranche_projects = sorted(tranche_by_project.items(), key=lambda x: x[1]['released'], reverse=True)[:10]
    tranche_project_labels = [item[0] for item in sorted_tranche_projects]
    tranche_released_amounts = [item[1]['released'] for item in sorted_tranche_projects]
    tranche_liquidated_amounts = [item[1]['liquidated'] for item in sorted_tranche_projects]

    # -----------------------------
    # DOST Equipment Tracking
    # -----------------------------
    equipment_queryset = ProjectEquipment.objects.all()
    if selected_year:
        try:
            year_int = int(selected_year)
            equipment_queryset = equipment_queryset.filter(project__year=year_int)
        except ValueError:
            pass
    
    # Equipment ownership statistics
    equipment_ownership_counts = {
        'dost_owned': equipment_queryset.filter(ownership_status='dost_owned').count(),
        'beneficiary_owned': equipment_queryset.filter(ownership_status='beneficiary_owned').count(),
        'transfer_pending': equipment_queryset.filter(ownership_status='transfer_pending').count(),
    }
    total_equipment = sum(equipment_ownership_counts.values())
    
    # Equipment with property tags assigned
    equipment_with_tags = equipment_queryset.exclude(property_tag_number__isnull=True).exclude(property_tag_number='').count()
    equipment_tag_rate = round((equipment_with_tags / total_equipment * 100), 2) if total_equipment > 0 else 0
    
    # Equipment by category
    equipment_by_category = {}
    for eq in equipment_queryset.select_related('budget_allocation__equipment_item__category')[:100]:
        try:
            category_name = eq.budget_allocation.equipment_item.category.name if eq.budget_allocation and eq.budget_allocation.equipment_item and eq.budget_allocation.equipment_item.category else 'Uncategorized'
        except:
            category_name = 'Uncategorized'
        if category_name not in equipment_by_category:
            equipment_by_category[category_name] = 0
        equipment_by_category[category_name] += 1

    # -----------------------------
    # TNA Status Summary (Beneficiaries)
    # -----------------------------
    beneficiaries = User.objects.filter(role='beneficiary')
    tna_status_counts = {
        'not_started': beneficiaries.filter(tna_status='not_started').count(),
        'in_progress': beneficiaries.filter(tna_status='in_progress').count(),
        'completed': beneficiaries.filter(tna_status='completed').count(),
        'expired': beneficiaries.filter(tna_status='expired').count(),
    }
    total_beneficiaries = sum(tna_status_counts.values())
    tna_completion_rate = round((tna_status_counts['completed'] / total_beneficiaries * 100), 2) if total_beneficiaries > 0 else 0

    context = {
        # Financial
        'total_budget': total_budget_int,
        'total_spent': total_spent_int,
        'total_remaining': total_remaining_int,
        'utilization_rate': utilization_rate,
        
        # Proposals
        'proposal_status_counts': json.dumps(proposal_status_counts),
        'total_proposals': total_proposals,
        
        # Projects
        'project_status_counts': json.dumps(project_status_counts),
        'total_projects': total_projects,
        'projects_new': projects_new,
        'projects_ongoing': projects_ongoing,
        'projects_completed': projects_completed,
        'projects_terminated': projects_terminated,
        'project_completion_rate': project_completion_rate,
        'municipality_counts': json.dumps(municipality_counts),
        'top_municipality_name': top_municipality_name,
        'top_municipality_count': top_municipality_count,
        
        # Users
        'user_role_counts': json.dumps(user_role_counts),
        'total_users': total_users,
        
    # Equipment allocations summary
    'total_allocations': total_allocations,
    'total_allocated_items': int(total_allocated_items),
    'total_delivered_items': int(total_delivered_items),
    'allocations_pending': allocations_pending,
    'allocations_delivered': allocations_delivered,
    'allocation_status_counts': json.dumps(allocation_status_counts),

    # Tasks
    'total_tasks': total_tasks,
    'tasks_pending': tasks_pending,
    'tasks_in_progress': tasks_in_progress,
    'tasks_completed': tasks_completed,
    'tasks_delayed': tasks_delayed,
    'task_status_counts': json.dumps(task_status_counts),

    # Charts data
    'scatter_labels': json.dumps(scatter_labels),
    'scatter_data': json.dumps(scatter_data),
    'top_equipment_labels': json.dumps(top_equipment_labels),
    'top_equipment_values': json.dumps(top_equipment_values),
    'monthly_projects': json.dumps(monthly_projects),
        'proposal_municipality_counts': json.dumps(proposal_municipality_counts),
        'project_program_counts': json.dumps(program_counts),
        'project_type_counts': json.dumps(type_counts),
        'top_proponent_labels': json.dumps(top_proponent_labels),
        'top_proponent_values': json.dumps(top_proponent_values),
        
        # DOST Tranche Summary
        'tranche_status_counts': json.dumps(tranche_status_counts),
        'total_tranches': total_tranches,
        'total_tranche_amount': int(total_tranche_amount),
        'total_liquidated_amount': int(total_liquidated_amount),
        'liquidation_rate': liquidation_rate,
        'tranche_project_labels': json.dumps(tranche_project_labels),
        'tranche_released_amounts': json.dumps(tranche_released_amounts),
        'tranche_liquidated_amounts': json.dumps(tranche_liquidated_amounts),
        
        # DOST Equipment Tracking
        'equipment_ownership_counts': json.dumps(equipment_ownership_counts),
        'total_equipment': total_equipment,
        'equipment_with_tags': equipment_with_tags,
        'equipment_tag_rate': equipment_tag_rate,
        'equipment_by_category': json.dumps(equipment_by_category),
        
    # NOTE: TNA status removed from this report view - equipment/budget focused metrics are shown instead
    'total_budgets': total_budgets,
        
        # Filters
        'available_years': available_years,
        'available_municipalities': available_municipalities,
        'available_statuses': available_statuses,
        'selected_year': selected_year,
        'selected_municipality': selected_municipality,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': selected_status,
        
        'report_year': selected_year or 'All Years',
        'report_date': timezone.now().strftime('%B %d, %Y'),
        
        # User signature for PDF export
        'user_signature': DigitalSignature.objects.filter(user=request.user).order_by('-signed_at').first(),
    }

    return render(request, 'administrator/reports.html', context)
def administrator_settings_view(request):
    user = request.user  # currently logged-in user
    
    # Get user's existing signature
    user_signature = DigitalSignature.objects.filter(user=user).order_by('-signed_at').first()

    if request.method == "POST":
        # Capture old data
        old_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Update user fields
        user.first_name = request.POST.get('first_name')
        user.middle_name = request.POST.get('middle_name')
        user.last_name = request.POST.get('last_name')
        user.suffix = request.POST.get('suffix')
        user.sex = request.POST.get('sex')
        user.civil_status = request.POST.get('civil_status')
        user.contact_number = request.POST.get('contact_number')
        user.email = request.POST.get('email')
        user.address = request.POST.get('address')

        if not apply_profile_picture_update(request, user):
            return redirect('administrator_settings_url')

        user.save()
        
        # Handle signature upload
        if 'signature_image' in request.FILES:
            sig_file = request.FILES['signature_image']
            # Create or update signature
            if user_signature:
                user_signature.signature_image = sig_file
                user_signature.save()
            else:
                DigitalSignature.objects.create(
                    user=user,
                    signature_image=sig_file
                )

        # Capture new data
        new_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        messages.success(request, "Your profile has been updated successfully!")
        return redirect('administrator_settings_url')

    return render(request, 'administrator/settings.html', {
        'user': user,
        'user_signature': user_signature
    })


@login_required
def administrator_change_password_view(request):
    if request.method == "POST":
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        user = request.user

        # Check if old password matches
        if not user.check_password(old_password):
            messages.error(request, "Your old password is incorrect.")
            return redirect('administrator_settings_url')

        # Validate new passwords
        if new_password1 != new_password2:
            messages.error(request, "New passwords do not match.")
            return redirect('administrator_settings_url')

        # Capture old data (just note that password was changed)
        old_data = {"password": "********"}
        new_data = {"password": "********"}

        # Set and save new password
        user.set_password(new_password1)
        user.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        # Log the user out after successful password change
        logout(request)
        messages.success(request, "Your password has been changed. Please log in again.")
        return redirect('logout_url')

    return redirect('administrator_settings_url')




# -------------------------
# View Audit Logs
# -------------------------
@login_required
def administrator_audit_logs_view(request):
    import json
    # You can filter, sort, or paginate if needed
    logs = AuditLog.objects.all().order_by('-timestamp')  # newest first
    
    # Calculate action counts for stat cards
    create_count = logs.filter(action__iexact='create').count()
    update_count = logs.filter(action__iexact='update').count()
    delete_count = logs.filter(action__iexact='delete').count()
    login_count = logs.filter(action__iexact='login').count()
    
    # Serialize audit log data for JavaScript
    logs_json_data = {}
    for log in logs:
        old_data = log.old_data
        new_data = log.new_data

        # Backward compatibility: some historical rows stored JSON as strings.
        if isinstance(old_data, str):
            try:
                old_data = json.loads(old_data)
            except (TypeError, json.JSONDecodeError):
                old_data = {'raw': old_data}
        if isinstance(new_data, str):
            try:
                new_data = json.loads(new_data)
            except (TypeError, json.JSONDecodeError):
                new_data = {'raw': new_data}

        logs_json_data[str(log.id)] = {
            'action': log.get_action_display() if hasattr(log, 'get_action_display') else log.action.capitalize(),
            'model_name': log.model_name,
            'object_id': log.object_id,
            'old_data': old_data if old_data else None,
            'new_data': new_data if new_data else None,
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else None,
            'user': log.user.get_full_name() if log.user else 'System',
            'reason': log.reason if hasattr(log, 'reason') and log.reason else None,
            'ip_address': log.ip_address if hasattr(log, 'ip_address') and log.ip_address else None,
        }

    context = {
        'logs': logs,
        'logs_json': json.dumps(logs_json_data),
        'create_count': create_count,
        'update_count': update_count,
        'delete_count': delete_count,
        'login_count': login_count,
    }
    return render(request, 'administrator/audit_logs.html', context)


# =========================================
# EXCEL EXPORT VIEWS
# =========================================

@login_required
def export_projects_excel(request):
    """Export all projects to Excel file - matches the Projects table structure"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers - matching the projects table exactly
    headers = [
        # Identification
        "No.", "Code", "Year",
        # Basic Info
        "Project Title", "Agency/Grantee", "Program", "Type", "Status", "Remarks",
        # Location
        "Municipality", "Province", "District",
        # Beneficiaries
        "Beneficiary", "Beneficiary Address", "Contact", "Proponent", 
        "No. Beneficiaries", "Male", "Female", "Total", "Senior Citizen", "PWD",
        # Financials
        "Fund Source", "Approved Budget", "Total Project Cost", "Counterpart Fund", 
        "Internal Managed Fund", "Total Released",
        # Tranches
        "1st Tranche", "2nd Tranche", "3rd Tranche",
        # Dates & Timeline
        "Start Date", "End Date", "Release Date", "Completion Date", "Duration", "Extension Date",
        # Liquidation
        "Check/ADA No.", "Liquidation Status", "Liquidation Date", "Amount Liquidated",
        # Tech & Interventions
        "Technologies Availed", "Interventions",
        # Documents & Status
        "TAFR", "PAR", "Terminal Report", "Invoice/Receipt", "Equipment List",
        # Donation
        "Donated", "Donation Date", "Donation Status",
        # Other
        "PME Visit", "Women's Group", "Inspection/Tagging Date", "Receipt by Grantee"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    projects = Project.objects.all().order_by('no', '-date_created')
    
    for row_num, project in enumerate(projects, 2):
        data = [
            # Identification
            project.no or '-',
            project.project_code or '-',
            project.year or '-',
            # Basic Info
            project.project_title or '-',
            project.agency_grantee or '-',
            project.program or '-',
            project.type_of_project or '-',
            project.status or '-',
            project.remarks or '-',
            # Location
            project.mun or '-',
            project.province or '-',
            project.district or '-',
            # Beneficiaries
            project.beneficiary or '-',
            project.beneficiary_address or '-',
            project.contact_details or '-',
            project.proponent_details or '-',
            project.no_of_beneficiaries or '-',
            project.male or '-',
            project.female or '-',
            project.total_beneficiaries or '-',
            project.senior_citizen or '-',
            project.pwd or '-',
            # Financials
            project.fund_source or '-',
            float(project.funds or 0) if project.funds else '-',
            float(project.total_project_cost or 0) if project.total_project_cost else '-',
            float(project.counterpart_funds or 0) if project.counterpart_funds else '-',
            float(project.internally_managed_fund or 0) if project.internally_managed_fund else '-',
            float(project.total_funds_released or 0) if project.total_funds_released else '-',
            # Tranches
            float(project.first_tranche or 0) if project.first_tranche else '-',
            float(project.second_tranche or 0) if project.second_tranche else '-',
            float(project.third_tranche or 0) if project.third_tranche else '-',
            # Dates & Timeline
            project.project_start.strftime('%Y-%m-%d') if project.project_start else '-',
            project.project_end.strftime('%Y-%m-%d') if project.project_end else '-',
            project.date_of_release.strftime('%Y-%m-%d') if project.date_of_release else '-',
            project.date_of_completion.strftime('%Y-%m-%d') if project.date_of_completion else '-',
            project.original_project_duration or '-',
            project.extension_date or '-',
            # Liquidation
            project.check_ada_no or '-',
            project.status_of_liquidation or '-',
            project.date_of_liquidation.strftime('%Y-%m-%d') if project.date_of_liquidation else '-',
            float(project.amount_liquidated or 0) if project.amount_liquidated else '-',
            # Tech & Interventions
            project.availed_technologies or '-',
            project.interventions or '-',
            # Documents & Status
            project.tafr or '-',
            project.par or '-',
            project.terminal_report or '-',
            project.invoice_receipt or '-',
            project.list_of_eqpt or '-',
            # Donation
            project.donated or '-',
            project.date_of_donation.strftime('%Y-%m-%d') if project.date_of_donation else '-',
            project.donation_status or '-',
            # Other
            project.pme_visit or '-',
            project.womens_group or '-',
            project.date_of_inspection_tagging.strftime('%Y-%m-%d') if project.date_of_inspection_tagging else '-',
            project.acknowledgment_receipt_by_grantee or '-',
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    column_widths = [
        6, 15, 6,  # Identification
        40, 25, 12, 15, 12, 20,  # Basic Info
        12, 12, 8,  # Location
        25, 30, 20, 25, 8, 6, 6, 6, 6, 6,  # Beneficiaries
        15, 15, 15, 15, 15, 15,  # Financials
        12, 12, 12,  # Tranches
        12, 12, 12, 12, 15, 15,  # Dates
        15, 15, 12, 15,  # Liquidation
        30, 30,  # Tech
        10, 10, 12, 12, 20,  # Documents
        10, 12, 15,  # Donation
        10, 12, 12, 15  # Other
    ]
    for i, width in enumerate(column_widths, 1):
        if i <= len(column_widths):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="projects_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_budgets_excel(request):
    """Export all budgets to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budgets"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Fiscal Year", "Fund Source", "Total Amount", "Remaining Amount",
        "Status", "Date Allocated", "Created By"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    budgets = Budget.objects.all().order_by('-fiscal_year')
    
    for row_num, budget in enumerate(budgets, 2):
        created_by = budget.created_by.full_name() if budget.created_by else '-'
        data = [
            budget.fiscal_year,
            budget.fund_source or '-',
            float(budget.total_amount or 0),
            float(budget.remaining_amount or 0),
            budget.status or '-',
            budget.date_allocated.strftime('%Y-%m-%d') if budget.date_allocated else '-',
            created_by
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    column_widths = [12, 30, 18, 18, 12, 15, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="budgets_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_proposals_excel(request):
    """Export all proposals to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposals"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Title", "Status", "Proposed Amount", "Approved Amount",
        "Submitted By", "Submission Date", "Processed By", "Beneficiary",
        "Location", "Municipality", "Province"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    proposals = Proposal.objects.all().order_by('-submission_date')
    
    for row_num, proposal in enumerate(proposals, 2):
        submitted_by = proposal.submitted_by.full_name() if proposal.submitted_by else '-'
        processed_by = proposal.processed_by.full_name() if proposal.processed_by else '-'
        beneficiary = proposal.beneficiary.full_name() if proposal.beneficiary else '-'
        
        data = [
            proposal.title or '-',
            proposal.get_status_display() or '-',
            float(proposal.proposed_amount or 0),
            float(proposal.approved_amount or 0) if proposal.approved_amount else '-',
            submitted_by,
            proposal.submission_date.strftime('%Y-%m-%d') if proposal.submission_date else '-',
            processed_by,
            beneficiary,
            proposal.location or '-',
            proposal.municipality or '-',
            proposal.province or '-'
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    column_widths = [40, 15, 18, 18, 25, 15, 25, 25, 25, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="proposals_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_tasks_excel(request):
    """Export all tasks to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Task Title", "Project", "Description", "Assigned To",
        "Status", "Due Date", "Completion Date", "Location"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    tasks = Task.objects.all().select_related('project', 'assigned_to').order_by('-due_date')
    
    for row_num, task in enumerate(tasks, 2):
        assigned_to = task.assigned_to.full_name() if task.assigned_to else '-'
        project_title = task.project.project_title if task.project else '-'
        
        data = [
            task.title or '-',
            project_title,
            task.description or '-',
            assigned_to,
            task.get_status_display() or '-',
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '-',
            task.completion_date.strftime('%Y-%m-%d') if task.completion_date else '-',
            task.location_name or '-'
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    column_widths = [30, 40, 50, 25, 15, 12, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tasks_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_master_report_excel(request):
    """
    Export a comprehensive master report with all data in separate sheets:
    - Summary Dashboard
    - Budgets
    - Proposals
    - Projects
    - Tasks
    """
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0072CE", end_color="0072CE", fill_type="solid")
    title_font = Font(bold=True, size=14, color="0072CE")
    subtitle_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '₱#,##0.00'
    
    def style_header_row(ws, row, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # =====================
    # SHEET 1: SUMMARY
    # =====================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Get summary data
    total_budgets = Budget.objects.count()
    total_budget_amount = Budget.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    total_remaining = Budget.objects.aggregate(total=Sum('remaining_amount'))['total'] or 0
    total_spent = float(total_budget_amount) - float(total_remaining)
    
    total_proposals = Proposal.objects.count()
    pending_proposals = Proposal.objects.filter(status='pending').count()
    approved_proposals = Proposal.objects.filter(status='approved').count()
    rejected_proposals = Proposal.objects.filter(status='rejected').count()
    
    total_projects = Project.objects.count()
    ongoing_projects = Project.objects.filter(status__iexact='ongoing').count()
    completed_projects = Project.objects.filter(status__iexact='completed').count()
    
    total_tasks = Task.objects.count()
    pending_tasks = Task.objects.filter(status='pending').count()
    completed_tasks = Task.objects.filter(status='completed').count()
    
    # Title
    ws_summary['A1'] = "DOST Biliran - Master Report Summary"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws_summary['A2'].font = Font(italic=True, color="666666")
    
    # Budget Summary
    ws_summary['A4'] = "BUDGET SUMMARY"
    ws_summary['A4'].font = subtitle_font
    ws_summary['A5'] = "Total Budgets:"
    ws_summary['B5'] = total_budgets
    ws_summary['A6'] = "Total Allocated:"
    ws_summary['B6'] = float(total_budget_amount)
    ws_summary['B6'].number_format = currency_format
    ws_summary['A7'] = "Total Spent:"
    ws_summary['B7'] = total_spent
    ws_summary['B7'].number_format = currency_format
    ws_summary['A8'] = "Total Remaining:"
    ws_summary['B8'] = float(total_remaining)
    ws_summary['B8'].number_format = currency_format
    
    # Proposal Summary
    ws_summary['A10'] = "PROPOSAL SUMMARY"
    ws_summary['A10'].font = subtitle_font
    ws_summary['A11'] = "Total Proposals:"
    ws_summary['B11'] = total_proposals
    ws_summary['A12'] = "Pending:"
    ws_summary['B12'] = pending_proposals
    ws_summary['A13'] = "Approved:"
    ws_summary['B13'] = approved_proposals
    ws_summary['A14'] = "Rejected:"
    ws_summary['B14'] = rejected_proposals
    
    # Project Summary
    ws_summary['A16'] = "PROJECT SUMMARY"
    ws_summary['A16'].font = subtitle_font
    ws_summary['A17'] = "Total Projects:"
    ws_summary['B17'] = total_projects
    ws_summary['A18'] = "Ongoing:"
    ws_summary['B18'] = ongoing_projects
    ws_summary['A19'] = "Completed:"
    ws_summary['B19'] = completed_projects
    
    # Task Summary
    ws_summary['A21'] = "TASK SUMMARY"
    ws_summary['A21'].font = subtitle_font
    ws_summary['A22'] = "Total Tasks:"
    ws_summary['B22'] = total_tasks
    ws_summary['A23'] = "Pending:"
    ws_summary['B23'] = pending_tasks
    ws_summary['A24'] = "Completed:"
    ws_summary['B24'] = completed_tasks
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 20
    
    # =====================
    # SHEET 2: BUDGETS
    # =====================
    ws_budgets = wb.create_sheet("Budgets")
    budget_headers = ["Fiscal Year", "Fund Source", "Total Amount", "Remaining Amount", "Spent", "Status", "Date Allocated"]
    
    for col, header in enumerate(budget_headers, 1):
        ws_budgets.cell(row=1, column=col, value=header)
    style_header_row(ws_budgets, 1, len(budget_headers))
    
    budgets = Budget.objects.all().order_by('-fiscal_year')
    for row_num, budget in enumerate(budgets, 2):
        spent = float(budget.total_amount or 0) - float(budget.remaining_amount or 0)
        data = [
            budget.fiscal_year,
            budget.fund_source or '-',
            float(budget.total_amount or 0),
            float(budget.remaining_amount or 0),
            spent,
            budget.status or '-',
            budget.date_allocated.strftime('%Y-%m-%d') if budget.date_allocated else '-'
        ]
        for col, value in enumerate(data, 1):
            cell = ws_budgets.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in [3, 4, 5]:
                cell.number_format = currency_format
    
    for i, width in enumerate([12, 30, 18, 18, 18, 12, 15], 1):
        ws_budgets.column_dimensions[get_column_letter(i)].width = width
    ws_budgets.freeze_panes = 'A2'
    
    # =====================
    # SHEET 3: PROPOSALS
    # =====================
    ws_proposals = wb.create_sheet("Proposals")
    proposal_headers = ["Title", "Status", "Proposed Amount", "Approved Amount", "Submitted By", 
                        "Submission Date", "Processed By", "Beneficiary", "Municipality", "Province"]
    
    for col, header in enumerate(proposal_headers, 1):
        ws_proposals.cell(row=1, column=col, value=header)
    style_header_row(ws_proposals, 1, len(proposal_headers))
    
    proposals = Proposal.objects.all().select_related('submitted_by', 'processed_by', 'beneficiary').order_by('-submission_date')
    for row_num, proposal in enumerate(proposals, 2):
        data = [
            proposal.title or '-',
            proposal.get_status_display() or '-',
            float(proposal.proposed_amount or 0),
            float(proposal.approved_amount or 0) if proposal.approved_amount else '-',
            proposal.submitted_by.full_name() if proposal.submitted_by else '-',
            proposal.submission_date.strftime('%Y-%m-%d') if proposal.submission_date else '-',
            proposal.processed_by.full_name() if proposal.processed_by else '-',
            proposal.beneficiary.full_name() if proposal.beneficiary else '-',
            proposal.municipality or '-',
            proposal.province or '-'
        ]
        for col, value in enumerate(data, 1):
            cell = ws_proposals.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in [3, 4] and isinstance(value, float):
                cell.number_format = currency_format
    
    for i, width in enumerate([40, 15, 18, 18, 25, 15, 25, 25, 15, 15], 1):
        ws_proposals.column_dimensions[get_column_letter(i)].width = width
    ws_proposals.freeze_panes = 'A2'
    
    # =====================
    # SHEET 4: PROJECTS (Full 54-column structure matching projects table)
    # =====================
    ws_projects = wb.create_sheet("Projects")
    project_headers = [
        # Identification
        "No.", "Code", "Year",
        # Basic Info
        "Project Title", "Agency/Grantee", "Program", "Type", "Status", "Remarks",
        # Location
        "Municipality", "Province", "District",
        # Beneficiaries
        "Beneficiary", "Beneficiary Address", "Contact", "Proponent", 
        "No. Beneficiaries", "Male", "Female", "Total", "Senior Citizen", "PWD",
        # Financials
        "Fund Source", "Approved Budget", "Total Project Cost", "Counterpart Fund", 
        "Internal Managed Fund", "Total Released",
        # Tranches
        "1st Tranche", "2nd Tranche", "3rd Tranche",
        # Dates & Timeline
        "Start Date", "End Date", "Release Date", "Completion Date", "Duration", "Extension Date",
        # Liquidation
        "Check/ADA No.", "Liquidation Status", "Liquidation Date", "Amount Liquidated",
        # Tech & Interventions
        "Technologies Availed", "Interventions",
        # Documents & Status
        "TAFR", "PAR", "Terminal Report", "Invoice/Receipt", "Equipment List",
        # Donation
        "Donated", "Donation Date", "Donation Status",
        # Other
        "PME Visit", "Women's Group", "Inspection/Tagging Date", "Receipt by Grantee"
    ]
    
    for col, header in enumerate(project_headers, 1):
        ws_projects.cell(row=1, column=col, value=header)
    style_header_row(ws_projects, 1, len(project_headers))
    
    projects = Project.objects.all().order_by('no', '-date_created')
    for row_num, project in enumerate(projects, 2):
        data = [
            # Identification
            project.no or '-',
            project.project_code or '-',
            project.year or '-',
            # Basic Info
            project.project_title or '-',
            project.agency_grantee or '-',
            project.program or '-',
            project.type_of_project or '-',
            project.status or '-',
            project.remarks or '-',
            # Location
            project.mun or '-',
            project.province or '-',
            project.district or '-',
            # Beneficiaries
            project.beneficiary or '-',
            project.beneficiary_address or '-',
            project.contact_details or '-',
            project.proponent_details or '-',
            project.no_of_beneficiaries or '-',
            project.male or '-',
            project.female or '-',
            project.total_beneficiaries or '-',
            project.senior_citizen or '-',
            project.pwd or '-',
            # Financials
            project.fund_source or '-',
            float(project.funds or 0) if project.funds else '-',
            float(project.total_project_cost or 0) if project.total_project_cost else '-',
            float(project.counterpart_funds or 0) if project.counterpart_funds else '-',
            float(project.internally_managed_fund or 0) if project.internally_managed_fund else '-',
            float(project.total_funds_released or 0) if project.total_funds_released else '-',
            # Tranches
            float(project.first_tranche or 0) if project.first_tranche else '-',
            float(project.second_tranche or 0) if project.second_tranche else '-',
            float(project.third_tranche or 0) if project.third_tranche else '-',
            # Dates & Timeline
            project.project_start.strftime('%Y-%m-%d') if project.project_start else '-',
            project.project_end.strftime('%Y-%m-%d') if project.project_end else '-',
            project.date_of_release.strftime('%Y-%m-%d') if project.date_of_release else '-',
            project.date_of_completion.strftime('%Y-%m-%d') if project.date_of_completion else '-',
            project.original_project_duration or '-',
            project.extension_date or '-',
            # Liquidation
            project.check_ada_no or '-',
            project.status_of_liquidation or '-',
            project.date_of_liquidation.strftime('%Y-%m-%d') if project.date_of_liquidation else '-',
            float(project.amount_liquidated or 0) if project.amount_liquidated else '-',
            # Tech & Interventions
            project.availed_technologies or '-',
            project.interventions or '-',
            # Documents & Status
            project.tafr or '-',
            project.par or '-',
            project.terminal_report or '-',
            project.invoice_receipt or '-',
            project.list_of_eqpt or '-',
            # Donation
            project.donated or '-',
            project.date_of_donation.strftime('%Y-%m-%d') if project.date_of_donation else '-',
            project.donation_status or '-',
            # Other
            project.pme_visit or '-',
            project.womens_group or '-',
            project.date_of_inspection_tagging.strftime('%Y-%m-%d') if project.date_of_inspection_tagging else '-',
            project.acknowledgment_receipt_by_grantee or '-',
        ]
        for col, value in enumerate(data, 1):
            cell = ws_projects.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            # Apply currency format to financial columns (24-31 and 41)
            if col in [24, 25, 26, 27, 28, 29, 30, 31, 41] and isinstance(value, float):
                cell.number_format = currency_format
    
    # Column widths for all 54 columns
    project_widths = [
        6, 15, 6,  # Identification
        40, 25, 12, 15, 12, 20,  # Basic Info
        12, 12, 8,  # Location
        25, 30, 20, 25, 8, 6, 6, 6, 6, 6,  # Beneficiaries
        15, 15, 15, 15, 15, 15,  # Financials
        12, 12, 12,  # Tranches
        12, 12, 12, 12, 15, 15,  # Dates
        15, 15, 12, 15,  # Liquidation
        30, 30,  # Tech
        10, 10, 12, 12, 20,  # Documents
        10, 12, 15,  # Donation
        10, 12, 12, 15  # Other
    ]
    for i, width in enumerate(project_widths, 1):
        if i <= len(project_widths):
            ws_projects.column_dimensions[get_column_letter(i)].width = width
    ws_projects.freeze_panes = 'A2'
    
    # =====================
    # SHEET 5: TASKS
    # =====================
    ws_tasks = wb.create_sheet("Tasks")
    task_headers = ["Task Title", "Project", "Assigned To", "Status", "Due Date", "Completion Date", "Location"]
    
    for col, header in enumerate(task_headers, 1):
        ws_tasks.cell(row=1, column=col, value=header)
    style_header_row(ws_tasks, 1, len(task_headers))
    
    tasks = Task.objects.all().select_related('project', 'assigned_to').order_by('-due_date')
    for row_num, task in enumerate(tasks, 2):
        data = [
            task.title or '-',
            task.project.project_title if task.project else '-',
            task.assigned_to.full_name() if task.assigned_to else '-',
            task.get_status_display() or '-',
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '-',
            task.completion_date.strftime('%Y-%m-%d') if task.completion_date else '-',
            task.location_name or '-'
        ]
        for col, value in enumerate(data, 1):
            cell = ws_tasks.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
    
    for i, width in enumerate([30, 40, 25, 15, 12, 15, 30], 1):
        ws_tasks.column_dimensions[get_column_letter(i)].width = width
    ws_tasks.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="DOST_Master_Report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response






























# Administrator Extension Requests
@login_required
def administrator_extension_requests_view(request):
    """List all extension requests for administrator approval"""
    extension_requests = ExtensionRequest.objects.select_related(
        'proposal', 'proponent', 'proposal__project'
    ).order_by('-date_submitted')
    
    # Status counts for summary cards
    pending_count = extension_requests.filter(status='pending').count()
    approved_count = extension_requests.filter(status='approved').count()
    rejected_count = extension_requests.filter(status='rejected').count()
    total_count = extension_requests.count()
    
    context = {
        'extension_requests': extension_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'total_count': total_count,
    }
    return render(request, 'administrator/extension-requests.html', context)


@login_required
@require_POST
def administrator_extension_requests_bulk_approve_view(request):
    """Bulk approve multiple extension requests"""
    request_ids = request.POST.getlist('request_ids')
    approved_days = request.POST.get('approved_days', 30)
    admin_notes = request.POST.get('admin_notes', '')
    
    if not request_ids:
        messages.error(request, "No extension requests selected.")
        return redirect('administrator_extension_requests_url')
    
    approved_count = 0
    for request_id in request_ids:
        try:
            ext_request = ExtensionRequest.objects.get(pk=request_id, status='pending')
            ext_request.status = 'approved'
            ext_request.approved_days = approved_days
            ext_request.remarks = admin_notes
            ext_request.approved_by = request.user
            ext_request.date_approved = timezone.now()
            ext_request.save()
            
            # Update project end date if applicable
            if ext_request.proposal.project and approved_days:
                project = ext_request.proposal.project
                if project.date_of_completion:
                    project.date_of_completion += timedelta(days=int(approved_days))
                    project.save()
            
            approved_count += 1
            
            # Audit Log
            AuditLog.objects.create(
                user=request.user,
                action="bulk_approve",
                model_name="ExtensionRequest",
                object_id=str(ext_request.pk),
                old_data={"status": "pending"},
                new_data={"status": "approved", "approved_days": approved_days}
            )
        except ExtensionRequest.DoesNotExist:
            continue
    
    messages.success(request, f"Successfully approved {approved_count} extension request(s).")
    return redirect('administrator_extension_requests_url')


@login_required
@require_POST
def administrator_extension_requests_bulk_reject_view(request):
    """Bulk reject multiple extension requests"""
    request_ids = request.POST.getlist('request_ids')
    rejection_reason = request.POST.get('rejection_reason', '')
    
    if not request_ids:
        messages.error(request, "No extension requests selected.")
        return redirect('administrator_extension_requests_url')
    
    rejected_count = 0
    for request_id in request_ids:
        try:
            ext_request = ExtensionRequest.objects.get(pk=request_id, status='pending')
            ext_request.status = 'rejected'
            ext_request.remarks = rejection_reason
            ext_request.approved_by = request.user
            ext_request.date_approved = timezone.now()
            ext_request.save()
            
            rejected_count += 1
            
            # Audit Log
            AuditLog.objects.create(
                user=request.user,
                action="bulk_reject",
                model_name="ExtensionRequest",
                object_id=str(ext_request.pk),
                old_data={"status": "pending"},
                new_data={"status": "rejected", "reason": rejection_reason}
            )
        except ExtensionRequest.DoesNotExist:
            continue
    
    messages.success(request, f"Successfully rejected {rejected_count} extension request(s).")
    return redirect('administrator_extension_requests_url')


@login_required
def administrator_extension_requests_approve_view(request, pk):
    """Approve an extension request"""
    extension_request = get_object_or_404(ExtensionRequest, pk=pk)
    
    if request.method == 'POST':
        approved_days = request.POST.get('approved_days')
        admin_notes = request.POST.get('admin_notes')
        
        extension_request.status = 'approved'
        extension_request.approved_days = approved_days
        extension_request.remarks = admin_notes
        extension_request.approved_by = request.user
        extension_request.date_approved = timezone.now()
        extension_request.save()
        
        # Update the proposal/project end date if applicable
        if extension_request.proposal.project and approved_days:
            project = extension_request.proposal.project
            if project.date_of_completion:
                project.date_of_completion += timedelta(days=int(approved_days))
                project.save()
        
        # Create notification for the proponent
        proponent_url = reverse('proponent_extension_requests_url')
        Notification.objects.create(
            sender=request.user,
            receiver=extension_request.proponent,
            message=f'Your extension request for "{extension_request.proposal.title}" has been approved for {approved_days} days',
            link=proponent_url
        )
        
        messages.success(request, 'Extension request approved successfully!')
        return redirect('administrator_extension_requests_url')
    
    context = {
        'extension_request': extension_request,
    }
    return render(request, 'administrator/extension-requests-approve.html', context)


@login_required
def administrator_extension_requests_reject_view(request, pk):
    """Reject an extension request"""
    extension_request = get_object_or_404(ExtensionRequest, pk=pk)
    
    if request.method == 'POST':
        admin_notes = request.POST.get('admin_notes')
        
        extension_request.status = 'rejected'
        extension_request.remarks = admin_notes
        extension_request.approved_by = request.user
        extension_request.date_approved = timezone.now()
        extension_request.save()
        
        # Create notification for the proponent
        proponent_url = reverse('proponent_extension_requests_url')
        Notification.objects.create(
            sender=request.user,
            receiver=extension_request.proponent,
            message=f'Your extension request for "{extension_request.proposal.title}" has been rejected',
            link=proponent_url
        )
        
        messages.success(request, 'Extension request rejected.')
        return redirect('administrator_extension_requests_url')
    
    context = {
        'extension_request': extension_request,
    }
    return render(request, 'administrator/extension-requests-reject.html', context)


@login_required
def administrator_extension_requests_edit_view(request, pk):
    """Edit an extension request"""
    extension_request = get_object_or_404(ExtensionRequest, pk=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason')
        requested_extension_days = request.POST.get('requested_extension_days')
        attachment = request.FILES.get('attachment')
        
        extension_request.reason = reason
        extension_request.requested_extension_days = requested_extension_days
        if attachment:
            extension_request.letter = attachment
        extension_request.save()
        
        messages.success(request, 'Extension request updated successfully!')
        return redirect('administrator_extension_requests_url')
    
    context = {
        'extension_request': extension_request,
    }
    return render(request, 'administrator/extension-requests-edit.html', context)


@login_required
def administrator_extension_requests_delete_view(request, pk):
    """Delete an extension request"""
    extension_request = get_object_or_404(ExtensionRequest, pk=pk)
    
    if request.method == 'POST':
        extension_request.delete()
        messages.success(request, 'Extension request deleted successfully!')
        return redirect('administrator_extension_requests_url')
    
    context = {
        'extension_request': extension_request,
    }
    return render(request, 'administrator/extension-requests-delete.html', context)


@login_required
def administrator_extension_requests_add_view(request):
    """Admin creates an extension request on behalf of a proponent"""
    if request.method == 'POST':
        proponent_id = request.POST.get('proponent')
        proposal_id = request.POST.get('proposal')
        reason = request.POST.get('reason')
        requested_extension_days = request.POST.get('requested_extension_days')
        attachment = request.FILES.get('attachment')
        
        proponent = get_object_or_404(User, pk=proponent_id)
        proposal = get_object_or_404(Proposal, pk=proposal_id)
        
        extension_request = ExtensionRequest.objects.create(
            proponent=proponent,
            proposal=proposal,
            reason=reason,
            requested_extension_days=requested_extension_days,
            status='pending'
        )
        
        if attachment:
            extension_request.letter = attachment
            extension_request.save()
        
        messages.success(request, f'Extension request created successfully for {proponent.full_name}!')
        return redirect('administrator_extension_requests_url')
    
    return redirect('administrator_extension_requests_url')


# Staff
@login_required
def staff_dashboard_view(request):
    # -----------------------------
    # Tasks assigned to the current staff user, plus tasks under projects they lead
    # -----------------------------
    tasks_qs = Task.objects.select_related('project', 'assigned_to').filter(
        Q(assigned_to=request.user) | Q(project__project_leader=request.user)
    ).distinct()

    tasks = []
    for t in tasks_qs:
        project = t.project
        project_title = getattr(project, 'project_title', None) or str(project)
        assigned_to_name = t.assigned_to.full_name() if t.assigned_to else ''
        tasks.append({
            'id': t.id,
            'title': t.title or '',
            'description': t.description or '',
            'project_title': project_title,
            'assigned_to': assigned_to_name,
            'status': t.status or '',
            'due_date': t.due_date.isoformat() if t.due_date else '',
            'project_status': project.status if project else '',
            'project_municipality': project.mun if project else '',
            'project_province': project.province if project else '',
            'project_beneficiary': project.beneficiary if project else '',
            'project_proponent': project.proponent_details if project else '',
            'project_funds': float(project.funds) if project and project.funds else 0,
            'latitude': float(project.latitude) if project and project.latitude is not None else None,
            'longitude': float(project.longitude) if project and project.longitude is not None else None,
        })

    # Task counts for the current staff
    my_tasks_count = tasks_qs.count()
    pending_tasks = tasks_qs.filter(status='pending').count()
    in_progress_tasks = tasks_qs.filter(status='in_progress').count()
    completed_tasks = tasks_qs.filter(status='completed').count()
    delayed_tasks = tasks_qs.filter(status='delayed').count()

    # Projects the staff is involved in (through assigned tasks)
    my_project_ids = tasks_qs.values_list('project_id', flat=True).distinct()
    my_projects = Project.objects.filter(id__in=my_project_ids)
    my_projects_count = my_projects.count()

    projects_list = []
    for project in my_projects:
        projects_list.append({
            'id': project.id,
            'title': project.project_title or '',
            'status': project.status or '',
            'municipality': project.mun or '',
            'province': project.province or '',
            'beneficiary': project.beneficiary or '',
            'proponent': project.proponent_details or '',
            'funds': float(project.funds) if project.funds else 0,
            'latitude': float(project.latitude) if project.latitude is not None else None,
            'longitude': float(project.longitude) if project.longitude is not None else None,
        })

    context = {
        'tasks_json': json.dumps(tasks),
        'projects_json': json.dumps(projects_list),
        'proposals_json': json.dumps([]),
        'my_tasks_count': my_tasks_count,
        'pending_tasks': pending_tasks,
        'in_progress_tasks': in_progress_tasks,
        'completed_tasks': completed_tasks,
        'delayed_tasks': delayed_tasks,
        'my_projects_count': my_projects_count,
        'my_projects': my_projects,
        'tasks': tasks_qs,
    }

    return render(request, 'staff/dashboard.html', context)


@login_required
def staff_quick_actions_view(request):
    if request.user.role not in ['dost_staff', 'staff']:
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    task_scope = Task.objects.filter(
        Q(assigned_to=request.user) | Q(project__project_leader=request.user)
    ).distinct()

    context = {
        'my_pending_tasks': task_scope.filter(status='pending').count(),
        'my_in_progress_tasks': task_scope.filter(status='in_progress').count(),
        'my_overdue_tasks': task_scope.filter(status='delayed').count(),
        'my_projects_count': Project.objects.filter(id__in=task_scope.values_list('project_id', flat=True)).distinct().count(),
    }
    return render(request, 'staff/quick_actions.html', context)

# ----------------------------
# staff User Management Views
# ----------------------------


def staff_users_view(request):
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'staff/users.html', {'users': users})


@login_required
def mark_project_completed(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == "POST":
        project.status = 'completed'
        project.completion_date = timezone.now()
        project.save(update_fields=['status', 'completion_date'])
        messages.success(request, f"Project '{project.project_title}' marked as completed.")
    return redirect(request.META.get('HTTP_REFERER', 'administrator:dashboard'))

# --------------------------
# staff: Budget List
# --------------------------
def staff_budgets_view(request):
    budgets = Budget.objects.all().order_by('-date_created')
    return render(request, 'staff/budgets.html', {'budgets': budgets})

# -------------------------
# List Proposals
# -------------------------
def staff_proposals_view(request):
    # Only proposals submitted by the logged-in user
    proposals = Proposal.objects.select_related('submitted_by', 'processed_by', 'budget') \
                                .filter(submitted_by=request.user)
    
    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])

    # Fetch all proponents for the "Add Proposal" form
    proponents = User.objects.filter(role='proponent')

    # Fetch all beneficiaries
    beneficiaries = User.objects.filter(role='beneficiary')  # <- plural, match template

    context = {
        'proposals': proposals,
        'budgets': budgets,
        'proponents': proponents,
        'beneficiaries': beneficiaries,  # <- pass it to template
    }
    return render(request, 'staff/proposals.html', context)


def staff_proposals_add_view(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        proposed_amount = request.POST.get('proposed_amount') or 0
        budget_id = request.POST.get('budget')
        document = request.FILES.get('document')

        processed_by_id = request.POST.get('processed_by')
        processed_by = User.objects.get(pk=processed_by_id) if processed_by_id else None

        beneficiary_id = request.POST.get('beneficiary')
        beneficiary = User.objects.get(pk=beneficiary_id) if beneficiary_id else None

        budget = Budget.objects.get(pk=budget_id) if budget_id else None

        proposal = Proposal(
            title=title,
            description=description,
            submitted_by=request.user,
            processed_by=processed_by,
            proposed_amount=Decimal(proposed_amount),
            budget=budget,
            beneficiary=beneficiary,
            document=document            
        )

        try:
            proposal.full_clean()
            proposal.save()

            # -------------------------
            # Audit Log
            # -------------------------
            AuditLog.objects.create(
                user=request.user,
                action="create",
                model_name="Proposal",
                object_id=str(proposal.pk),
                old_data=None,
                new_data={
                    "title": proposal.title,
                    "description": proposal.description,
                    "proposed_amount": str(proposal.proposed_amount),
                    "budget": budget.fund_source if budget else None,
                    "processed_by": processed_by.get_full_name() if processed_by else None,
                    "beneficiary": beneficiary.get_full_name() if beneficiary else None,
                    "document": proposal.document.url if proposal.document else None
                }
            )
            # -------------------------

            # -------------------------
            # Notifications
            # 1. Notify all admins
            admins = User.objects.filter(role='admin')
            admin_link = reverse('administrator_proposals_url')
            for admin in admins:
                Notification.objects.create(
                    sender=request.user,
                    receiver=admin,
                    message=f"New proposal submitted: {proposal.title}",
                    link=admin_link
                )

            # 2. Notify the processed_by user (proponent)
            if processed_by:
                proponent_link = reverse('proponent_proposals_url')
                Notification.objects.create(
                    sender=request.user,
                    receiver=processed_by,
                    message=f"You have been assigned to process proposal: {proposal.title}",
                    link=proponent_link
                )

            # 3. Notify the beneficiary
            if beneficiary:
                beneficiary_link = reverse('beneficiary_proposals_url')  # <-- create a URL for beneficiary view
                Notification.objects.create(
                    sender=request.user,
                    receiver=beneficiary,
                    message=f"You are listed as beneficiary for proposal: {proposal.title}",
                    link=beneficiary_link
                )
            # -------------------------

            messages.success(request, 'Proposal added successfully.')

        except ValidationError as e:
            messages.error(request, e)

    return redirect('staff_proposals_url')


@login_required
def staff_proposals_update_view(request, pk):
    proposal = get_object_or_404(Proposal, pk=pk)
    
    if request.method == 'POST':
        # Capture old data for audit
        old_data = {
            "title": proposal.title,
            "description": proposal.description,
            "proposed_amount": str(proposal.proposed_amount),
            "approved_amount": str(proposal.approved_amount),
            "budget": proposal.budget.fund_source if proposal.budget else None,
            "processed_by": proposal.processed_by.get_full_name() if proposal.processed_by else None,
            "beneficiary": proposal.beneficiary.get_full_name() if proposal.beneficiary else None,
            "document": proposal.document.url if proposal.document else None,
            "status": proposal.status,
        }

        # Update fields
        proposal.title = request.POST.get('title')
        proposal.description = request.POST.get('description')
        proposal.proposed_amount = Decimal(request.POST.get('proposed_amount') or 0)

        approved_amount = request.POST.get('approved_amount')
        if approved_amount:
            proposal.approved_amount = Decimal(approved_amount)

        budget_id = request.POST.get('budget')
        proposal.budget = Budget.objects.get(pk=budget_id) if budget_id else None

        # Update processed_by
        processed_by_id = request.POST.get('processed_by')
        if processed_by_id:
            try:
                proposal.processed_by = User.objects.get(pk=processed_by_id)
            except User.DoesNotExist:
                proposal.processed_by = None
        else:
            proposal.processed_by = None

        # Update beneficiary
        beneficiary_id = request.POST.get('beneficiary')
        if beneficiary_id:
            try:
                proposal.beneficiary = User.objects.get(pk=beneficiary_id)
            except User.DoesNotExist:
                proposal.beneficiary = None
        else:
            proposal.beneficiary = None

        # Safely update status
        proposal.status = request.POST.get('status') or proposal.status

        # Update document if uploaded
        document = request.FILES.get('document')
        if document:
            proposal.document = document

        try:
            # Validate and save
            proposal.full_clean()
            proposal.save()

            # Audit log
            new_data = {
                "title": proposal.title,
                "description": proposal.description,
                "proposed_amount": str(proposal.proposed_amount),
                "approved_amount": str(proposal.approved_amount),
                "processed_by": proposal.processed_by.get_full_name() if proposal.processed_by else None,
                "beneficiary": proposal.beneficiary.get_full_name() if proposal.beneficiary else None,
                "status": proposal.status,
                "document": proposal.document.url if proposal.document else None,
            }

            AuditLog.objects.create(
                user=request.user,
                action="update",
                model_name="Proposal",
                object_id=str(proposal.pk),
                old_data=old_data,
                new_data=new_data
            )

            # -------------------------
            # Notifications
            # -------------------------
            admin_link = reverse('administrator_task_list_url')
            admins = User.objects.filter(role='admin')
            for admin in admins:
                Notification.objects.create(
                    sender=request.user,
                    receiver=admin,
                    message=f"Proposal '{proposal.title}' has been updated.",
                    link=admin_link
                )

            # Notify processed_by user
            if proposal.processed_by:
                proponent_link = reverse('proponent_task_list_url')
                Notification.objects.create(
                    sender=request.user,
                    receiver=proposal.processed_by,
                    message=f"You have been assigned to process proposal: {proposal.title}",
                    link=proponent_link
                )

            # Notify beneficiary
            if proposal.beneficiary:
                beneficiary_link = reverse('beneficiary_proposals_url')
                Notification.objects.create(
                    sender=request.user,
                    receiver=proposal.beneficiary,
                    message=f"You are listed as beneficiary for proposal: {proposal.title}",
                    link=beneficiary_link
                )

            # Notify submitter if different from processed_by
            if proposal.submitted_by != proposal.processed_by:
                Notification.objects.create(
                    sender=request.user,
                    receiver=proposal.submitted_by,
                    message=f"Your proposal '{proposal.title}' has been updated.",
                    link=proponent_link  # or a dedicated submitter link
                )

            messages.success(request, 'Proposal updated successfully.')

        except ValidationError as e:
            errors = []
            if hasattr(e, 'message_dict'):
                for field, msgs in e.message_dict.items():
                    for msg in msgs:
                        errors.append(f"{field.replace('_', ' ').capitalize()}: {msg}")
            elif hasattr(e, 'messages'):
                errors.extend(e.messages)
            messages.error(request, "Failed to update proposal. " + "; ".join(errors))

    return redirect('staff_proposals_url')


from datetime import date, datetime
def staff_projects_view(request):
    projects = Project.objects.select_related('proposal', 'budget', 'project_leader').all()
    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])
    proposals = Proposal.objects.filter(status='approved')

    # Compute display_status for template
    for project in projects:
        if project.status == 'proposal':
            project.display_status = 'proposal'
        elif project.status == 'terminated':
            project.display_status = 'terminated'
        elif project.end_date and timezone.now().date() >= project.end_date:
            project.display_status = 'completed'
        else:
            project.display_status = 'ongoing'

    context = {
        'projects': projects,
        'budgets': budgets,
        'proposals': proposals,
    }
    return render(request, 'staff/projects.html', context)

# -------------------------
# Staff: Task Views
# -------------------------
def staff_task_list_view(request):
    # Staff can see tasks assigned to them and tasks under projects they manage
    tasks = Task.objects.select_related(
        'project',
        'assigned_to',
        'project__proposal',
        'project__proposal__processed_by'
    ).filter(
        Q(assigned_to=request.user) | Q(project__project_leader=request.user)
    ).distinct()

    context = {
        'tasks': tasks,
        'status_choices': Task.STATUS_CHOICES,
    }
    return render(request, 'staff/tasks.html', context)

# -------------------------
# Create Task
# -------------------------
# -------------------------
def staff_task_create_view(request):
    if request.method == 'POST':
        description = request.POST.get('description')
        project_id = request.POST.get('project')
        due_date_str = request.POST.get('due_date')  # string from form
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        location_name = request.POST.get('location_name')

        project = Project.objects.get(id=project_id)

        # Convert due_date string to date object safely
        due_date = None
        if due_date_str:
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid due date format. Use YYYY-MM-DD.")
                return redirect('staff_task_list_url')

        # Assign the task to the logged-in user
        assigned_to = request.user

        task = Task.objects.create(
            title=f"Task for {project.project_title}",
            description=description,
            project=project,
            assigned_to=assigned_to,
            due_date=due_date,
            latitude=latitude if latitude else None,
            longitude=longitude if longitude else None,
            location_name=location_name,
        )
        # -------------------------
        # Audit Log
        # -------------------------
        new_data = {
            "title": task.title,
            "description": task.description,
            "project": task.project.project_title if task.project else None,
            "assigned_to": task.assigned_to.username if task.assigned_to else None,
            "due_date": str(task.due_date) if task.due_date else None,
            "status": task.status,
            "latitude": task.latitude,
            "longitude": task.longitude,
            "location_name": task.location_name,
        }

        AuditLog.objects.create(
            user=request.user,
            action="create",
            model_name="Task",
            object_id=str(task.pk),
            old_data=None,
            new_data=new_data
        )
        # -------------------------

        # -------------------------
        # Notifications
        # -------------------------
        admin_users = User.objects.filter(is_staff=True, is_superuser=True)
        admin_link = reverse('administrator_task_list_url')

        for admin in admin_users:
            Notification.objects.create(
                sender=request.user,
                receiver=admin,
                message=f"New task '{task.title}' has been created.",
                link=admin_link
            )

        # Notify project.proposal.processed_by safely
        proposal = getattr(task.project, 'proposal', None)
        if proposal and proposal.processed_by:
            proponent_link = reverse('proponent_task_list_url')
            Notification.objects.create(
                sender=request.user,
                receiver=proposal.processed_by,
                message=f"New task '{task.title}' has been created for your proposal.",
                link=proponent_link
            )
        # -------------------------

        messages.success(request, "Task created successfully!")
        return redirect('staff_task_list_url')

# -------------------------
# Edit Task
# -------------------------
def staff_task_edit_view(request):
    if request.method == 'POST':
        task_id = request.POST.get('id')
        task = get_object_or_404(Task, id=task_id)

        # Allow edits when task is assigned to the staff user or they lead the project.
        can_edit_task = task.assigned_to == request.user or (
            task.project and task.project.project_leader_id == request.user.id
        )
        if not can_edit_task:
            messages.error(request, "You can only edit tasks assigned to you or your managed projects.")
            return redirect('staff_task_list_url')

        # Capture old data
        old_data = {
            "status": task.status,
        }

        # Update only the status field (staff can only change status)
        old_status = task.status
        task.status = request.POST.get('status')

        # Make sure assigned_to remains the logged-in user
        task.assigned_to = request.user
        task.save()

        # -------------------------
        # Audit Log
        # -------------------------
        new_data = {
            "status": task.status,
        }

        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="Task",
            object_id=str(task.pk),
            old_data=old_data,
            new_data=new_data
        )
        # -------------------------

        # -------------------------
        # Notifications
        # -------------------------
        admin_users = User.objects.filter(is_staff=True, is_superuser=True)
        admin_link = reverse('administrator_task_list_url')

        for admin in admin_users:
            Notification.objects.create(
                sender=request.user,
                receiver=admin,
                message=f"Task '{task.title}' status has been updated to '{task.status}' by assigned staff.",
                link=admin_link
            )

        # Notify project.proposal.processed_by safely
        proposal = getattr(task.project, 'proposal', None)
        if proposal and proposal.processed_by:
            proponent_link = reverse('proponent_task_list_url')
            Notification.objects.create(
                sender=request.user,
                receiver=proposal.processed_by,
                message=f"Task '{task.title}' status has been updated to '{task.status}' for your proposal.",
                link=proponent_link
            )
        # -------------------------

        messages.success(request, "Task updated successfully!")
        return redirect('staff_task_list_url')


def mark_task_completed_view(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    # Ensure staff can only mark their own tasks as completed
    if task.assigned_to != request.user:
        messages.error(request, "You can only mark your own tasks as completed.")
        return redirect('staff_task_list_url')

    if request.method == "POST":

        # Update task
        task.status = "completed"
        task.completion_date = timezone.now().date()
        task.save()

        # NEW DATA for audit
        new_data = {
            "status": task.status,
            "completion_date": str(task.completion_date),
        }

        # Audit log entry
        AuditLog.objects.create(
            user=request.user,
            action="update",   # use "update" instead of "create"
            model_name="Task",
            object_id=str(task.pk),
            old_data=None,
            new_data=new_data
        )

        messages.success(request, f"Task '{task.title}' marked as completed.")
        return redirect(request.META.get("HTTP_REFERER", "staff_task_list_url"))

    return redirect("staff_task_list_url")

# -------------------------
# Task Dependency Management
# -------------------------
@login_required
def task_dependencies_view(request, task_id):
    """View to manage dependencies for a specific task"""
    task = get_object_or_404(Task, id=task_id)

    # Check permissions - users can only manage dependencies for tasks they have access to
    if request.user.role == 'dost_staff' and task.assigned_to != request.user:
        messages.error(request, "You can only manage dependencies for tasks assigned to you.")
        return redirect('staff_task_list_url')
    elif request.user.role == 'admin':
        pass  # Admins can manage all dependencies
    else:
        messages.error(request, "You don't have permission to manage task dependencies.")
        return redirect('staff_task_list_url')

    # Get all tasks in the same project (excluding the current task)
    project_tasks = Task.objects.filter(project=task.project).exclude(id=task.id)

    # Get current dependencies
    predecessors = task.get_predecessors()
    successors = task.get_successors()

    # Get available tasks to add as dependencies
    available_predecessors = project_tasks.exclude(
        id__in=[p.id for p in predecessors]
    ).exclude(id=task.id)

    context = {
        'task': task,
        'predecessors': predecessors,
        'successors': successors,
        'available_predecessors': available_predecessors,
        'is_blocked': task.is_blocked(),
        'blocking_tasks': task.get_blocking_tasks(),
    }

    return render(request, 'staff/task_dependencies.html', context)

@login_required
def add_task_dependency_view(request):
    """AJAX view to add a dependency between tasks"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})

    try:
        predecessor_id = request.POST.get('predecessor_id')
        successor_id = request.POST.get('successor_id')

        predecessor = get_object_or_404(Task, id=predecessor_id)
        successor = get_object_or_404(Task, id=successor_id)

        # Check permissions
        if request.user.role == 'dost_staff':
            if successor.assigned_to != request.user:
                return JsonResponse({'success': False, 'error': 'You can only manage dependencies for tasks assigned to you.'})
        elif request.user.role != 'admin':
            return JsonResponse({'success': False, 'error': 'Permission denied.'})

        # Add the dependency
        successor.add_dependency(predecessor)

        # Audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="Task",
            object_id=str(successor.pk),
            old_data={"dependencies": "none"},
            new_data={"dependency_added": f"Depends on task {predecessor.title}"}
        )

        return JsonResponse({
            'success': True,
            'message': f'Dependency added: {successor.title} now depends on {predecessor.title}',
            'is_blocked': successor.is_blocked()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def remove_task_dependency_view(request):
    """AJAX view to remove a dependency between tasks"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})

    try:
        predecessor_id = request.POST.get('predecessor_id')
        successor_id = request.POST.get('successor_id')

        predecessor = get_object_or_404(Task, id=predecessor_id)
        successor = get_object_or_404(Task, id=successor_id)

        # Check permissions
        if request.user.role == 'dost_staff':
            if successor.assigned_to != request.user:
                return JsonResponse({'success': False, 'error': 'You can only manage dependencies for tasks assigned to you.'})
        elif request.user.role != 'admin':
            return JsonResponse({'success': False, 'error': 'Permission denied.'})

        # Remove the dependency
        successor.remove_dependency(predecessor)

        # Audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="Task",
            object_id=str(successor.pk),
            old_data={"dependency_removed": f"Removed dependency on task {predecessor.title}"},
            new_data={"dependencies": "updated"}
        )

        return JsonResponse({
            'success': True,
            'message': f'Dependency removed: {successor.title} no longer depends on {predecessor.title}',
            'is_blocked': successor.is_blocked()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def task_dependency_graph_view(request, project_id):
    """View to display dependency graph for a project"""
    project = get_object_or_404(Project, id=project_id)

    # Check permissions
    if request.user.role not in ['admin', 'dost_staff']:
        messages.error(request, "You don't have permission to view task dependencies.")
        return redirect('staff_task_list_url')

    if request.user.role == 'dost_staff':
        # Staff can only see projects they have tasks in
        if not Task.objects.filter(project=project, assigned_to=request.user).exists():
            messages.error(request, "You don't have access to this project's tasks.")
            return redirect('staff_task_list_url')

    tasks = Task.objects.filter(project=project).prefetch_related('predecessor_dependencies', 'successor_dependencies')

    # Build dependency data for visualization
    nodes = []
    edges = []

    for task in tasks:
        nodes.append({
            'id': task.id,
            'label': task.title[:30] + '...' if len(task.title) > 30 else task.title,
            'status': task.status,
            'is_blocked': task.is_blocked(),
            'assigned_to': task.assigned_to.username if task.assigned_to else 'Unassigned'
        })

        # Add edges for dependencies
        for dep in task.predecessor_dependencies.all():
            edges.append({
                'from': dep.predecessor.id,
                'to': dep.successor.id,
                'label': 'depends on'
            })

    context = {
        'project': project,
        'tasks': tasks,
        'nodes': json.dumps(nodes),
        'edges': json.dumps(edges),
    }

# staff Reports

@login_required
def staff_reports_view(request):
    from decimal import Decimal
    from django.db.models import Sum, Count, Q
    import json
    from datetime import datetime
    from django.utils import timezone

    # Get filter parameters
    selected_year = request.GET.get('year')
    selected_municipality = request.GET.get('municipality')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_status = request.GET.get('status')
    
    # Don't set default year - only filter if user explicitly selects a year
    current_year = datetime.now().year

    # Base querysets for filtering
    project_queryset = Project.objects.all()
    proposal_queryset = Proposal.objects.all()
    budget_queryset = Budget.objects.all()

    # Apply filters
    if selected_year:
        try:
            year_int = int(selected_year)
            project_queryset = project_queryset.filter(year=year_int)
            proposal_queryset = proposal_queryset.filter(submission_date__year=year_int)
            budget_queryset = budget_queryset.filter(fiscal_year=year_int)
        except ValueError:
            pass

    if selected_municipality and selected_municipality != '':
        project_queryset = project_queryset.filter(mun__iexact=selected_municipality)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__gte=start) | 
                Q(date_of_completion__gte=start) |
                Q(approval_date__date__gte=start)
            )
            proposal_queryset = proposal_queryset.filter(submission_date__date__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__lte=end) | 
                Q(date_of_completion__lte=end) |
                Q(approval_date__date__lte=end)
            )
            proposal_queryset = proposal_queryset.filter(submission_date__date__lte=end)
        except ValueError:
            pass

    if selected_status and selected_status != '':
        project_queryset = project_queryset.filter(status__iexact=selected_status)

    # -----------------------------
    # Total Budget
    # -----------------------------
    total_budget = budget_queryset.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_spent = BudgetTransaction.objects.filter(transaction_type='deduction').aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    total_remaining = total_budget - total_spent

    total_budget_int = int(total_budget)
    total_spent_int = int(total_spent)
    total_remaining_int = int(total_remaining)

    # -----------------------------
    # Proposal Status Counts
    # -----------------------------
    proposal_counts_qs = proposal_queryset.values('status').annotate(total=Count('id'))
    proposal_status_counts = {'pending': 0, 'approved': 0, 'rejected': 0}
    for p in proposal_counts_qs:
        if p['status'] in proposal_status_counts:
            proposal_status_counts[p['status']] = p['total']

    # -----------------------------
    # Project Approved Amounts
    # -----------------------------
    projects = project_queryset.all()
    scatter_labels = [p.project_title for p in projects]
    scatter_data = [int(p.funds or 0) for p in projects]

    # -----------------------------
    # Tasks per Project (Pending & Completed)
    # -----------------------------
    step_labels = []
    pending_counts = []
    completed_counts = []

    for p in projects:
        step_labels.append(f"{p.project_title} ({p.project_start.year if p.project_start else 'N/A'})")
        pending_counts.append(p.tasks.filter(status='pending').count())
        completed_counts.append(p.tasks.filter(status='completed').count())

    # Get filter options for the form
    available_years = sorted(list(Project.objects.values_list('year', flat=True).distinct().exclude(year__isnull=True)))
    available_municipalities = sorted(list(Project.objects.values_list('mun', flat=True).distinct().exclude(mun__isnull=True).exclude(mun='')))
    available_statuses = sorted(list(Project.objects.values_list('status', flat=True).distinct().exclude(status__isnull=True).exclude(status='')))

    context = {
        'total_budget': total_budget_int,
        'total_spent': total_spent_int,
        'total_remaining': total_remaining_int,
        'proposal_status_counts': json.dumps(proposal_status_counts),
        'scatter_labels': json.dumps(scatter_labels),
        'scatter_data': json.dumps(scatter_data),
        'step_labels': json.dumps(step_labels),
        'pending_counts': json.dumps(pending_counts),
        'completed_counts': json.dumps(completed_counts),
        
        # Filters
        'available_years': available_years,
        'available_municipalities': available_municipalities,
        'available_statuses': available_statuses,
        'selected_year': selected_year,
        'selected_municipality': selected_municipality,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': selected_status,
        
        'report_year': int(selected_year) if selected_year else 2025,
    }

    return render(request, 'staff/reports.html', context)



@login_required
def staff_settings_view(request):
    user = request.user  # currently logged-in user

    if request.method == "POST":
        # Capture old data
        old_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Update user fields
        user.first_name = request.POST.get('first_name')
        user.middle_name = request.POST.get('middle_name')
        user.last_name = request.POST.get('last_name')
        user.suffix = request.POST.get('suffix')
        user.sex = request.POST.get('sex')
        user.civil_status = request.POST.get('civil_status')
        user.contact_number = request.POST.get('contact_number')
        user.email = request.POST.get('email')
        user.address = request.POST.get('address')

        if not apply_profile_picture_update(request, user):
            return redirect('staff_settings_url')

        user.save()

        # Capture new data
        new_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        messages.success(request, "Your profile has been updated successfully!")
        return redirect('staff_settings_url')

    return render(request, 'staff/settings.html', {'user': user})


@login_required
def staff_change_password_view(request):
    if request.method == "POST":
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        user = request.user

        # Check if old password matches
        if not user.check_password(old_password):
            messages.error(request, "Your old password is incorrect.")
            return redirect('staff_settings_url')

        # Validate new passwords
        if new_password1 != new_password2:
            messages.error(request, "New passwords do not match.")
            return redirect('staff_settings_url')

        # Capture old data (just note that password was changed)
        old_data = {"password": "********"}
        new_data = {"password": "********"}

        # Set and save new password
        user.set_password(new_password1)
        user.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        # Log the user out after successful password change
        logout(request)
        messages.success(request, "Your password has been changed. Please log in again.")
        return redirect('logout_url')

    return redirect('staff_settings_url')




# -------------------------
# View Audit Logs
# -------------------------
@login_required
def staff_audit_logs_view(request):
    # Only show logs that belong to the logged-in user
    logs = AuditLog.objects.filter(user=request.user).order_by('-timestamp')

    context = {
        'logs': logs,
    }
    return render(request, 'staff/audit_logs.html', context)

































# proponent
@login_required
def proponent_dashboard_view(request):
    user = request.user

    # Get proposals submitted by this proponent (processed_by is the proponent)
    proposals_qs = Proposal.objects.filter(proponent=user)
    
    # Get projects from the proponent's proposals
    projects_qs = Project.objects.filter(proposal__proponent=user)
    
    # Get tasks assigned to the logged-in proponent
    tasks_qs = Task.objects.filter(assigned_to=user)
    
    # Proposal status counts
    pending_proposals = proposals_qs.filter(status='pending').count()
    for_review_proposals = proposals_qs.filter(status='for_review').count()
    approved_proposals = proposals_qs.filter(status='approved').count()
    rejected_proposals = proposals_qs.filter(status='rejected').count()
    needs_revision_proposals = proposals_qs.filter(status='needs_revision').count()
    
    # Project status counts
    ongoing_projects = 0
    completed_projects = 0
    terminated_projects = 0
    for project in projects_qs:
        if project.status == 'terminated':
            terminated_projects += 1
        elif project.end_date and timezone.now().date() >= project.end_date:
            completed_projects += 1
        else:
            ongoing_projects += 1
    
    # DOST Compliance: Liquidation Progress
    from decimal import Decimal
    total_funds_released = Decimal('0.00')
    total_liquidated = Decimal('0.00')
    for p in projects_qs:
        total_funds_released += p.total_funds_released or Decimal('0.00')
        total_liquidated += p.amount_liquidated or Decimal('0.00')
    
    liquidation_rate = round((float(total_liquidated) / float(total_funds_released) * 100), 2) if total_funds_released > 0 else 0
    remaining_to_liquidate = float(total_funds_released - total_liquidated)
    
    # DOST Compliance: Tranche Summary
    tranche_data = []
    for p in projects_qs:
        project_tranches = TrancheRelease.objects.filter(project=p)
        for tr in project_tranches:
            tranche_data.append({
                'project_title': p.project_title[:30] + '...' if len(p.project_title or '') > 30 else (p.project_title or 'Untitled'),
                'tranche_number': tr.tranche_number,
                'amount': float(tr.amount or 0),
                'status': tr.status,
                'status_display': tr.get_status_display(),
                'liquidated_amount': float(tr.liquidated_amount or 0),
                'liquidation_percentage': tr.liquidation_percentage if hasattr(tr, 'liquidation_percentage') else 0,
                'release_date': tr.release_date.isoformat() if tr.release_date else '',
            })
    
    # DOST Compliance: Equipment Accountability
    equipment_list = []
    for p in projects_qs:
        for eq in p.equipment_deliveries.select_related('budget_allocation__equipment_item').all():
            equipment_list.append({
                'project_title': p.project_title[:25] + '...' if len(p.project_title or '') > 25 else (p.project_title or 'Untitled'),
                'name': eq.budget_allocation.equipment_item.name if eq.budget_allocation and eq.budget_allocation.equipment_item else 'Unknown',
                'quantity': eq.delivered_quantity,
                'property_tag': eq.property_tag_number or 'Not Assigned',
                'ownership_status': eq.ownership_status or 'dost_owned',
                'ownership_status_display': eq.get_ownership_status_display() if hasattr(eq, 'get_ownership_status_display') else eq.ownership_status,
                'lease_start_date': eq.lease_start_date.isoformat() if eq.lease_start_date else '',
                'ownership_end_date': eq.ownership_end_date.isoformat() if eq.ownership_end_date else '',
                'ownership_progress': eq.ownership_progress_percentage or 0,
                'days_remaining': eq.days_until_ownership if eq.days_until_ownership else 0,
                'is_eligible_for_transfer': eq.is_eligible_for_transfer if hasattr(eq, 'is_eligible_for_transfer') else False,
            })
    
    total_equipment = len(equipment_list)
    equipment_with_tags = len([eq for eq in equipment_list if eq['property_tag'] != 'Not Assigned'])
    equipment_eligible_transfer = len([eq for eq in equipment_list if eq['is_eligible_for_transfer']])
    
    # Build projects JSON for map
    projects_list = []
    for p in projects_qs:
        projects_list.append({
            'id': p.id,
            'title': p.project_title or '',
            'status': p.status or '',
            'latitude': float(p.latitude) if p.latitude is not None else None,
            'longitude': float(p.longitude) if p.longitude is not None else None,
        })

    proposals_list = []
    for prop in proposals_qs:
        proposals_list.append({
            'title': prop.title or '',
            'municipality': prop.municipality or '',
            'province': prop.province or '',
            'status': prop.status or '',
            'submitted_by': prop.submitted_by.full_name() if prop.submitted_by else '',
            'proposed_amount': float(prop.proposed_amount) if prop.proposed_amount else 0,
            'submission_date': prop.submission_date.isoformat() if prop.submission_date else '',
        })
    
    # Get my_projects for the template
    my_projects = projects_qs.order_by('-project_start')
    
    context = {
        'my_proposals_count': proposals_qs.count(),
        'my_projects_count': projects_qs.count(),
        'my_tasks_count': tasks_qs.count(),
        'pending_proposals': pending_proposals,
        'for_review_proposals': for_review_proposals,
        'approved_proposals': approved_proposals,
        'rejected_proposals': rejected_proposals,
        'needs_revision_proposals': needs_revision_proposals,
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'terminated_projects': terminated_projects,
        'proposals': proposals_qs.order_by('-submission_date')[:5],
        'projects': projects_qs.order_by('-project_start')[:5],
        'projects_json': json.dumps(projects_list),
        'proposals_json': json.dumps(proposals_list),
        'my_projects': my_projects,
        
        # DOST Compliance Data
        'total_funds_released': float(total_funds_released),
        'total_liquidated': float(total_liquidated),
        'liquidation_rate': liquidation_rate,
        'remaining_to_liquidate': remaining_to_liquidate,
        'tranche_data': tranche_data,
        'tranche_data_json': json.dumps(tranche_data),
        'equipment_list': equipment_list,
        'equipment_list_json': json.dumps(equipment_list),
        'total_equipment': total_equipment,
        'equipment_with_tags': equipment_with_tags,
        'equipment_eligible_transfer': equipment_eligible_transfer,
    }
    return render(request, 'proponent/dashboard.html', context)

# ----------------------------
# proponent User Management Views
# ----------------------------


def proponent_users_view(request):
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'proponent/users.html', {'users': users})


# --------------------------
# proponent: Budget List
# --------------------------
def proponent_budgets_view(request):
    budgets = Budget.objects.all().order_by('-date_created')
    return render(request, 'proponent/budgets.html', {'budgets': budgets})

# -------------------------
# List Proposals
# -------------------------
def proponent_proposals_view(request):
    # Only proposals submitted by the logged-in user
    proposals = Proposal.objects.select_related('submitted_by', 'processed_by', 'budget', 'proponent') \
                                .filter(proponent=request.user)
    
    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])

    # Fetch all proponents for the "Add Proposal" form
    proponents = User.objects.filter(role='proponent')

    context = {
        'proposals': proposals,
        'budgets': budgets,
        'proponents': proponents,
    }
    return render(request, 'proponent/proposals.html', context)




def proponent_proposals_add_view(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        proposed_amount = request.POST.get('proposed_amount') or 0
        budget_id = request.POST.get('budget')
        document = request.FILES.get('document')

        processed_by_id = request.POST.get('processed_by')
        processed_by = User.objects.get(pk=processed_by_id) if processed_by_id else None

        budget = Budget.objects.get(pk=budget_id) if budget_id else None

        proposal = Proposal(
            title=title,
            description=description,
            submitted_by=request.user,
            processed_by=processed_by,
            proposed_amount=Decimal(proposed_amount),
            budget=budget,
            document=document,
            proponent=request.user
        )

        try:
            proposal.full_clean()
            proposal.save()

            # -------------------------
            # Audit Log
            # -------------------------
            from .models import AuditLog

            AuditLog.objects.create(
                user=request.user,
                action="create",
                model_name="Proposal",
                object_id=str(proposal.pk),
                old_data=None,
                new_data={
                    "title": proposal.title,
                    "description": proposal.description,
                    "proposed_amount": str(proposal.proposed_amount),
                    "budget": budget.fund_source if budget else None,
                    "processed_by": processed_by.get_full_name() if processed_by else None,
                    "document": proposal.document.url if proposal.document else None
                }
            )
            # -------------------------

            # 🔔 NOTIFICATIONS
            # 1. Notify all admins
            admins = User.objects.filter(role='admin')
            for admin in admins:
                Notification.objects.create(
                    sender=request.user,
                    receiver=admin,
                    message=f"New proposal submitted: {proposal.title}",
                    link=f"/administrator/proposals/{proposal.id}/"
                )

            # 2. Notify the processed_by user
            if processed_by:
                Notification.objects.create(
                    sender=request.user,
                    receiver=processed_by,
                    message=f"You have been assigned to process proposal: {proposal.title}",
                    link=f"/administrator/proposals/{proposal.id}/"
                )

            messages.success(request, 'Proposal added successfully.')

        except ValidationError as e:
            messages.error(request, e)

    return redirect('proponent_proposals_url')

# -------------------------
# Update Proposal
# -------------------------
@login_required
def proponent_proposals_update_view(request, pk):
    proposal = get_object_or_404(Proposal, pk=pk)
    
    if request.method == 'POST':
        # Capture old data for audit
        old_data = {
            "title": proposal.title,
            "description": proposal.description,
            "proposed_amount": str(proposal.proposed_amount),
            "approved_amount": str(proposal.approved_amount),
            "budget": proposal.budget.fund_source if proposal.budget else None,
            "processed_by": proposal.processed_by.get_full_name() if proposal.processed_by else None,
            "document": proposal.document.url if proposal.document else None,
            "status": proposal.status,
        }

        # Update fields
        proposal.title = request.POST.get('title')
        proposal.description = request.POST.get('description')
        proposal.proposed_amount = Decimal(request.POST.get('proposed_amount') or 0)

        approved_amount = request.POST.get('approved_amount')
        if approved_amount:
            proposal.approved_amount = Decimal(approved_amount)

        budget_id = request.POST.get('budget')
        proposal.budget = Budget.objects.get(pk=budget_id) if budget_id else None

        # Update processed_by
        processed_by_id = request.POST.get('processed_by')
        if processed_by_id:
            try:
                proposal.processed_by = User.objects.get(pk=processed_by_id)
            except User.DoesNotExist:
                proposal.processed_by = None
        else:
            proposal.processed_by = None

        # Safely update status
        proposal.status = request.POST.get('status') or proposal.status

        # Update document if uploaded
        document = request.FILES.get('document')
        if document:
            proposal.document = document

        try:
            # Validate and save
            proposal.full_clean()
            proposal.save()

            # Audit log
            new_data = {
                "title": proposal.title,
                "description": proposal.description,
                "proposed_amount": str(proposal.proposed_amount),
                "approved_amount": str(proposal.approved_amount),
                "budget": proposal.budget.fund_source if proposal.budget else None,
                "status": proposal.status,
                "processed_by": proposal.processed_by.get_full_name() if proposal.processed_by else None,
                "document": proposal.document.url if proposal.document else None,
            }

            AuditLog.objects.create(
                user=request.user,
                action="update",
                model_name="Proposal",
                object_id=str(proposal.pk),
                old_data=old_data,
                new_data=new_data
            )

            # -------------------------
            # Notifications to submitter & processed_by
            # -------------------------
            receivers = [proposal.submitted_by]
            if proposal.processed_by and proposal.processed_by != proposal.submitted_by:
                receivers.append(proposal.processed_by)

            for receiver in receivers:
                Notification.objects.create(
                    sender=request.user,
                    receiver=receiver,
                    message=f"Proposal '{proposal.title}' has been updated.",
                    link=f"/proponent/proposals/"  # Change link as needed
                )

            messages.success(request, 'Proposal updated successfully.')
        except ValidationError as e:
            # Collect errors
            errors = []
            if hasattr(e, 'message_dict'):
                for field, msgs in e.message_dict.items():
                    for msg in msgs:
                        errors.append(f"{field.replace('_', ' ').capitalize()}: {msg}")
            elif hasattr(e, 'messages'):
                errors.extend(e.messages)
            messages.error(request, "Failed to update proposal. " + "; ".join(errors))

    return redirect('proponent_proposals_url')


from datetime import date

def proponent_projects_view(request):
    # Only projects where the logged-in user is the proponent
    projects = Project.objects.select_related('proposal', 'budget', 'project_leader') \
                              .filter(proposal__proponent=request.user)

    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])
    proposals = Proposal.objects.filter(status='approved')

    # Compute display_status for template
    for project in projects:
        if project.status == 'proposal':
            project.display_status = 'proposal'
        elif project.status == 'terminated':
            project.display_status = 'terminated'
        elif project.end_date and timezone.now().date() >= project.end_date:
            project.display_status = 'completed'
        else:
            project.display_status = 'ongoing'

    context = {
        'projects': projects,
        'budgets': budgets,
        'proposals': proposals,
    }
    return render(request, 'proponent/projects.html', context)


# -------------------------
# proponent: Task Views
# -------------------------

@login_required
def proponent_task_list_view(request):
    # Filter tasks where the task's project is linked to a proposal
    # and the proposal's processed_by is the current logged-in user
    tasks = Task.objects.select_related('project', 'assigned_to', 'project__proposal') \
        .filter(project__proposal__processed_by=request.user)
    
    projects = Project.objects.filter(proposal__processed_by=request.user)
    users = User.objects.all()

    context = {
        'tasks': tasks,
        'projects': projects,
        'users': users,
        'status_choices': Task.STATUS_CHOICES,
    }
    return render(request, 'proponent/tasks.html', context)

# proponent Reports
@login_required
def proponent_reports_view(request):
    from datetime import datetime
    from django.utils import timezone

    # Get filter parameters
    selected_year = request.GET.get('year')
    selected_municipality = request.GET.get('municipality')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_status = request.GET.get('status')
    
    # Don't set default year - only filter if user explicitly selects a year
    current_year = datetime.now().year

    # Initialize geocoder
    geolocator = Nominatim(user_agent="admin_reports")

    # Base querysets for filtering
    budget_queryset = Budget.objects.all()
    project_queryset = Project.objects.all()

    # Apply filters
    if selected_year:
        try:
            year_int = int(selected_year)
            budget_queryset = budget_queryset.filter(fiscal_year=year_int)
            project_queryset = project_queryset.filter(year=year_int)
        except ValueError:
            pass

    if selected_municipality and selected_municipality != '':
        project_queryset = project_queryset.filter(mun__iexact=selected_municipality)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__gte=start) | 
                Q(date_of_completion__gte=start) |
                Q(approval_date__date__gte=start)
            )
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__lte=end) | 
                Q(date_of_completion__lte=end) |
                Q(approval_date__date__lte=end)
            )
        except ValueError:
            pass

    if selected_status and selected_status != '':
        project_queryset = project_queryset.filter(status__iexact=selected_status)

    # Fetch budgets ordered by fiscal year, filtered
    budgets = budget_queryset.order_by('fiscal_year')

    report_data = []
    fiscal_years = []
    total_amounts = []
    remaining_amounts = []
    spent_amounts = []
    project_task_grouped = []

    for budget in budgets:
        # Filter projects for this budget based on our filters
        projects = budget.projects.filter(id__in=project_queryset.values_list('id', flat=True))
        total_projects = projects.count()
        tasks_count = Task.objects.filter(project__in=projects).count()

        # Count projects by status
        status_counts = {
            'proposal': projects.filter(status='proposal').count(),
            'ongoing': projects.filter(status='ongoing').count(),
            'completed': projects.filter(status='completed').count(),
            'terminated': projects.filter(status='terminated').count(),
        }

        # Prepare report row
        report_data.append({
            'fiscal_year': budget.fiscal_year,
            'total_budget': float(budget.total_amount),
            'spent_budget': float(budget.total_amount - budget.remaining_amount),
            'remaining_budget': float(budget.remaining_amount),
            'total_projects': total_projects,
            **status_counts,
            'total_tasks': tasks_count,
        })

        # For Chart.js
        fiscal_years.append(budget.fiscal_year)
        total_amounts.append(float(budget.total_amount))
        remaining_amounts.append(float(budget.remaining_amount))
        spent_amounts.append(float(budget.total_amount - budget.remaining_amount))

        # Grouped Project-Task data
        for project in projects:
            task_list = []
            for task in project.tasks.all():
                location_name = "N/A"
                if task.latitude and task.longitude:
                    try:
                        location = geolocator.reverse((task.latitude, task.longitude), exactly_one=True)
                        if location:
                            location_name = location.address
                    except Exception as e:
                        location_name = "N/A"

                task_list.append({
                    'task_title': task.title,
                    'location_name': location_name,
                    'task_status': task.status,
                    'assigned_to': task.assigned_to.get_full_name() if task.assigned_to else 'Unassigned',
                })

            project_task_grouped.append({
                'project_title': project.project_title,
                'approved_budget': float(project.funds or 0),
                'tasks': task_list
            })

    # Get filter options for the form
    available_years = sorted(list(Project.objects.values_list('year', flat=True).distinct().exclude(year__isnull=True)))
    available_municipalities = sorted(list(Project.objects.values_list('mun', flat=True).distinct().exclude(mun__isnull=True).exclude(mun='')))
    available_statuses = sorted(list(Project.objects.values_list('status', flat=True).distinct().exclude(status__isnull=True).exclude(status='')))

    context = {
        'report_data': report_data,
        'fiscal_years': fiscal_years,
        'total_amounts': total_amounts,
        'remaining_amounts': remaining_amounts,
        'spent_amounts': spent_amounts,
        'project_task_grouped': project_task_grouped,
        
        # Filters
        'available_years': available_years,
        'available_municipalities': available_municipalities,
        'available_statuses': available_statuses,
        'selected_year': selected_year,
        'selected_municipality': selected_municipality,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': selected_status,
    }

    return render(request, 'proponent/reports.html', context)


@login_required
def proponent_settings_view(request):
    user = request.user  # currently logged-in user

    if request.method == "POST":
        # Capture old data
        old_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Update user fields
        user.first_name = request.POST.get('first_name')
        user.middle_name = request.POST.get('middle_name')
        user.last_name = request.POST.get('last_name')
        user.suffix = request.POST.get('suffix')
        user.sex = request.POST.get('sex')
        user.civil_status = request.POST.get('civil_status')
        user.contact_number = request.POST.get('contact_number')
        user.email = request.POST.get('email')
        user.address = request.POST.get('address')

        if not apply_profile_picture_update(request, user):
            return redirect('beneficiary_settings_url')

        user.save()

        # Capture new data
        new_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        messages.success(request, "Your profile has been updated successfully!")
        return redirect('proponent_settings_url')

    return render(request, 'proponent/settings.html', {'user': user})


@login_required
def proponent_change_password_view(request):
    if request.method == "POST":
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        user = request.user

        # Check if old password matches
        if not user.check_password(old_password):
            messages.error(request, "Your old password is incorrect.")
            return redirect('proponent_settings_url')

        # Validate new passwords
        if new_password1 != new_password2:
            messages.error(request, "New passwords do not match.")
            return redirect('proponent_settings_url')

        # Capture old data (just note that password was changed)
        old_data = {"password": "********"}
        new_data = {"password": "********"}

        # Set and save new password
        user.set_password(new_password1)
        user.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        # Log the user out after successful password change
        logout(request)
        messages.success(request, "Your password has been changed. Please log in again.")
        return redirect('logout_url')

    return redirect('proponent_settings_url')


# -------------------------
# View Audit Logs
# -------------------------
@login_required
def proponent_audit_logs_view(request):
    # Only show logs that belong to the logged-in user
    logs = AuditLog.objects.filter(user=request.user).order_by('-timestamp')

    context = {
        'logs': logs,
    }
    return render(request, 'proponent/audit_logs.html', context)































# proponent Extension Requests
@login_required
@login_required
def proponent_extension_requests_view(request):
    """List extension requests submitted by the logged-in proponent"""
    extension_requests = ExtensionRequest.objects.select_related('proposal', 'proposal__project') \
                                                .filter(proponent=request.user) \
                                                .order_by('-date_submitted')
    
    context = {
        'extension_requests': extension_requests,
    }
    return render(request, 'proponent/extension-requests.html', context)


@login_required
def proponent_extension_requests_add_view(request):
    """Add a new extension request for a proponent's project/proposal"""
    if request.method == 'POST':
        proposal_id = request.POST.get('proposal')
        reason = request.POST.get('reason')
        requested_extension_days = request.POST.get('requested_extension_days')
        attachment = request.FILES.get('attachment')
        
        proposal = get_object_or_404(Proposal, pk=proposal_id, proponent=request.user)
        
        extension_request = ExtensionRequest(
            proposal=proposal,
            proponent=request.user,
            reason=reason,
            requested_extension_days=requested_extension_days,
            letter=attachment,
            status='pending'
        )
        extension_request.save()
        
        # Create notification for administrators
        from django.urls import reverse
        admin_url = reverse('administrator_extension_requests_url')
        admin_users = User.objects.filter(role='administrator')
        for admin in admin_users:
            Notification.objects.create(
                sender=request.user,
                receiver=admin,
                message=f'New extension request submitted by {request.user.get_full_name()} for proposal "{proposal.title}"',
                link=admin_url
            )
        
        messages.success(request, 'Extension request submitted successfully!')
        return redirect('proponent_extension_requests_url')
    
    # Get proposals/projects that can have extension requests
    # Only proposals that have been approved and converted to projects, or ongoing projects
    proposals = Proposal.objects.filter(
        proponent=request.user,
        status__in=['approved', 'ongoing']
    ).select_related('project')
    
    context = {
        'proposals': proposals,
    }
    return render(request, 'proponent/extension-requests-add.html', context)


# beneficiary
@login_required
def beneficiary_dashboard_view(request):
    user = request.user

    # Get proposals where the logged-in user is the beneficiary
    proposals_qs = Proposal.objects.filter(beneficiary=user)
    
    # Get projects where the logged-in user is the beneficiary
    projects_qs = Project.objects.filter(proposal__beneficiary=user)
    
    # Get tasks assigned to the logged-in beneficiary
    tasks_qs = Task.objects.filter(assigned_to=user)
    
    # Proposal status counts
    pending_proposals = proposals_qs.filter(status='pending').count()
    for_review_proposals = proposals_qs.filter(status='for_review').count()
    approved_proposals = proposals_qs.filter(status='approved').count()
    rejected_proposals = proposals_qs.filter(status='rejected').count()
    needs_revision_proposals = proposals_qs.filter(status='needs_revision').count()
    
    # Project status counts
    ongoing_projects = 0
    completed_projects = 0
    terminated_projects = 0
    proposal_projects = 0
    for project in projects_qs:
        if project.status == 'proposal':
            proposal_projects += 1
        elif project.status == 'terminated':
            terminated_projects += 1
        elif project.end_date and timezone.now().date() >= project.end_date:
            completed_projects += 1
        else:
            ongoing_projects += 1
    
    # DOST Compliance: TNA Status for the beneficiary
    tna_status = user.tna_status or 'not_started'
    tna_status_display = dict(User.TNA_STATUS_CHOICES).get(tna_status, 'Not Started') if hasattr(User, 'TNA_STATUS_CHOICES') else tna_status
    tna_completion_date = user.tna_completion_date
    tna_notes = user.tna_notes or ''
    
    # Calculate TNA progress
    tna_progress = 0
    if tna_status == 'completed':
        tna_progress = 100
    elif tna_status == 'in_progress':
        tna_progress = 50
    elif tna_status == 'expired':
        tna_progress = 100  # Show as full but with different color
    
    # DOST Compliance: Equipment assigned to beneficiary's projects with ownership timeline
    equipment_list = []
    for p in projects_qs:
        for eq in p.equipment_deliveries.select_related('budget_allocation__equipment_item').all():
            equipment_list.append({
                'project_title': p.project_title[:25] + '...' if len(p.project_title or '') > 25 else (p.project_title or 'Untitled'),
                'name': eq.budget_allocation.equipment_item.name if eq.budget_allocation and eq.budget_allocation.equipment_item else 'Unknown',
                'quantity': eq.delivered_quantity,
                'property_tag': eq.property_tag_number or 'Not Assigned',
                'ownership_status': eq.ownership_status or 'dost_owned',
                'ownership_status_display': eq.get_ownership_status_display() if hasattr(eq, 'get_ownership_status_display') else eq.ownership_status,
                'lease_start_date': eq.lease_start_date.strftime('%b %d, %Y') if eq.lease_start_date else 'Not Set',
                'ownership_end_date': eq.ownership_end_date.strftime('%b %d, %Y') if eq.ownership_end_date else 'Not Set',
                'ownership_progress': eq.ownership_progress_percentage or 0,
                'days_remaining': eq.days_until_ownership if eq.days_until_ownership else 0,
                'is_eligible_for_transfer': eq.is_eligible_for_transfer if hasattr(eq, 'is_eligible_for_transfer') else False,
            })
    
    total_equipment = len(equipment_list)
    equipment_dost_owned = len([eq for eq in equipment_list if eq['ownership_status'] == 'dost_owned'])
    equipment_transferred = len([eq for eq in equipment_list if eq['ownership_status'] == 'beneficiary_owned'])
    equipment_eligible_transfer = len([eq for eq in equipment_list if eq['is_eligible_for_transfer']])
    
    # Build projects JSON for map
    projects_list = []
    for p in projects_qs:
        projects_list.append({
            'id': p.id,
            'title': p.project_title or '',
            'status': p.status or '',
            'latitude': float(p.latitude) if p.latitude is not None else None,
            'longitude': float(p.longitude) if p.longitude is not None else None,
        })

    proposals_list = []
    for proposal in proposals_qs:
        proposals_list.append({
            'id': proposal.id,
            'title': proposal.title or '',
            'status': proposal.status or '',
            'latitude': float(proposal.latitude) if proposal.latitude is not None else None,
            'longitude': float(proposal.longitude) if proposal.longitude is not None else None,
            'submission_date': proposal.submission_date.strftime('%Y-%m-%d') if proposal.submission_date else None,
        })
    
    # Get my_projects for the template
    my_projects = projects_qs.order_by('-project_start')
    
    context = {
        'my_proposals_count': proposals_qs.count(),
        'pending_proposals': pending_proposals,
        'for_review_proposals': for_review_proposals,
        'approved_proposals': approved_proposals,
        'rejected_proposals': rejected_proposals,
        'needs_revision_proposals': needs_revision_proposals,
        'my_projects_count': projects_qs.count(),
        'my_tasks_count': tasks_qs.count(),
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'terminated_projects': terminated_projects,
        'proposal_projects': proposal_projects,
        'proposals': proposals_qs.order_by('-submission_date')[:5],
        'projects': projects_qs.order_by('-project_start')[:5],
        'projects_json': json.dumps(projects_list),
        'proposals_json': json.dumps(proposals_list),
        'my_projects': my_projects,
        
        # DOST Compliance: TNA Status
        'tna_status': tna_status,
        'tna_status_display': tna_status_display,
        'tna_completion_date': tna_completion_date,
        'tna_notes': tna_notes,
        'tna_progress': tna_progress,
        
        # DOST Compliance: Equipment Ownership
        'equipment_list': equipment_list,
        'equipment_list_json': json.dumps(equipment_list),
        'total_equipment': total_equipment,
        'equipment_dost_owned': equipment_dost_owned,
        'equipment_transferred': equipment_transferred,
        'equipment_eligible_transfer': equipment_eligible_transfer,
    }
    return render(request, 'beneficiary/dashboard.html', context)

# ----------------------------
# beneficiary User Management Views
# ----------------------------


# -------------------------
# List Proposals
# -------------------------
def beneficiary_proposals_view(request):
    # Only proposals submitted by the logged-in user
    proposals = Proposal.objects.select_related('submitted_by', 'processed_by', 'budget') \
                                .filter(beneficiary=request.user)
    
    # Get budgets that are available for allocation (includes legacy 'active' status)
    budgets = Budget.objects.filter(status__in=['available', 'partially_allocated', 'active'])

    # Fetch all beneficiarys for the "Add Proposal" form
    beneficiarys = User.objects.filter(role='beneficiary')

    context = {
        'proposals': proposals,
        'budgets': budgets,
        'beneficiarys': beneficiarys,
    }
    return render(request, 'beneficiary/proposals.html', context)



from datetime import date

def beneficiary_projects_view(request):
    # Only projects where the logged-in user is the project leader
    projects = Project.objects.select_related('proposal', 'budget', 'project_leader') \
                              .filter(proposal__beneficiary=request.user)

    budgets = Budget.objects.filter(status='active')
    proposals = Proposal.objects.filter(status='approved')

    # Compute display_status for template
    for project in projects:
        if project.status == 'proposal':
            project.display_status = 'proposal'
        elif project.status == 'terminated':
            project.display_status = 'terminated'
        elif project.end_date and date.today() >= project.end_date:
            project.display_status = 'completed'
        else:
            project.display_status = 'ongoing'

    context = {
        'projects': projects,
        'budgets': budgets,
        'proposals': proposals,
    }
    return render(request, 'beneficiary/projects.html', context)


# -------------------------
# beneficiary: Task Views
# -------------------------

@login_required
def beneficiary_task_list_view(request):
    # Filter tasks where the task's project is linked to a proposal
    # and the proposal's processed_by is the current logged-in user
    tasks = Task.objects.select_related('project', 'assigned_to', 'project__proposal') \
        .filter(project__proposal__beneficiary=request.user)
    
    projects = Project.objects.filter(proposal__beneficiary=request.user)
    users = User.objects.all()

    context = {
        'tasks': tasks,
        'projects': projects,
        'users': users,
        'status_choices': Task.STATUS_CHOICES,
    }
    return render(request, 'beneficiary/tasks.html', context)


# beneficiary Reports
@login_required
def beneficiary_reports_view(request):
    from datetime import datetime
    from django.utils import timezone

    # Get filter parameters
    selected_year = request.GET.get('year')
    selected_municipality = request.GET.get('municipality')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_status = request.GET.get('status')
    
    # Don't set default year - only filter if user explicitly selects a year
    current_year = datetime.now().year

    # Initialize geocoder
    geolocator = Nominatim(user_agent="admin_reports")

    # Base querysets for filtering
    budget_queryset = Budget.objects.all()
    project_queryset = Project.objects.all()

    # Apply filters
    if selected_year:
        try:
            year_int = int(selected_year)
            budget_queryset = budget_queryset.filter(fiscal_year=year_int)
            project_queryset = project_queryset.filter(year=year_int)
        except ValueError:
            pass

    if selected_municipality and selected_municipality != '':
        project_queryset = project_queryset.filter(mun__iexact=selected_municipality)

    if start_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__gte=start) | 
                Q(date_of_completion__gte=start) |
                Q(approval_date__date__gte=start)
            )
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            project_queryset = project_queryset.filter(
                Q(project_start__lte=end) | 
                Q(date_of_completion__lte=end) |
                Q(approval_date__date__lte=end)
            )
        except ValueError:
            pass

    if selected_status and selected_status != '':
        project_queryset = project_queryset.filter(status__iexact=selected_status)

    # Fetch budgets ordered by fiscal year, filtered
    budgets = budget_queryset.order_by('fiscal_year')

    report_data = []
    fiscal_years = []
    total_amounts = []
    remaining_amounts = []
    spent_amounts = []
    project_task_grouped = []

    for budget in budgets:
        # Filter projects for this budget based on our filters
        projects = budget.projects.filter(id__in=project_queryset.values_list('id', flat=True))
        total_projects = projects.count()
        tasks_count = Task.objects.filter(project__in=projects).count()

        # Count projects by status
        status_counts = {
            'proposal': projects.filter(status='proposal').count(),
            'ongoing': projects.filter(status='ongoing').count(),
            'completed': projects.filter(status='completed').count(),
            'terminated': projects.filter(status='terminated').count(),
        }

        # Prepare report row
        report_data.append({
            'fiscal_year': budget.fiscal_year,
            'total_budget': float(budget.total_amount),
            'spent_budget': float(budget.total_amount - budget.remaining_amount),
            'remaining_budget': float(budget.remaining_amount),
            'total_projects': total_projects,
            **status_counts,
            'total_tasks': tasks_count,
        })

        # For Chart.js
        fiscal_years.append(budget.fiscal_year)
        total_amounts.append(float(budget.total_amount))
        remaining_amounts.append(float(budget.remaining_amount))
        spent_amounts.append(float(budget.total_amount - budget.remaining_amount))

        # Grouped Project-Task data
        for project in projects:
            task_list = []
            for task in project.tasks.all():
                location_name = "N/A"
                if task.latitude and task.longitude:
                    try:
                        location = geolocator.reverse((task.latitude, task.longitude), exactly_one=True)
                        if location:
                            location_name = location.address
                    except Exception as e:
                        location_name = "N/A"

                task_list.append({
                    'task_title': task.title,
                    'location_name': location_name,
                    'task_status': task.status,
                    'assigned_to': task.assigned_to.get_full_name() if task.assigned_to else 'Unassigned',
                })

            project_task_grouped.append({
                'project_title': project.project_title,
                'approved_budget': float(project.funds or 0),
                'tasks': task_list
            })

    # Get filter options for the form
    available_years = sorted(list(Project.objects.values_list('year', flat=True).distinct().exclude(year__isnull=True)))
    available_municipalities = sorted(list(Project.objects.values_list('mun', flat=True).distinct().exclude(mun__isnull=True).exclude(mun='')))
    available_statuses = sorted(list(Project.objects.values_list('status', flat=True).distinct().exclude(status__isnull=True).exclude(status='')))

    context = {
        'report_data': report_data,
        'fiscal_years': fiscal_years,
        'total_amounts': total_amounts,
        'remaining_amounts': remaining_amounts,
        'spent_amounts': spent_amounts,
        'project_task_grouped': project_task_grouped,
        
        # Filters
        'available_years': available_years,
        'available_municipalities': available_municipalities,
        'available_statuses': available_statuses,
        'selected_year': selected_year,
        'selected_municipality': selected_municipality,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': selected_status,
    }

    return render(request, 'beneficiary/reports.html', context)


@login_required
def beneficiary_settings_view(request):
    user = request.user  # currently logged-in user

    if request.method == "POST":
        # Capture old data
        old_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Update user fields
        user.first_name = request.POST.get('first_name')
        user.middle_name = request.POST.get('middle_name')
        user.last_name = request.POST.get('last_name')
        user.suffix = request.POST.get('suffix')
        user.sex = request.POST.get('sex')
        user.civil_status = request.POST.get('civil_status')
        user.contact_number = request.POST.get('contact_number')
        user.email = request.POST.get('email')
        user.address = request.POST.get('address')

        if 'profile_picture' in request.FILES:
            user.profile_picture = request.FILES['profile_picture']

        user.save()

        # Capture new data
        new_data = {
            "first_name": user.first_name,
            "middle_name": user.middle_name,
            "last_name": user.last_name,
            "suffix": user.suffix,
            "sex": user.sex,
            "civil_status": user.civil_status,
            "contact_number": user.contact_number,
            "email": user.email,
            "address": user.address,
            "profile_picture": user.profile_picture.url if user.profile_picture else None
        }

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        messages.success(request, "Your profile has been updated successfully!")
        return redirect('beneficiary_settings_url')

    return render(request, 'beneficiary/settings.html', {'user': user})


@login_required
def beneficiary_change_password_view(request):
    if request.method == "POST":
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')

        user = request.user

        # Check if old password matches
        if not user.check_password(old_password):
            messages.error(request, "Your old password is incorrect.")
            return redirect('beneficiary_settings_url')

        # Validate new passwords
        if new_password1 != new_password2:
            messages.error(request, "New passwords do not match.")
            return redirect('beneficiary_settings_url')

        # Capture old data (just note that password was changed)
        old_data = {"password": "********"}
        new_data = {"password": "********"}

        # Set and save new password
        user.set_password(new_password1)
        user.save()

        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="User",
            object_id=user.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )

        # Log the user out after successful password change
        logout(request)
        messages.success(request, "Your password has been changed. Please log in again.")
        return redirect('logout_url')

    return redirect('beneficiary_settings_url')


# -------------------------
# View Audit Logs
# -------------------------
@login_required
def beneficiary_audit_logs_view(request):
    # Only show logs that belong to the logged-in user
    logs = AuditLog.objects.filter(user=request.user).order_by('-timestamp')

    context = {
        'logs': logs,
    }
    return render(request, 'beneficiary/audit_logs.html', context)


# =========================================
# FORMS MODULE - ADMINISTRATOR
# =========================================

@login_required
def administrator_forms_view(request):
    """List all form templates for administrator"""
    forms = FormTemplate.objects.all().order_by('-date_uploaded')
    categories = FormTemplate.CATEGORY_CHOICES
    
    context = {
        'forms': forms,
        'categories': categories,
    }
    return render(request, 'administrator/forms.html', context)


@login_required
def administrator_forms_add_view(request):
    """Add a new form template"""
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        category = request.POST.get('category', 'other')
        file = request.FILES.get('file')
        
        if not title:
            messages.error(request, "Title is required.")
            return redirect('administrator_forms_url')
        
        if not file:
            messages.error(request, "File is required.")
            return redirect('administrator_forms_url')
        
        # Check file extension
        allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx']
        file_ext = file.name.lower()[file.name.rfind('.'):]
        if file_ext not in allowed_extensions:
            messages.error(request, "Only PDF, DOC, DOCX, XLS, and XLSX files are allowed.")
            return redirect('administrator_forms_url')
        
        form_template = FormTemplate.objects.create(
            title=title,
            description=description,
            category=category,
            file=file,
            uploaded_by=request.user
        )
        
        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="create",
            model_name="FormTemplate",
            object_id=form_template.id,
            old_data=None,
            new_data=json.dumps({
                'title': title,
                'description': description,
                'category': category,
                'file': file.name
            })
        )
        
        messages.success(request, f"Form '{title}' has been uploaded successfully.")
        return redirect('administrator_forms_url')
    
    return redirect('administrator_forms_url')


@login_required
def administrator_forms_edit_view(request, form_id):
    """Edit an existing form template"""
    form_template = get_object_or_404(FormTemplate, id=form_id)
    
    if request.method == 'POST':
        old_data = {
            'title': form_template.title,
            'description': form_template.description,
            'category': form_template.category,
            'file': form_template.file.name if form_template.file else None
        }
        
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        category = request.POST.get('category', 'other')
        new_file = request.FILES.get('file')
        
        if not title:
            messages.error(request, "Title is required.")
            return redirect('administrator_forms_url')
        
        form_template.title = title
        form_template.description = description
        form_template.category = category
        
        if new_file:
            # Check file extension
            allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx']
            file_ext = new_file.name.lower()[new_file.name.rfind('.'):]
            if file_ext not in allowed_extensions:
                messages.error(request, "Only PDF, DOC, DOCX, XLS, and XLSX files are allowed.")
                return redirect('administrator_forms_url')
            form_template.file = new_file
        
        form_template.save()
        
        new_data = {
            'title': form_template.title,
            'description': form_template.description,
            'category': form_template.category,
            'file': form_template.file.name if form_template.file else None
        }
        
        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="update",
            model_name="FormTemplate",
            object_id=form_template.id,
            old_data=json.dumps(old_data),
            new_data=json.dumps(new_data)
        )
        
        messages.success(request, f"Form '{title}' has been updated successfully.")
        return redirect('administrator_forms_url')
    
    return redirect('administrator_forms_url')


@login_required
def administrator_forms_delete_view(request, form_id):
    """Delete a form template"""
    form_template = get_object_or_404(FormTemplate, id=form_id)
    
    if request.method == 'POST':
        old_data = {
            'title': form_template.title,
            'description': form_template.description,
            'category': form_template.category,
            'file': form_template.file.name if form_template.file else None
        }
        
        title = form_template.title
        form_id_for_log = form_template.id
        
        # Delete the file from storage
        if form_template.file:
            form_template.file.delete(save=False)
        
        form_template.delete()
        
        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action="delete",
            model_name="FormTemplate",
            object_id=form_id_for_log,
            old_data=json.dumps(old_data),
            new_data=None
        )
        
        messages.success(request, f"Form '{title}' has been deleted successfully.")
    
    return redirect('administrator_forms_url')


@login_required
def form_download_view(request, form_id):
    """Download a form template file"""
    form_template = get_object_or_404(FormTemplate, id=form_id)
    
    if not form_template.file:
        raise Http404("File not found.")
    
    # Increment download count
    form_template.download_count += 1
    form_template.save(update_fields=['download_count'])
    
    # Get file extension for content type
    file_name = form_template.file.name
    ext = file_name.lower().split('.')[-1]
    
    content_types = {
        'pdf': 'application/pdf',
        'doc': 'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'xls': 'application/vnd.ms-excel',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    
    content_type = content_types.get(ext, 'application/octet-stream')
    
    response = HttpResponse(form_template.file.read(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{form_template.title}.{ext}"'
    
    return response


# =========================================
# FORMS MODULE - STAFF
# =========================================

@login_required
def staff_forms_view(request):
    """List all form templates for staff to download"""
    forms = FormTemplate.objects.all().order_by('-date_uploaded')
    categories = FormTemplate.CATEGORY_CHOICES
    
    context = {
        'forms': forms,
        'categories': categories,
    }
    return render(request, 'staff/forms.html', context)


# =========================================
# FORMS MODULE - PROPONENT
# =========================================

@login_required
def proponent_forms_view(request):
    """List all form templates for proponent to download"""
    forms = FormTemplate.objects.all().order_by('-date_uploaded')
    categories = FormTemplate.CATEGORY_CHOICES
    
    context = {
        'forms': forms,
        'categories': categories,
    }
    return render(request, 'proponent/forms.html', context)


# =========================================
# FORMS MODULE - BENEFICIARY
# =========================================

@login_required
def beneficiary_forms_view(request):
    """List all form templates for beneficiary to download"""
    forms = FormTemplate.objects.all().order_by('-date_uploaded')
    categories = FormTemplate.CATEGORY_CHOICES
    
    context = {
        'forms': forms,
        'categories': categories,
    }
    return render(request, 'beneficiary/forms.html', context)

# -------------------------
# Personal Task Views (Staff Personal Checklists)
# -------------------------

@login_required
def staff_personal_tasks_view(request):
    """Staff personal task checklist view"""
    if request.user.role not in ['dost_staff', 'admin']:
        messages.error(request, "Access denied. Only staff members can access personal tasks.")
        return redirect('home')
    
    # Get user's personal tasks
    personal_tasks = PersonalTask.objects.filter(user=request.user).order_by('-created_at')
    
    # Get projects the user is assigned to for task association
    user_projects = Project.objects.filter(
        Q(project_leader=request.user) | 
        Q(tasks__assigned_to=request.user)
    ).distinct()
    
    context = {
        'personal_tasks': personal_tasks,
        'user_projects': user_projects,
        'status_choices': PersonalTask.STATUS_CHOICES,
    }
    return render(request, 'staff/personal_tasks.html', context)

@login_required
def staff_personal_task_create_view(request):
    """Create a new personal task"""
    if request.user.role not in ['dost_staff', 'admin']:
        messages.error(request, "Access denied.")
        return redirect('home')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        project_id = request.POST.get('project_id')
        priority = request.POST.get('priority', 'medium')
        due_date = request.POST.get('due_date')
        checklist_items = request.POST.getlist('checklist_items[]')
        
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            messages.error(request, "Invalid project selected.")
            return redirect('staff_personal_tasks_url')
        
        # Create checklist from form data
        checklist = []
        for item_text in checklist_items:
            if item_text.strip():  # Only add non-empty items
                checklist.append({'text': item_text.strip(), 'completed': False})
        
        if not checklist:
            messages.error(request, "Please add at least one checklist item.")
            return redirect('staff_personal_tasks_url')
        
        personal_task = PersonalTask.objects.create(
            user=request.user,
            project=project,
            title=title,
            checklist=checklist,
            priority=priority,
            due_date=due_date if due_date else None,
        )
        
        messages.success(request, "Personal task created successfully!")
        return redirect('staff_personal_tasks_url')
    
    # GET request - show form
    user_projects = Project.objects.filter(
        Q(project_leader=request.user) | 
        Q(tasks__assigned_to=request.user)
    ).distinct()
    
    context = {
        'user_projects': user_projects,
    }
    return render(request, 'staff/personal_task_form.html', context)

@login_required
def staff_personal_task_toggle_view(request, task_id):
    """Toggle personal task completion status"""
    if request.user.role not in ['dost_staff', 'admin']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        task = PersonalTask.objects.get(id=task_id, user=request.user)
        
        if task.status == 'completed':
            task.mark_pending()
            status = 'pending'
        else:
            task.mark_completed()
            status = 'completed'
        
        return JsonResponse({
            'success': True,
            'status': status,
            'completed_at': task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else None
        })
        
    except PersonalTask.DoesNotExist:
        return JsonResponse({'error': 'Task not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def staff_personal_task_edit_view(request, task_id):
    """Edit a personal task"""
    if request.user.role not in ['dost_staff', 'admin']:
        messages.error(request, "Access denied.")
        return redirect('home')
    
    try:
        task = PersonalTask.objects.get(id=task_id, user=request.user)
    except PersonalTask.DoesNotExist:
        messages.error(request, "Task not found.")
        return redirect('staff_personal_tasks_url')
    
    if request.method == 'POST':
        task.title = request.POST.get('title')
        project_id = request.POST.get('project_id')
        task.priority = request.POST.get('priority', 'medium')
        task.due_date = request.POST.get('due_date') or None
        checklist_items = request.POST.getlist('checklist_items[]')
        
        try:
            task.project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            messages.error(request, "Invalid project selected.")
            return redirect('staff_personal_tasks_url')
        
        # Update checklist from form data
        checklist = []
        for item_text in checklist_items:
            if item_text.strip():  # Only add non-empty items
                checklist.append({'text': item_text.strip(), 'completed': False})
        
        if not checklist:
            messages.error(request, "Please add at least one checklist item.")
            return redirect('staff_personal_tasks_url')
        
        task.checklist = checklist
        task.save()
        messages.success(request, "Personal task updated successfully!")
        return redirect('staff_personal_tasks_url')
    
    # GET request - show form
    user_projects = Project.objects.filter(
        Q(project_leader=request.user) | 
        Q(tasks__assigned_to=request.user)
    ).distinct()
    
    context = {
        'task': task,
        'user_projects': user_projects,
    }
    return render(request, 'staff/personal_task_form.html', context)

@login_required
def staff_personal_task_toggle_checklist_item_view(request, task_id):
    """Toggle completion status of a checklist item"""
    if request.user.role not in ['dost_staff', 'admin']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        task = PersonalTask.objects.get(id=task_id, user=request.user)
        
        # Try to get item_index from POST data (form) or JSON
        item_index = request.POST.get('item_index')
        if item_index is None:
            # Try JSON data
            import json
            data = json.loads(request.body)
            item_index = data.get('item_index')
        
        item_index = int(item_index)
        
        if 0 <= item_index < len(task.checklist):
            task.toggle_checklist_item(item_index)
            return JsonResponse({
                'success': True,
                'status': task.status,
                'completed_at': task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else None
            })
        else:
            return JsonResponse({'error': 'Invalid item index'}, status=400)
            
    except PersonalTask.DoesNotExist:
        return JsonResponse({'error': 'Task not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def staff_personal_task_delete_view(request, task_id):
    """Delete a personal task"""
    if request.user.role not in ['dost_staff', 'admin']:
        messages.error(request, "Access denied.")
        return redirect('home')
    
    try:
        task = PersonalTask.objects.get(id=task_id, user=request.user)
        task.delete()
        messages.success(request, "Personal task deleted successfully!")
    except PersonalTask.DoesNotExist:
        messages.error(request, "Task not found.")
    
    return redirect('staff_personal_tasks_url')

# ================================
# REPORT EXPORT FUNCTIONALITY
# ================================
@login_required
def export_user_productivity_pdf(request):
    """Export user productivity report as PDF"""
    if request.user.role not in ['admin', 'dost_staff']:
        messages.error(request, "Access denied.")
        return redirect('home')

    # Get user productivity data
    from django.db.models import Sum, Count

    users = User.objects.filter(role='dost_staff', status='active')

    user_data = []
    for user in users:
        tasks_assigned = Task.objects.filter(assigned_to=user).count()
        tasks_completed = Task.objects.filter(assigned_to=user, status='completed').count()
        personal_tasks = PersonalTask.objects.filter(user=user).count()
        personal_tasks_completed = PersonalTask.objects.filter(user=user, status='completed').count()
        total_time = TaskTimeEntry.objects.filter(user=user).aggregate(total=Sum('hours_spent'))['total'] or Decimal('0.00')
        completion_rate = (tasks_completed / tasks_assigned * 100) if tasks_assigned > 0 else 0

        user_data.append({
            'user': user,
            'tasks_assigned': tasks_assigned,
            'tasks_completed': tasks_completed,
            'personal_tasks': personal_tasks,
            'personal_tasks_completed': personal_tasks_completed,
            'total_time_spent': float(total_time),
            'completion_rate': round(completion_rate, 1),
        })

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=getSampleStyleSheet()['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    elements.append(Paragraph("User Productivity Report", title_style))
    elements.append(Spacer(1, 12))

    # Table data
    data = [['User', 'Tasks Assigned', 'Tasks Completed', 'Personal Tasks', 'Personal Completed', 'Time Spent (hrs)', 'Completion Rate (%)']]

    for user_info in user_data:
        data.append([
            user_info['user'].full_name(),
            str(user_info['tasks_assigned']),
            str(user_info['tasks_completed']),
            str(user_info['personal_tasks']),
            str(user_info['personal_tasks_completed']),
            f"{user_info['total_time_spent']:.1f}",
            f"{user_info['completion_rate']:.1f}%"
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="user_productivity_report.pdf"'
    return response


@login_required
def export_project_progress_excel(request):
    """Export project progress report as Excel"""
    if request.user.role not in ['admin', 'dost_staff']:
        messages.error(request, "Access denied.")
        return redirect('home')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Get project data
    projects = Project.objects.all().prefetch_related('tasks')

    project_data = []
    for project in projects:
        tasks = project.tasks.all()
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status='completed').count()
        progress_percentage = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        total_time = TaskTimeEntry.objects.filter(task__project=project).aggregate(
            total=Sum('hours_spent')
        )['total'] or Decimal('0.00')

        project_data.append({
            'project': project,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'progress_percentage': round(progress_percentage, 1),
            'total_time_spent': float(total_time),
        })

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Progress Report"

    # Header
    headers = ['Project Title', 'Total Tasks', 'Completed Tasks', 'Progress (%)', 'Time Spent (hrs)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_num, project_info in enumerate(project_data, 2):
        ws.cell(row=row_num, column=1).value = project_info['project'].project_title or f"Project {project_info['project'].no}"
        ws.cell(row=row_num, column=2).value = project_info['total_tasks']
        ws.cell(row=row_num, column=3).value = project_info['completed_tasks']
        ws.cell(row=row_num, column=4).value = project_info['progress_percentage']
        ws.cell(row=row_num, column=5).value = project_info['total_time_spent']

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="project_progress_report.xlsx"'
    wb.save(response)

# -------------------------------
# Communication Hub Views
# -------------------------------

@login_required
def administrator_communication_hub_view(request):
    """Main communication hub dashboard"""
    user = request.user

    cache_key = f'admin_comm_hub_payload_{user.id}_v1'
    cached_payload = cache.get(cache_key)
    if cached_payload:
        unread_messages = cached_payload.get('unread_messages', 0)
        group_chat_ids = cached_payload.get('group_chat_member_ids', [])
        announcement_ids = cached_payload.get('announcement_ids', [])
        message_ids = cached_payload.get('message_ids', [])

        user_group_chats = GroupChatMember.objects.filter(
            id__in=group_chat_ids,
            user=user,
            is_active=True
        ).select_related('group_chat').order_by('-group_chat__updated_at')

        recent_announcements = Announcement.objects.filter(
            id__in=announcement_ids
        ).order_by('-created_at')

        recent_messages = Message.objects.filter(
            id__in=message_ids
        ).select_related('sender', 'recipient').order_by('-created_at')

        context = {
            'unread_messages': unread_messages,
            'user_group_chats': user_group_chats,
            'recent_announcements': recent_announcements,
            'recent_messages': recent_messages,
        }
        return render(request, 'administrator/communication_hub.html', context)

    # Get unread message count
    unread_messages = Message.objects.filter(recipient=user, is_read=False).count()

    # Get user's group chats
    user_group_chats = GroupChatMember.objects.filter(
        user=user,
        is_active=True
    ).select_related('group_chat').order_by('-group_chat__updated_at')

    # Get recent announcements
    # First get all active announcements, then filter by target roles/users in Python
    # since SQLite doesn't support contains lookup on JSONFields
    all_recent_announcements = Announcement.objects.filter(
        Q(is_active=True) &
        (Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now()))
    ).prefetch_related('target_users').order_by('-created_at')[:20]  # Get more to filter

    # Filter announcements that target this user
    recent_announcements = []
    for announcement in all_recent_announcements:
        # Check if announcement targets everyone (empty target_roles and no target_users)
        if not announcement.target_roles and not announcement.target_users.exists():
            recent_announcements.append(announcement)
        # Check if announcement targets user's role
        elif user.role in announcement.target_roles:
            recent_announcements.append(announcement)
        # Check if announcement targets this specific user
        elif any(target_user.id == user.id for target_user in announcement.target_users.all()):
            recent_announcements.append(announcement)

        if len(recent_announcements) >= 5:
            break

    recent_announcements = recent_announcements[:5]

    # Get recent messages
    recent_messages = Message.objects.filter(
        Q(sender=user) | Q(recipient=user)
    ).select_related('sender', 'recipient').order_by('-created_at')[:10]

    context = {
        'unread_messages': unread_messages,
        'user_group_chats': user_group_chats,
        'recent_announcements': recent_announcements,
        'recent_messages': recent_messages,
    }

    cache.set(
        cache_key,
        {
            'unread_messages': unread_messages,
            'group_chat_member_ids': [member.id for member in user_group_chats],
            'announcement_ids': [announcement.id for announcement in recent_announcements],
            'message_ids': [message.id for message in recent_messages],
        },
        45,
    )

    return render(request, 'administrator/communication_hub.html', context)

@login_required
def administrator_messages_view(request):
    """Messages inbox view - shows conversations with people"""
    user = request.user
    conversations = _build_user_conversations(user)

    context = {
        'conversations': conversations,
    }

    return render(request, 'administrator/messages.html', context)


@login_required
def administrator_conversation_view(request, partner_id):
    """Chat-style conversation view with a specific user"""
    user = request.user

    try:
        partner = User.objects.get(id=partner_id)
    except User.DoesNotExist:
        raise Http404("User not found")

    # Get IDs of messages deleted by this user
    from .models import DeletedMessage
    deleted_message_ids = DeletedMessage.objects.filter(
        user=user
    ).values_list('message_id', flat=True)

    # Check if there's any conversation history between these users
    conversation_exists = Message.objects.filter(
        (Q(sender=user) & Q(recipient=partner)) |
        (Q(sender=partner) & Q(recipient=user))
    ).exclude(id__in=deleted_message_ids).exists()

    if not conversation_exists:
        # No conversation exists, redirect to compose new message
        return redirect('administrator_compose_message_url')

    # Get all messages in this conversation (both directions), excluding deleted ones
    conversation_messages = Message.objects.filter(
        (Q(sender=user) & Q(recipient=partner)) |
        (Q(sender=partner) & Q(recipient=user))
    ).exclude(id__in=deleted_message_ids).select_related('sender', 'recipient').order_by('created_at')

    # Mark all messages from partner as read
    marked_read = Message.objects.filter(
        sender=partner,
        recipient=user,
        is_read=False
    ).update(is_read=True)
    if marked_read:
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    if request.method == 'POST':
        reply_content = request.POST.get('reply_content')
        reply_attachment = request.FILES.get('reply_attachment')

        if reply_content:
            try:
                # Create new message in this conversation
                message = Message.objects.create(
                    sender=user,
                    recipient=partner,
                    subject=f"Message from {user.get_full_name()}",
                    content=reply_content,
                    attachment=reply_attachment,
                    message_type='direct'
                )
                _invalidate_admin_comm_hub_cache_for_users([user.id, partner.id])

                # Log the action
                attachment_info = f" with attachment '{reply_attachment.name}'" if reply_attachment else ""
                AuditLog.objects.create(
                    user=request.user,
                    action='Message Sent',
                    details=f"Sent message to {partner.get_full_name()}{attachment_info}"
                )

                messages.success(request, 'Message sent successfully!')
                return redirect('administrator_conversation_url', partner_id=partner.id)

            except Exception as e:
                messages.error(request, f'Error sending message: {str(e)}')

    context = {
        'partner': partner,
        'conversation_messages': conversation_messages,
    }

    return render(request, 'administrator/conversation.html', context)


@login_required
def administrator_compose_message_view(request):
    """Compose new message view"""
    if request.method == 'POST':
        recipient_id = request.POST.get('recipient')
        subject = request.POST.get('subject')
        content = request.POST.get('content')
        attachment = request.FILES.get('attachment')

        try:
            recipient = User.objects.get(id=recipient_id, status='active')
            message = Message.objects.create(
                sender=request.user,
                recipient=recipient,
                subject=subject,
                content=content,
                attachment=attachment,
                message_type='direct'
            )
            _invalidate_admin_comm_hub_cache_for_users([request.user.id, recipient.id])

            # Log the action
            attachment_info = f" with attachment '{attachment.name}'" if attachment else ""
            AuditLog.objects.create(
                user=request.user,
                action='Message Sent',
                details=f"Sent message to {recipient.get_full_name()}: {subject}{attachment_info}"
            )

            messages.success(request, f'Message sent to {recipient.get_full_name()} successfully!')
            return redirect('administrator_messages_url')

        except User.DoesNotExist:
            messages.error(request, 'Recipient not found.')
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')

    # Get list of active users (excluding current user)
    users = User.objects.filter(status='active').exclude(id=request.user.id).order_by('first_name', 'last_name')

    context = {
        'users': users,
    }

    return render(request, 'administrator/compose_message.html', context)

@login_required
def administrator_message_detail_view(request, message_id):
    """View message details and reply"""
    try:
        message = Message.objects.select_related('sender', 'recipient', 'parent_message').get(
            Q(id=message_id) & (Q(sender=request.user) | Q(recipient=request.user))
        )
    except Message.DoesNotExist:
        raise Http404("Message not found")

    # Mark as read if recipient is viewing
    if message.recipient == request.user and not message.is_read:
        message.mark_as_read()
        _invalidate_admin_comm_hub_cache_for_users([request.user.id])

    # Get message thread (parent and replies)
    thread_messages = []
    if message.parent_message:
        # This is a reply, get the entire thread
        root_message = message.parent_message
        while root_message.parent_message:
            root_message = root_message.parent_message
        thread_messages = Message.objects.filter(
            Q(id=root_message.id) | Q(parent_message=root_message) | Q(parent_message__parent_message=root_message)
        ).select_related('sender', 'recipient').order_by('created_at')
    else:
        # This is the root message
        thread_messages = Message.objects.filter(
            Q(id=message.id) | Q(parent_message=message) | Q(parent_message__parent_message=message)
        ).select_related('sender', 'recipient').order_by('created_at')

    if request.method == 'POST':
        reply_content = request.POST.get('reply_content')
        reply_attachment = request.FILES.get('reply_attachment')
        if reply_content:
            try:
                reply = Message.objects.create(
                    sender=request.user,
                    recipient=message.sender if message.recipient == request.user else message.recipient,
                    subject=f"Re: {message.subject}",
                    content=reply_content,
                    attachment=reply_attachment,
                    message_type='direct',
                    parent_message=message
                )
                _invalidate_admin_comm_hub_cache_for_users([reply.sender_id, reply.recipient_id])

                # Log the action
                attachment_info = f" with attachment '{reply_attachment.name}'" if reply_attachment else ""
                AuditLog.objects.create(
                    user=request.user,
                    action='Message Reply Sent',
                    details=f"Replied to message from {message.sender.get_full_name()}{attachment_info}"
                )

                messages.success(request, 'Reply sent successfully!')
                return redirect('administrator_message_detail_url', message_id=message.id)

            except Exception as e:
                messages.error(request, f'Error sending reply: {str(e)}')

    context = {
        'message': message,
        'thread_messages': thread_messages,
    }

    return render(request, 'administrator/message_detail.html', context)

@login_required
def administrator_group_chats_view(request):
    """Group chats list view"""
    user = request.user

    # Get user's group chats
    user_chats = GroupChatMember.objects.filter(
        user=user,
        is_active=True
    ).select_related('group_chat').order_by('-group_chat__updated_at')

    # Get available projects for creating new group chats
    projects = Project.objects.filter(status__in=['ongoing', 'planning']).order_by('project_title')

    context = {
        'user_chats': user_chats,
        'projects': projects,
    }

    return render(request, 'administrator/group_chats.html', context)

@login_required
def administrator_group_chat_detail_view(request, chat_id):
    """Group chat detail view"""
    try:
        chat_member = GroupChatMember.objects.select_related('group_chat').get(
            group_chat_id=chat_id,
            user=request.user,
            is_active=True
        )
        chat = chat_member.group_chat
    except GroupChatMember.DoesNotExist:
        raise Http404("Group chat not found or access denied")

    # Update last seen
    chat_member.last_seen_at = timezone.now()
    chat_member.save(update_fields=['last_seen_at'])

    # Get recent messages
    messages_qs = GroupChatMessage.objects.filter(
        group_chat=chat
    ).select_related('sender').order_by('-created_at')[:50]
    chat_messages = list(reversed(messages_qs))

    # Get active members
    active_members = GroupChatMember.objects.filter(
        group_chat=chat,
        is_active=True
    ).select_related('user').order_by('user__first_name')

    if request.method == 'POST':
        content = request.POST.get('content')
        if content and content.strip():
            try:
                GroupChatMessage.objects.create(
                    group_chat=chat,
                    sender=request.user,
                    content=content.strip()
                )
                _invalidate_comm_hub_cache_for_group_chat(chat)

                # Update chat's updated_at
                chat.updated_at = timezone.now()
                chat.save(update_fields=['updated_at'])

                # Log the action
                AuditLog.objects.create(
                    user=request.user,
                    action='Group Chat Message Sent',
                    details=f"Sent message in group chat: {chat.name}"
                )

                return redirect('administrator_group_chat_detail_url', chat_id=chat.id)

            except Exception as e:
                messages.error(request, f'Error sending message: {str(e)}')

    context = {
        'chat': chat,
        'chat_messages': chat_messages,
        'active_members': active_members,
        'is_admin': chat_member.role == 'admin',
    }

    return render(request, 'administrator/group_chat_detail.html', context)

@login_required
def administrator_create_group_chat_view(request):
    """Create new group chat"""
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        project_id = request.POST.get('project')
        member_ids = request.POST.getlist('members')

        try:
            # Create the group chat
            group_chat = GroupChat.objects.create(
                name=name,
                description=description,
                project_id=project_id if project_id else None,
                created_by=request.user
            )

            # Add creator as admin
            GroupChatMember.objects.create(
                group_chat=group_chat,
                user=request.user,
                role='admin'
            )

            # Add other members
            for member_id in member_ids:
                if member_id != str(request.user.id):
                    try:
                        user = User.objects.get(id=member_id, status='active')
                        GroupChatMember.objects.create(
                            group_chat=group_chat,
                            user=user,
                            role='member'
                        )
                    except User.DoesNotExist:
                        continue

            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='Group Chat Created',
                details=f"Created group chat: {name}"
            )

            messages.success(request, f'Group chat "{name}" created successfully!')
            return redirect('administrator_group_chat_detail_url', chat_id=group_chat.id)

        except Exception as e:
            messages.error(request, f'Error creating group chat: {str(e)}')

    # Get available users
    users = User.objects.filter(status='active').exclude(id=request.user.id).order_by('first_name', 'last_name')
    projects = Project.objects.filter(status__in=['ongoing', 'planning']).order_by('project_title')

    context = {
        'users': users,
        'projects': projects,
    }

    return render(request, 'administrator/create_group_chat.html', context)

@login_required
def administrator_manage_group_chat_members_view(request, chat_id):
    """Manage group chat members - add/remove/change roles"""
    try:
        # Check if user is admin of this chat
        chat_member = GroupChatMember.objects.get(
            group_chat_id=chat_id,
            user=request.user,
            role='admin',
            is_active=True
        )
        chat = chat_member.group_chat
    except GroupChatMember.DoesNotExist:
        messages.error(request, 'Access denied. Only chat administrators can manage members.')
        return redirect('administrator_group_chat_detail_url', chat_id=chat_id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_member':
            user_id = request.POST.get('user_id')
            role = request.POST.get('role', 'member')

            if not user_id:
                messages.error(request, 'Please select a user to add.')
            else:
                try:
                    user = User.objects.get(id=user_id)
                    # Check if user is already an active member
                    if not GroupChatMember.objects.filter(group_chat=chat, user=user, is_active=True).exists():
                        # Check if user has an inactive membership that we can reactivate
                        existing_membership = GroupChatMember.objects.filter(group_chat=chat, user=user, is_active=False).first()
                        if existing_membership:
                            # Reactivate existing membership
                            existing_membership.is_active = True
                            existing_membership.role = role  # Update role if changed
                            existing_membership.save()
                            messages.success(request, f'{user.get_full_name()} re-added to the group chat.')
                        else:
                            # Create new membership
                            # Activate user if they're deactivated
                            if user.status != 'active':
                                user.status = 'active'
                                user.save()
                                messages.info(request, f'Activated and added {user.get_full_name()} to the group chat.')
                            else:
                                messages.success(request, f'{user.get_full_name()} added to the group chat.')
                            GroupChatMember.objects.create(
                                group_chat=chat,
                                user=user,
                                role=role
                            )
                    else:
                        messages.warning(request, f'{user.get_full_name()} is already a member of this chat.')
                except User.DoesNotExist:
                    messages.error(request, 'Selected user not found.')

        elif action == 'remove_member':
            member_id = request.POST.get('member_id')
            try:
                member = GroupChatMember.objects.get(id=member_id, group_chat=chat)
                if member.user != request.user:  # Can't remove yourself
                    member.is_active = False
                    member.save()
                    messages.success(request, f'{member.user.get_full_name()} removed from the group chat.')
                else:
                    messages.error(request, 'You cannot remove yourself from the chat.')
            except GroupChatMember.DoesNotExist:
                messages.error(request, 'Member not found.')

        elif action == 'change_role':
            member_id = request.POST.get('member_id')
            new_role = request.POST.get('role')
            try:
                member = GroupChatMember.objects.get(id=member_id, group_chat=chat)
                member.role = new_role
                member.save()
                messages.success(request, f'{member.user.get_full_name()} role changed to {new_role}.')
            except GroupChatMember.DoesNotExist:
                messages.error(request, 'Member not found.')

        return redirect('administrator_manage_group_chat_members_url', chat_id=chat_id)

    # Get current members
    current_members = GroupChatMember.objects.filter(
        group_chat=chat,
        is_active=True
    ).select_related('user').order_by('user__first_name')

    # Get available users to add (active users not already in chat)
    existing_user_ids = list(current_members.values_list('user_id', flat=True))
    all_active_users = User.objects.filter(status='active')
    available_users = all_active_users.exclude(id__in=existing_user_ids).order_by('first_name', 'last_name')

    # For debugging: also include inactive users if no active users available
    if not available_users.exists():
        inactive_users = User.objects.filter(status='deactivated').exclude(id__in=existing_user_ids).order_by('first_name', 'last_name')
        if inactive_users.exists():
            messages.info(request, f'No active users available. Showing {inactive_users.count()} inactive users for testing.')
            available_users = inactive_users

    context = {
        'chat': chat,
        'current_members': current_members,
        'available_users': available_users,
    }

    return render(request, 'administrator/manage_group_chat_members.html', context)

@login_required
def administrator_edit_group_chat_settings_view(request, chat_id):
    """Edit group chat settings"""
    try:
        # Check if user is admin of this chat
        chat_member = GroupChatMember.objects.get(
            group_chat_id=chat_id,
            user=request.user,
            role='admin',
            is_active=True
        )
        chat = chat_member.group_chat
    except GroupChatMember.DoesNotExist:
        messages.error(request, 'Access denied. Only chat administrators can edit settings.')
        return redirect('administrator_group_chat_detail_url', chat_id=chat_id)

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')

        if name and name.strip():
            chat.name = name.strip()
            chat.description = description.strip() if description else ''
            chat.save()
            messages.success(request, 'Group chat settings updated successfully.')
            return redirect('administrator_group_chat_detail_url', chat_id=chat_id)
        else:
            messages.error(request, 'Chat name is required.')

    context = {
        'chat': chat,
    }

    return render(request, 'administrator/edit_group_chat_settings.html', context)


# Staff Group Chat Views
@login_required
def staff_group_chats_view(request):
    """List all group chats for staff user"""
    if request.user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    user_group_chats = GroupChatMember.objects.filter(
        user=request.user,
        is_active=True
    ).select_related('group_chat').annotate(
        member_count=Count('group_chat__members', filter=Q(group_chat__members__is_active=True))
    ).order_by('-group_chat__updated_at')

    context = {
        'user_group_chats': user_group_chats,
    }

    return render(request, 'staff/group_chats.html', context)

@login_required
@login_required
def staff_group_chat_detail_view(request, chat_id):
    """View group chat details and messages for staff"""
    if request.user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    try:
        chat_member = GroupChatMember.objects.select_related('group_chat').get(
            group_chat_id=chat_id,
            user=request.user,
            is_active=True
        )
        chat = chat_member.group_chat
    except GroupChatMember.DoesNotExist:
        messages.error(request, 'Access denied. You are not a member of this chat.')
        return redirect('staff_group_chats_url')

    # Get chat messages
    chat_messages = GroupChatMessage.objects.filter(
        group_chat=chat
    ).select_related('sender').order_by('created_at')

    # Get active members
    active_members = GroupChatMember.objects.filter(
        group_chat=chat,
        is_active=True
    ).select_related('user')

    # Check if user is admin
    is_admin = chat_member.role == 'admin'

    if request.method == 'POST':
        content = request.POST.get('content')
        if content and content.strip():
            GroupChatMessage.objects.create(
                group_chat=chat,
                sender=request.user,
                content=content.strip()
            )
            _invalidate_comm_hub_cache_for_group_chat(chat)
            # Update chat's updated_at
            chat.save()
            return redirect('staff_group_chat_detail_url', chat_id=chat.id)

    context = {
        'chat': chat,
        'chat_messages': chat_messages,
        'active_members': active_members,
        'is_admin': is_admin,
    }

    return render(request, 'staff/group_chat_detail.html', context)

@login_required
def staff_create_group_chat_view(request):
    """Create new group chat for staff"""
    if request.user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        project_id = request.POST.get('project')
        member_ids = request.POST.getlist('members')

        try:
            # Create the group chat
            group_chat = GroupChat.objects.create(
                name=name,
                description=description,
                project_id=project_id if project_id else None,
                created_by=request.user
            )

            # Add creator as admin
            GroupChatMember.objects.create(
                group_chat=group_chat,
                user=request.user,
                role='admin'
            )

            # Add other members
            for member_id in member_ids:
                if member_id != str(request.user.id):
                    try:
                        user = User.objects.get(id=member_id, status='active')
                        GroupChatMember.objects.create(
                            group_chat=group_chat,
                            user=user,
                            role='member'
                        )
                    except User.DoesNotExist:
                        continue

            messages.success(request, f'Group chat "{name}" created successfully!')
            return redirect('staff_group_chat_detail_url', chat_id=group_chat.id)

        except Exception as e:
            messages.error(request, f'Error creating group chat: {str(e)}')

    # Get available users (active users except current user)
    users = User.objects.filter(status='active').exclude(id=request.user.id).order_by('first_name', 'last_name')
    # Get projects that this staff user is involved with
    projects = Project.objects.filter(
        project_leader=request.user
    ).order_by('project_title')


    context = {
        'users': users,
        'projects': projects,
    }

    return render(request, 'staff/create_group_chat.html', context)

# Proponent Group Chat Views
@login_required
@login_required
def proponent_group_chats_view(request):
    """List all group chats for proponent user"""
    if request.user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    user_group_chats = GroupChatMember.objects.filter(
        user=request.user,
        is_active=True
    ).select_related('group_chat').annotate(
        member_count=Count('group_chat__members', filter=Q(group_chat__members__is_active=True))
    ).order_by('-group_chat__updated_at')

    context = {
        'user_group_chats': user_group_chats,
    }

    return render(request, 'proponent/group_chats.html', context)

@login_required
def proponent_group_chat_detail_view(request, chat_id):
    """View group chat details and messages for proponent"""
    if request.user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    try:
        chat_member = GroupChatMember.objects.select_related('group_chat').get(
            group_chat_id=chat_id,
            user=request.user,
            is_active=True
        )
        chat = chat_member.group_chat
    except GroupChatMember.DoesNotExist:
        messages.error(request, 'Access denied. You are not a member of this chat.')
        return redirect('proponent_group_chats_url')

    # Get chat messages
    chat_messages = GroupChatMessage.objects.filter(
        group_chat=chat
    ).select_related('sender').order_by('created_at')

    # Get active members
    active_members = GroupChatMember.objects.filter(
        group_chat=chat,
        is_active=True
    ).select_related('user')

    # Check if user is admin
    is_admin = chat_member.role == 'admin'

    if request.method == 'POST':
        content = request.POST.get('content')
        if content and content.strip():
            GroupChatMessage.objects.create(
                group_chat=chat,
                sender=request.user,
                content=content.strip()
            )
            _invalidate_comm_hub_cache_for_group_chat(chat)
            # Update chat's updated_at
            chat.save()
            return redirect('proponent_group_chat_detail_url', chat_id=chat.id)

    context = {
        'chat': chat,
        'chat_messages': chat_messages,
        'active_members': active_members,
        'is_admin': is_admin,
    }

    return render(request, 'proponent/group_chat_detail.html', context)

@login_required
def proponent_create_group_chat_view(request):
    """Create new group chat for proponent - DISABLED: Only Admin/Staff can create group chats"""
    messages.error(request, 'Only administrators and staff can create group chats.')
    return redirect('proponent_group_chats_url')

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        project_id = request.POST.get('project')
        member_ids = request.POST.getlist('members')

        try:
            # Create the group chat
            group_chat = GroupChat.objects.create(
                name=name,
                description=description,
                project_id=project_id if project_id else None,
                created_by=request.user
            )

            # Add creator as admin
            GroupChatMember.objects.create(
                group_chat=group_chat,
                user=request.user,
                role='admin'
            )

            # Add other members
            for member_id in member_ids:
                if member_id != str(request.user.id):
                    try:
                        user = User.objects.get(id=member_id, status='active')
                        GroupChatMember.objects.create(
                            group_chat=group_chat,
                            user=user,
                            role='member'
                        )
                    except User.DoesNotExist:
                        continue

            messages.success(request, f'Group chat "{name}" created successfully!')
            return redirect('proponent_group_chat_detail_url', chat_id=group_chat.id)

        except Exception as e:
            messages.error(request, f'Error creating group chat: {str(e)}')

    # Get available users (active users except current user)
    users = User.objects.filter(status='active').exclude(id=request.user.id).order_by('first_name', 'last_name')
    # Get projects that this proponent user is involved with
    projects = Project.objects.filter(
        Q(created_by=request.user) |
        Q(proponent=request.user)
    ).distinct().order_by('project_title')

    context = {
        'users': users,
        'projects': projects,
    }

    return render(request, 'proponent/create_group_chat.html', context)

# Beneficiary Group Chat Views
@login_required
def beneficiary_group_chats_view(request):
    """List all group chats for beneficiary user"""
    if request.user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('beneficiary_dashboard_url')

    user_group_chats = GroupChatMember.objects.filter(
        user=request.user,
        is_active=True
    ).select_related('group_chat').annotate(
        member_count=Count('group_chat__members', filter=Q(group_chat__members__is_active=True))
    ).order_by('-group_chat__updated_at')

    context = {
        'user_group_chats': user_group_chats,
    }

    return render(request, 'beneficiary/group_chats.html', context)

@login_required
def beneficiary_group_chat_detail_view(request, chat_id):
    """View group chat details and messages for beneficiary"""
    if request.user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('beneficiary_dashboard_url')

    try:
        chat_member = GroupChatMember.objects.select_related('group_chat').get(
            group_chat_id=chat_id,
            user=request.user,
            is_active=True
        )
        chat = chat_member.group_chat
    except GroupChatMember.DoesNotExist:
        messages.error(request, 'Access denied. You are not a member of this chat.')
        return redirect('beneficiary_group_chats_url')

    # Get chat messages
    chat_messages = GroupChatMessage.objects.filter(
        group_chat=chat
    ).select_related('sender').order_by('created_at')

    # Get active members
    active_members = GroupChatMember.objects.filter(
        group_chat=chat,
        is_active=True
    ).select_related('user')

    # Check if user is admin
    is_admin = chat_member.role == 'admin'

    if request.method == 'POST':
        content = request.POST.get('content')
        if content and content.strip():
            GroupChatMessage.objects.create(
                group_chat=chat,
                sender=request.user,
                content=content.strip()
            )
            _invalidate_comm_hub_cache_for_group_chat(chat)
            # Update chat's updated_at
            chat.save()
            return redirect('beneficiary_group_chat_detail_url', chat_id=chat.id)

    context = {
        'chat': chat,
        'chat_messages': chat_messages,
        'active_members': active_members,
        'is_admin': is_admin,
    }

    return render(request, 'beneficiary/group_chat_detail.html', context)

@login_required
def beneficiary_create_group_chat_view(request):
    """Create new group chat for beneficiary - DISABLED: Only Admin/Staff can create group chats"""
    messages.error(request, 'Only administrators and staff can create group chats.')
    return redirect('beneficiary_group_chats_url')

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        project_id = request.POST.get('project')
        member_ids = request.POST.getlist('members')

        try:
            # Create the group chat
            group_chat = GroupChat.objects.create(
                name=name,
                description=description,
                project_id=project_id if project_id else None,
                created_by=request.user
            )

            # Add creator as admin
            GroupChatMember.objects.create(
                group_chat=group_chat,
                user=request.user,
                role='admin'
            )

            # Add other members
            for member_id in member_ids:
                if member_id != str(request.user.id):
                    try:
                        user = User.objects.get(id=member_id, status='active')
                        GroupChatMember.objects.create(
                            group_chat=group_chat,
                            user=user,
                            role='member'
                        )
                    except User.DoesNotExist:
                        continue

            messages.success(request, f'Group chat "{name}" created successfully!')
            return redirect('beneficiary_group_chat_detail_url', chat_id=group_chat.id)

        except Exception as e:
            messages.error(request, f'Error creating group chat: {str(e)}')

    # Get available users (active users except current user)
    users = User.objects.filter(status='active').exclude(id=request.user.id).order_by('first_name', 'last_name')
    # Get projects that this beneficiary user is involved with
    projects = Project.objects.filter(
        Q(beneficiaries=request.user)
    ).distinct().order_by('project_title')

    context = {
        'users': users,
        'projects': projects,
    }

    return render(request, 'beneficiary/create_group_chat.html', context)


# =========================================
# One-Sided Delete Views for Messages
# =========================================

@login_required
def delete_message_view(request, message_id):
    """Delete a single message for the current user only (one-sided)"""
    if request.method == 'POST':
        try:
            message = Message.objects.get(id=message_id)
            # Verify user is sender or recipient
            if request.user not in [message.sender, message.recipient]:
                return JsonResponse({'success': False, 'error': 'Access denied'})
            
            # Create one-sided delete record
            DeletedMessage.objects.get_or_create(
                user=request.user,
                message=message
            )
            _invalidate_admin_comm_hub_cache_for_users([request.user.id])
            
            return JsonResponse({'success': True, 'message': 'Message deleted for you'})
        except Message.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Message not found'})
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@login_required
def delete_conversation_view(request, partner_id):
    """Delete entire conversation with a user for the current user only (one-sided)"""
    if request.method == 'POST':
        try:
            partner = User.objects.get(id=partner_id)
            
            # Create or update one-sided delete record
            # All messages before this timestamp will be hidden for this user
            deleted_conv, created = DeletedConversation.objects.update_or_create(
                user=request.user,
                partner=partner,
                defaults={'delete_before': timezone.now()}
            )
            _invalidate_admin_comm_hub_cache_for_users([request.user.id])
            
            return JsonResponse({'success': True, 'message': 'Conversation deleted for you'})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@login_required  
def delete_group_chat_view(request, chat_id):
    """Delete/hide a group chat for the current user only (one-sided)"""
    if request.method == 'POST':
        try:
            group_chat = GroupChat.objects.get(id=chat_id)
            
            # Verify user is a member
            if not GroupChatMember.objects.filter(group_chat=group_chat, user=request.user).exists():
                return JsonResponse({'success': False, 'error': 'Access denied'})
            
            # Create one-sided delete record
            DeletedGroupChat.objects.update_or_create(
                user=request.user,
                group_chat=group_chat,
                defaults={'delete_before': timezone.now()}
            )
            _invalidate_comm_hub_cache_for_group_chat(group_chat)
            
            return JsonResponse({'success': True, 'message': 'Group chat deleted for you'})
        except GroupChat.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Group chat not found'})
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@login_required
def delete_group_message_view(request, message_id):
    """Delete a single group chat message for the current user only (one-sided)"""
    if request.method == 'POST':
        try:
            message = GroupChatMessage.objects.get(id=message_id)
            
            # Verify user is a member of the group chat
            if not GroupChatMember.objects.filter(group_chat=message.group_chat, user=request.user).exists():
                return JsonResponse({'success': False, 'error': 'Access denied'})
            
            # Create one-sided delete record
            DeletedGroupChatMessage.objects.get_or_create(
                user=request.user,
                message=message
            )
            _invalidate_comm_hub_cache_for_group_chat(message.group_chat)
            
            return JsonResponse({'success': True, 'message': 'Message deleted for you'})
        except GroupChatMessage.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Message not found'})
    return JsonResponse({'success': False, 'error': 'Invalid method'})


@login_required
def administrator_announcements_view(request):
    """Announcements management view"""
    user = request.user

    # Only admins can manage announcements
    if user.role != 'admin':
        messages.error(request, 'Access denied. Only administrators can manage announcements.')
        return redirect('administrator_dashboard_url')

    announcements = Announcement.objects.all().order_by('-created_at')

    context = {
        'announcements': announcements,
    }

    return render(request, 'administrator/announcements.html', context)

@login_required
def administrator_create_announcement_view(request):
    """Create new announcement"""
    user = request.user

    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('administrator_dashboard_url')

    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        priority = request.POST.get('priority', 'normal')
        expires_at = request.POST.get('expires_at')
        target_roles = request.POST.getlist('target_roles')
        target_users = request.POST.getlist('target_users')

        try:
            expires_datetime = None
            if expires_at:
                expires_datetime = timezone.datetime.fromisoformat(expires_at.replace('T', ' '))

            announcement = Announcement.objects.create(
                title=title,
                content=content,
                priority=priority,
                expires_at=expires_datetime,
                created_by=user,
                target_roles=target_roles if target_roles else [],
            )

            # Add target users if specified
            if target_users:
                announcement.target_users.set(target_users)

            _invalidate_admin_comm_hub_cache_for_users(_get_announcement_target_user_ids(announcement))

            # Log the action
            AuditLog.objects.create(
                user=user,
                action='Announcement Created',
                details=f"Created announcement: {title}"
            )

            messages.success(request, 'Announcement created successfully!')
            return redirect('administrator_announcements_url')

        except Exception as e:
            messages.error(request, f'Error creating announcement: {str(e)}')

    # Get all users for targeting
    users = User.objects.filter(status='active').order_by('first_name', 'last_name')
    roles = User.ROLE_CHOICES

    context = {
        'users': users,
        'roles': roles,
    }

    return render(request, 'administrator/create_announcement.html', context)


@login_required
def administrator_update_announcement_view(request, pk):
    """Update an existing announcement"""
    user = request.user

    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('administrator_dashboard_url')

    try:
        announcement = Announcement.objects.get(pk=pk)
    except Announcement.DoesNotExist:
        messages.error(request, 'Announcement not found.')
        return redirect('administrator_announcements_url')

    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        priority = request.POST.get('priority', 'normal')
        expires_at = request.POST.get('expires_at')
        target_roles = request.POST.getlist('target_roles')
        is_active = request.POST.get('is_active') == 'on'

        try:
            expires_datetime = None
            if expires_at:
                expires_datetime = timezone.datetime.fromisoformat(expires_at.replace('T', ' '))

            announcement.title = title
            announcement.content = content
            announcement.priority = priority
            announcement.expires_at = expires_datetime
            announcement.target_roles = target_roles if target_roles else []
            announcement.is_active = is_active
            announcement.save()

            _invalidate_admin_comm_hub_cache_for_users(_get_announcement_target_user_ids(announcement))

            # Log the action
            AuditLog.objects.create(
                user=user,
                action='Announcement Updated',
                details=f"Updated announcement: {title}"
            )

            messages.success(request, 'Announcement updated successfully!')
            return redirect('administrator_announcements_url')

        except Exception as e:
            messages.error(request, f'Error updating announcement: {str(e)}')

    # Get all users for targeting
    users = User.objects.filter(status='active').order_by('first_name', 'last_name')
    roles = User.ROLE_CHOICES

    context = {
        'announcement': announcement,
        'users': users,
        'roles': roles,
    }

    return render(request, 'administrator/edit_announcement.html', context)


@login_required
def administrator_delete_announcement_view(request, pk):
    """Delete an announcement"""
    user = request.user

    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('administrator_dashboard_url')

    try:
        announcement = Announcement.objects.get(pk=pk)
        affected_user_ids = _get_announcement_target_user_ids(announcement)
        title = announcement.title
        announcement.delete()
        _invalidate_admin_comm_hub_cache_for_users(affected_user_ids)

        # Log the action
        AuditLog.objects.create(
            user=user,
            action='Announcement Deleted',
            details=f"Deleted announcement: {title}"
        )

        messages.success(request, 'Announcement deleted successfully!')
    except Announcement.DoesNotExist:
        messages.error(request, 'Announcement not found.')
    except Exception as e:
        messages.error(request, f'Error deleting announcement: {str(e)}')

    return redirect('administrator_announcements_url')


# -------------------------------
# System Health Monitoring Views
# -------------------------------

def collect_system_metrics():
    """Collect current system health metrics"""
    import psutil
    import time
    from django.db import connection
    from django.contrib.sessions.models import Session

    metrics = []

    try:
        # CPU Usage
        cpu_percent = psutil.cpu_percent(interval=1)
        cpu_status = 'healthy' if cpu_percent < 70 else 'warning' if cpu_percent < 90 else 'critical'
        metrics.append(('cpu_usage', cpu_percent, 'percent', cpu_status))

        # Memory Usage
        memory = psutil.virtual_memory()
        memory_percent = memory.percent
        memory_status = 'healthy' if memory_percent < 70 else 'warning' if memory_percent < 90 else 'critical'
        metrics.append(('memory_usage', memory_percent, 'percent', memory_status))

        # Disk Usage
        disk = psutil.disk_usage('/')
        disk_percent = disk.percent
        disk_status = 'healthy' if disk_percent < 70 else 'warning' if disk_percent < 90 else 'critical'
        metrics.append(('disk_usage', disk_percent, 'percent', disk_status))

        # Database Connections
        db_connections = len(connection.queries)
        db_status = 'healthy' if db_connections < 50 else 'warning' if db_connections < 100 else 'critical'
        metrics.append(('database_connections', db_connections, 'count', db_status))

        # Active Users (rough estimate from sessions)
        active_sessions = Session.objects.filter(expire_date__gte=timezone.now()).count()
        active_users_status = 'healthy' if active_sessions < 50 else 'warning' if active_sessions < 100 else 'critical'
        metrics.append(('active_users', active_sessions, 'count', active_users_status))

        # Response Time (mock - would need actual monitoring)
        response_time = 0.1  # seconds
        response_status = 'healthy' if response_time < 1.0 else 'warning' if response_time < 5.0 else 'critical'
        metrics.append(('response_time', response_time, 'seconds', response_status))

        # Network Latency (mock)
        network_latency = 50  # ms
        network_status = 'healthy' if network_latency < 100 else 'warning' if network_latency < 500 else 'critical'
        metrics.append(('network_latency', network_latency, 'ms', network_status))

        # System Uptime
        uptime_seconds = time.time() - psutil.boot_time()
        uptime_hours = uptime_seconds / 3600
        uptime_status = 'healthy'
        metrics.append(('uptime', uptime_hours, 'hours', uptime_status))

        # Error Rate (mock - would need actual error tracking)
        error_rate = 0.01  # percent
        error_status = 'healthy' if error_rate < 1.0 else 'warning' if error_rate < 5.0 else 'critical'
        metrics.append(('error_rate', error_rate, 'percent', error_status))

        # Queue Length (mock - for background tasks)
        queue_length = 0
        queue_status = 'healthy' if queue_length < 10 else 'warning' if queue_length < 50 else 'critical'
        metrics.append(('queue_length', queue_length, 'count', queue_status))

    except Exception as e:
        # If system monitoring fails, return mock data
        print(f"System monitoring error: {e}")
        metrics = [
            ('cpu_usage', 45.0, 'percent', 'healthy'),
            ('memory_usage', 60.0, 'percent', 'healthy'),
            ('disk_usage', 55.0, 'percent', 'healthy'),
            ('database_connections', 5, 'count', 'healthy'),
            ('active_users', 12, 'count', 'healthy'),
            ('response_time', 0.2, 'seconds', 'healthy'),
            ('network_latency', 25, 'ms', 'healthy'),
            ('uptime', 168.0, 'hours', 'healthy'),
            ('error_rate', 0.0, 'percent', 'healthy'),
            ('queue_length', 0, 'count', 'healthy'),
        ]

    return metrics

@login_required
def administrator_system_health_view(request):
    """System health dashboard"""
    user = request.user

    # Only admins can access system health
    if user.role != 'admin':
        messages.error(request, 'Access denied. Only administrators can access system health.')
        return redirect('administrator_dashboard_url')

    # Collect current system metrics
    current_metrics = collect_system_metrics()

    # Save metrics to database
    for metric_type, value, unit, status in current_metrics:
        SystemHealth.objects.create(
            metric_type=metric_type,
            value=value,
            unit=unit,
            status=status
        )

    # Get latest health metrics
    latest_metrics = {}
    for metric_type, _ in SystemHealth.METRIC_TYPES:
        try:
            metric = SystemHealth.objects.filter(metric_type=metric_type).order_by('-recorded_at').first()
            if metric:
                latest_metrics[metric_type] = metric
        except:
            pass

    # Get system alerts
    alerts = []
    if latest_metrics.get('cpu_usage') and latest_metrics['cpu_usage'].status == 'critical':
        alerts.append({
            'type': 'critical',
            'message': f"CPU usage is critically high: {latest_metrics['cpu_usage'].value}%"
        })
    if latest_metrics.get('memory_usage') and latest_metrics['memory_usage'].status == 'critical':
        alerts.append({
            'type': 'critical',
            'message': f"Memory usage is critically high: {latest_metrics['memory_usage'].value}%"
        })
    if latest_metrics.get('disk_usage') and latest_metrics['disk_usage'].status == 'critical':
        alerts.append({
            'type': 'warning',
            'message': f"Disk usage is high: {latest_metrics['disk_usage'].value}%"
        })

    context = {
        'latest_metrics': latest_metrics,
        'alerts': alerts,
    }

    return render(request, 'administrator/system_health.html', context)

@login_required
def administrator_backup_management_view(request):
    """Backup management view"""
    user = request.user

    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('administrator_dashboard_url')

    backups = BackupStatus.objects.all().order_by('-created_at')

    context = {
        'backups': backups,
    }

    return render(request, 'administrator/backup_management.html', context)

@login_required
def administrator_maintenance_schedule_view(request):
    """Maintenance schedule management"""
    user = request.user

    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('administrator_dashboard_url')

    maintenance_tasks = MaintenanceSchedule.objects.all().order_by('scheduled_at')

    context = {
        'maintenance_tasks': maintenance_tasks,
    }

    return render(request, 'administrator/maintenance_schedule.html', context)

@login_required
def administrator_create_maintenance_task_view(request):
    """Create new maintenance task"""
    user = request.user

    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('administrator_dashboard_url')

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        maintenance_type = request.POST.get('maintenance_type')
        scheduled_at = request.POST.get('scheduled_at')
        estimated_duration_hours = request.POST.get('estimated_duration_hours')
        assigned_to_id = request.POST.get('assigned_to')
        is_recurring = request.POST.get('is_recurring') == 'on'
        recurrence_pattern = request.POST.get('recurrence_pattern')

        try:
            scheduled_datetime = timezone.datetime.fromisoformat(scheduled_at.replace('T', ' '))

            estimated_duration = None
            if estimated_duration_hours:
                estimated_duration = timezone.timedelta(hours=int(estimated_duration_hours))

            assigned_to = None
            if assigned_to_id:
                assigned_to = User.objects.get(id=assigned_to_id, status='active')

            MaintenanceSchedule.objects.create(
                title=title,
                description=description,
                maintenance_type=maintenance_type,
                scheduled_at=scheduled_datetime,
                estimated_duration=estimated_duration,
                created_by=user,
                assigned_to=assigned_to,
                is_recurring=is_recurring,
                recurrence_pattern=recurrence_pattern if is_recurring else None,
            )

            # Log the action
            AuditLog.objects.create(
                user=user,
                action='Maintenance Task Created',
                details=f"Created maintenance task: {title}"
            )

            messages.success(request, 'Maintenance task created successfully!')
            return redirect('administrator_maintenance_schedule_url')

        except Exception as e:
            messages.error(request, f'Error creating maintenance task: {str(e)}')

    # Get available users for assignment
    users = User.objects.filter(status='active').order_by('first_name', 'last_name')

    context = {
        'users': users,
        'maintenance_types': MaintenanceSchedule.MAINTENANCE_TYPES,
    }

    return render(request, 'administrator/create_maintenance_task.html', context)

# ===============================
# Messaging for Other Roles
# ===============================

@login_required
def staff_messages_view(request):
    """Messages inbox view for staff - shows conversations with people"""
    user = request.user
    if user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    conversations = _build_user_conversations(user)

    context = {
        'conversations': conversations,
    }

    return render(request, 'staff/messages.html', context)

@login_required
def staff_conversation_view(request, partner_id):
    """Chat-style conversation view for staff"""
    user = request.user
    if user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    try:
        partner = User.objects.get(id=partner_id)
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('staff_messages_url')

    # Get IDs of messages deleted by this user
    deleted_msg_ids = DeletedMessage.objects.filter(user=user).values('message_id')

    # Get all messages between user and partner (excluding deleted ones)
    conversation_messages = Message.objects.filter(
        (Q(sender=user) & Q(recipient=partner)) |
        (Q(sender=partner) & Q(recipient=user))
    ).exclude(id__in=Subquery(deleted_msg_ids)).select_related('sender', 'recipient').order_by('created_at')

    # Mark messages from partner as read
    marked_read = Message.objects.filter(
        sender=partner,
        recipient=user,
        is_read=False
    ).update(is_read=True)
    if marked_read:
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    # Handle new message submission
    if request.method == 'POST':
        form = MessageForm(request.POST, request.FILES)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = user
            message.recipient = partner
            message.save()
            _invalidate_admin_comm_hub_cache_for_users([user.id, partner.id])
            messages.success(request, 'Message sent successfully.')
            return redirect('staff_conversation_url', partner_id=partner_id)
    else:
        form = MessageForm()

    context = {
        'partner': partner,
        'conversation_messages': conversation_messages,
        'form': form,
    }

    return render(request, 'staff/conversation.html', context)

@login_required
def staff_compose_message_view(request):
    """Compose new message view for staff"""
    user = request.user
    if user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    if request.method == 'POST':
        recipient_id = request.POST.get('recipient')
        subject = request.POST.get('subject')
        content = request.POST.get('content')
        attachment = request.FILES.get('attachment')

        try:
            recipient = User.objects.get(id=recipient_id, status='active')
            message = Message.objects.create(
                sender=user,
                recipient=recipient,
                subject=subject,
                content=content,
                attachment=attachment,
                message_type='direct'
            )
            _invalidate_admin_comm_hub_cache_for_users([user.id, recipient.id])

            # Log the action
            attachment_info = f" with attachment '{attachment.name}'" if attachment else ""
            AuditLog.objects.create(
                user=user,
                action='Message Sent',
                details=f"Sent message to {recipient.get_full_name()}: {subject}{attachment_info}"
            )

            messages.success(request, f'Message sent to {recipient.get_full_name()} successfully!')
            return redirect('staff_messages_url')

        except User.DoesNotExist:
            messages.error(request, 'Recipient not found.')
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')

    # Get list of active users (excluding current user)
    users = User.objects.filter(status='active').exclude(id=user.id).order_by('first_name', 'last_name')

    context = {
        'users': users,
    }

    return render(request, 'staff/compose_message.html', context)

@login_required
def staff_message_detail_view(request, message_id):
    """View message details and reply for staff"""
    user = request.user
    if user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    try:
        message = Message.objects.select_related('sender', 'recipient', 'parent_message').get(
            Q(id=message_id) & (Q(sender=user) | Q(recipient=user))
        )
    except Message.DoesNotExist:
        raise Http404("Message not found")

    # Mark as read if recipient is viewing
    if message.recipient == user and not message.is_read:
        message.mark_as_read()
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    # Get message thread (parent and replies)
    thread_messages = []
    if message.parent_message:
        # This is a reply, get the entire thread
        root_message = message.parent_message
        while root_message.parent_message:
            root_message = root_message.parent_message
        thread_messages = Message.objects.filter(
            Q(id=root_message.id) | Q(parent_message=root_message) | Q(parent_message__parent_message=root_message)
        ).select_related('sender', 'recipient').order_by('created_at')
    else:
        # This is the root message
        thread_messages = Message.objects.filter(
            Q(id=message.id) | Q(parent_message=message) | Q(parent_message__parent_message=message)
        ).select_related('sender', 'recipient').order_by('created_at')

    if request.method == 'POST':
        reply_content = request.POST.get('reply_content')
        reply_attachment = request.FILES.get('reply_attachment')
        if reply_content:
            try:
                reply = Message.objects.create(
                    sender=user,
                    recipient=message.sender if message.recipient == user else message.recipient,
                    subject=f"Re: {message.subject}",
                    content=reply_content,
                    attachment=reply_attachment,
                    message_type='direct',
                    parent_message=message
                )
                _invalidate_admin_comm_hub_cache_for_users([reply.sender_id, reply.recipient_id])

                # Log the action
                attachment_info = f" with attachment '{reply_attachment.name}'" if reply_attachment else ""
                AuditLog.objects.create(
                    user=user,
                    action='Message Reply Sent',
                    details=f"Replied to message from {message.sender.get_full_name()}{attachment_info}"
                )

                messages.success(request, 'Reply sent successfully!')
                return redirect('staff_message_detail_url', message_id=message.id)

            except Exception as e:
                messages.error(request, f'Error sending reply: {str(e)}')

    context = {
        'message': message,
        'thread_messages': thread_messages,
    }

    return render(request, 'staff/message_detail.html', context)

@login_required
def staff_announcements_view(request):
    """View announcements for staff"""
    user = request.user
    if user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('staff_dashboard_url')

    # Get announcements for this user (all staff announcements or targeted ones)
    # Filter in Python since SQLite doesn't support JSON field lookups like __contains
    all_announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')
    announcements = []
    for announcement in all_announcements:
        target_roles = announcement.target_roles or []
        # Include if no target roles specified, or if dost_staff is in target roles
        if not target_roles or 'dost_staff' in target_roles:
            announcements.append(announcement)

    context = {
        'announcements': announcements,
    }

    return render(request, 'staff/announcements.html', context)

# Proponent messaging views
@login_required
def proponent_messages_view(request):
    """Messages inbox view for proponents - shows conversations with people"""
    user = request.user
    if user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    conversations = _build_user_conversations(user)

    context = {
        'conversations': conversations,
    }

    return render(request, 'proponent/messages.html', context)

@login_required
def proponent_conversation_view(request, partner_id):
    """Chat-style conversation view for proponents"""
    user = request.user
    if user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    try:
        partner = User.objects.get(id=partner_id)
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('proponent_messages_url')

    # Get IDs of messages deleted by this user
    deleted_msg_ids = DeletedMessage.objects.filter(user=user).values('message_id')

    # Get all messages between user and partner (excluding deleted ones)
    conversation_messages = Message.objects.filter(
        (Q(sender=user) & Q(recipient=partner)) |
        (Q(sender=partner) & Q(recipient=user))
    ).exclude(id__in=Subquery(deleted_msg_ids)).select_related('sender', 'recipient').order_by('created_at')

    # Mark messages from partner as read
    marked_read = Message.objects.filter(
        sender=partner,
        recipient=user,
        is_read=False
    ).update(is_read=True)
    if marked_read:
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    # Handle new message submission
    if request.method == 'POST':
        form = MessageForm(request.POST, request.FILES)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = user
            message.recipient = partner
            message.save()
            _invalidate_admin_comm_hub_cache_for_users([user.id, partner.id])
            messages.success(request, 'Message sent successfully.')
            return redirect('proponent_conversation_url', partner_id=partner_id)
    else:
        form = MessageForm()

    context = {
        'partner': partner,
        'conversation_messages': conversation_messages,
        'form': form,
    }

    return render(request, 'proponent/conversation.html', context)

@login_required
def proponent_compose_message_view(request):
    """Compose new message view for proponents"""
    user = request.user
    if user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    if request.method == 'POST':
        recipient_id = request.POST.get('recipient')
        subject = request.POST.get('subject')
        content = request.POST.get('content')
        attachment = request.FILES.get('attachment')

        try:
            recipient = User.objects.get(id=recipient_id, status='active')
            message = Message.objects.create(
                sender=user,
                recipient=recipient,
                subject=subject,
                content=content,
                attachment=attachment,
                message_type='direct'
            )
            _invalidate_admin_comm_hub_cache_for_users([user.id, recipient.id])

            # Log the action
            attachment_info = f" with attachment '{attachment.name}'" if attachment else ""
            AuditLog.objects.create(
                user=user,
                action='Message Sent',
                details=f"Sent message to {recipient.get_full_name()}: {subject}{attachment_info}"
            )

            messages.success(request, f'Message sent to {recipient.get_full_name()} successfully!')
            return redirect('proponent_messages_url')

        except User.DoesNotExist:
            messages.error(request, 'Recipient not found.')
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')

    # Get list of staff assigned to tasks in projects from this proponent's proposals
    # Plus all admins - proponents can message assigned staff AND all admins
    assigned_staff_ids = Task.objects.filter(
        project__proposal__proponent=user,
        assigned_to__isnull=False
    ).values_list('assigned_to', flat=True).distinct()
    
    # Get all admins plus staff assigned to proponent's projects
    users = User.objects.filter(
        Q(role='admin') |  # All admins
        Q(id__in=assigned_staff_ids, role='dost_staff'),  # Staff assigned to their projects
        status='active'
    ).exclude(id=user.id).order_by('first_name', 'last_name')

    context = {
        'users': users,
    }

    return render(request, 'proponent/compose_message.html', context)

@login_required
def proponent_message_detail_view(request, message_id):
    """View message details and reply for proponents"""
    user = request.user
    if user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    try:
        message = Message.objects.select_related('sender', 'recipient', 'parent_message').get(
            Q(id=message_id) & (Q(sender=user) | Q(recipient=user))
        )
    except Message.DoesNotExist:
        raise Http404("Message not found")

    # Mark as read if recipient is viewing
    if message.recipient == user and not message.is_read:
        message.mark_as_read()
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    # Get message thread (parent and replies)
    thread_messages = []
    if message.parent_message:
        # This is a reply, get the entire thread
        root_message = message.parent_message
        while root_message.parent_message:
            root_message = root_message.parent_message
        thread_messages = Message.objects.filter(
            Q(id=root_message.id) | Q(parent_message=root_message) | Q(parent_message__parent_message=root_message)
        ).select_related('sender', 'recipient').order_by('created_at')
    else:
        # This is the root message
        thread_messages = Message.objects.filter(
            Q(id=message.id) | Q(parent_message=message) | Q(parent_message__parent_message=message)
        ).select_related('sender', 'recipient').order_by('created_at')

    if request.method == 'POST':
        reply_content = request.POST.get('reply_content')
        reply_attachment = request.FILES.get('reply_attachment')
        if reply_content:
            try:
                reply = Message.objects.create(
                    sender=user,
                    recipient=message.sender if message.recipient == user else message.recipient,
                    subject=f"Re: {message.subject}",
                    content=reply_content,
                    attachment=reply_attachment,
                    message_type='direct',
                    parent_message=message
                )
                _invalidate_admin_comm_hub_cache_for_users([reply.sender_id, reply.recipient_id])

                # Log the action
                attachment_info = f" with attachment '{reply_attachment.name}'" if reply_attachment else ""
                AuditLog.objects.create(
                    user=user,
                    action='Message Reply Sent',
                    details=f"Replied to message from {message.sender.get_full_name()}{attachment_info}"
                )

                messages.success(request, 'Reply sent successfully!')
                return redirect('proponent_message_detail_url', message_id=message.id)

            except Exception as e:
                messages.error(request, f'Error sending reply: {str(e)}')

    context = {
        'message': message,
        'thread_messages': thread_messages,
    }

    return render(request, 'proponent/message_detail.html', context)

@login_required
def proponent_announcements_view(request):
    """View announcements for proponents"""
    user = request.user
    if user.role != 'proponent':
        messages.error(request, 'Access denied.')
        return redirect('proponent_dashboard_url')

    # Get announcements for this user (all proponent announcements or targeted ones)
    # Filter in Python since SQLite doesn't support JSON field lookups like __contains
    all_announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')
    announcements = []
    for announcement in all_announcements:
        target_roles = announcement.target_roles or []
        # Include if no target roles specified, or if proponent is in target roles
        if not target_roles or 'proponent' in target_roles:
            announcements.append(announcement)

    context = {
        'announcements': announcements,
    }

    return render(request, 'proponent/announcements.html', context)

# Beneficiary messaging views
@login_required
def beneficiary_messages_view(request):
    """Messages inbox view for beneficiaries - shows conversations with people"""
    user = request.user
    if user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    conversations = _build_user_conversations(user)

    context = {
        'conversations': conversations,
    }

    return render(request, 'beneficiary/messages.html', context)

@login_required
def beneficiary_conversation_view(request, partner_id):
    """Chat-style conversation view for beneficiaries"""
    user = request.user
    if user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    try:
        partner = User.objects.get(id=partner_id)
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('beneficiary_messages_url')

    # Get IDs of messages deleted by this user
    deleted_msg_ids = DeletedMessage.objects.filter(user=user).values('message_id')

    # Get all messages between user and partner (excluding deleted ones)
    conversation_messages = Message.objects.filter(
        (Q(sender=user) & Q(recipient=partner)) |
        (Q(sender=partner) & Q(recipient=user))
    ).exclude(id__in=Subquery(deleted_msg_ids)).select_related('sender', 'recipient').order_by('created_at')

    # Mark messages from partner as read
    marked_read = Message.objects.filter(
        sender=partner,
        recipient=user,
        is_read=False
    ).update(is_read=True)
    if marked_read:
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    # Handle new message submission
    if request.method == 'POST':
        form = MessageForm(request.POST, request.FILES)
        if form.is_valid():
            message = form.save(commit=False)
            message.sender = user
            message.recipient = partner
            message.save()
            _invalidate_admin_comm_hub_cache_for_users([user.id, partner.id])
            messages.success(request, 'Message sent successfully.')
            return redirect('beneficiary_conversation_url', partner_id=partner_id)
    else:
        form = MessageForm()

    context = {
        'partner': partner,
        'conversation_messages': conversation_messages,
        'form': form,
    }

    return render(request, 'beneficiary/conversation.html', context)

@login_required
def beneficiary_compose_message_view(request):
    """Compose new message view for beneficiaries"""
    user = request.user
    if user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    if request.method == 'POST':
        recipient_id = request.POST.get('recipient')
        subject = request.POST.get('subject')
        content = request.POST.get('content')
        attachment = request.FILES.get('attachment')

        try:
            recipient = User.objects.get(id=recipient_id, status='active')
            message = Message.objects.create(
                sender=user,
                recipient=recipient,
                subject=subject,
                content=content,
                attachment=attachment,
                message_type='direct'
            )
            _invalidate_admin_comm_hub_cache_for_users([user.id, recipient.id])

            # Log the action
            attachment_info = f" with attachment '{attachment.name}'" if attachment else ""
            AuditLog.objects.create(
                user=user,
                action='Message Sent',
                details=f"Sent message to {recipient.get_full_name()}: {subject}{attachment_info}"
            )

            messages.success(request, f'Message sent to {recipient.get_full_name()} successfully!')
            return redirect('beneficiary_messages_url')

        except User.DoesNotExist:
            messages.error(request, 'Recipient not found.')
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')

    # Get list of staff assigned to tasks in projects where this user is the beneficiary
    # Plus all admins - beneficiaries can message assigned staff AND all admins
    assigned_staff_ids = Task.objects.filter(
        project__proposal__beneficiary=user,
        assigned_to__isnull=False
    ).values_list('assigned_to', flat=True).distinct()
    
    # Get all admins plus staff assigned to beneficiary's projects
    users = User.objects.filter(
        Q(role='admin') |  # All admins
        Q(id__in=assigned_staff_ids, role='dost_staff'),  # Staff assigned to their projects
        status='active'
    ).exclude(id=user.id).order_by('first_name', 'last_name')

    context = {
        'users': users,
    }

    return render(request, 'beneficiary/compose_message.html', context)

@login_required
def beneficiary_message_detail_view(request, message_id):
    """View message details and reply for beneficiaries"""
    user = request.user
    if user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    try:
        message = Message.objects.select_related('sender', 'recipient', 'parent_message').get(
            Q(id=message_id) & (Q(sender=user) | Q(recipient=user))
        )
    except Message.DoesNotExist:
        raise Http404("Message not found")

    # Mark as read if recipient is viewing
    if message.recipient == user and not message.is_read:
        message.mark_as_read()
        _invalidate_admin_comm_hub_cache_for_users([user.id])

    # Get message thread (parent and replies)
    thread_messages = []
    if message.parent_message:
        # This is a reply, get the entire thread
        root_message = message.parent_message
        while root_message.parent_message:
            root_message = root_message.parent_message
        thread_messages = Message.objects.filter(
            Q(id=root_message.id) | Q(parent_message=root_message) | Q(parent_message__parent_message=root_message)
        ).select_related('sender', 'recipient').order_by('created_at')
    else:
        # This is the root message
        thread_messages = Message.objects.filter(
            Q(id=message.id) | Q(parent_message=message) | Q(parent_message__parent_message=message)
        ).select_related('sender', 'recipient').order_by('created_at')

    if request.method == 'POST':
        reply_content = request.POST.get('reply_content')
        reply_attachment = request.FILES.get('reply_attachment')
        if reply_content:
            try:
                reply = Message.objects.create(
                    sender=user,
                    recipient=message.sender if message.recipient == user else message.recipient,
                    subject=f"Re: {message.subject}",
                    content=reply_content,
                    attachment=reply_attachment,
                    message_type='direct',
                    parent_message=message
                )
                _invalidate_admin_comm_hub_cache_for_users([reply.sender_id, reply.recipient_id])

                # Log the action
                attachment_info = f" with attachment '{reply_attachment.name}'" if reply_attachment else ""
                AuditLog.objects.create(
                    user=user,
                    action='Message Reply Sent',
                    details=f"Replied to message from {message.sender.get_full_name()}{attachment_info}"
                )

                messages.success(request, 'Reply sent successfully!')
                return redirect('beneficiary_message_detail_url', message_id=message.id)

            except Exception as e:
                messages.error(request, f'Error sending reply: {str(e)}')

    context = {
        'message': message,
        'thread_messages': thread_messages,
    }

    return render(request, 'beneficiary/message_detail.html', context)

@login_required
def beneficiary_announcements_view(request):
    """View announcements for beneficiaries"""
    user = request.user
    if user.role != 'beneficiary':
        messages.error(request, 'Access denied.')
        return redirect('index_url')

    # Get announcements for this user (all beneficiary announcements or targeted ones)
    # Filter in Python since SQLite doesn't support JSON field lookups like __contains
    all_announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')
    announcements = []
    for announcement in all_announcements:
        target_roles = announcement.target_roles or []
        # Include if no target roles specified, or if beneficiary is in target roles
        if not target_roles or 'beneficiary' in target_roles:
            announcements.append(announcement)

    context = {
        'announcements': announcements,
    }

    return render(request, 'beneficiary/announcements.html', context)


# ===============================
# CALENDAR VIEWS (#17)
# ===============================
@login_required
def administrator_calendar_view(request):
    """Calendar view for administrator"""
    user = request.user
    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('index_url')
    
    # Get all calendar events visible to this admin
    events = CalendarEvent.objects.filter(
        Q(created_by=user) | Q(is_public=True) | Q(participants=user)
    ).distinct()
    
    # Get projects with deadlines
    projects = Project.objects.all()
    
    # Get tasks with due dates
    tasks = Task.objects.filter(due_date__isnull=False)
    
    today = timezone.localdate()
    next_7_days = today + timedelta(days=7)
    next_30_days = today + timedelta(days=30)

    visible_projects = Project.objects.filter(date_of_completion__isnull=False)
    visible_tasks = Task.objects.filter(due_date__isnull=False)

    upcoming_agenda = []

    for event in events.filter(start_date__gte=today, start_date__lte=next_30_days).order_by('start_date')[:8]:
        upcoming_agenda.append({
            'title': event.title,
            'date': event.start_date,
            'source': 'event',
            'type': event.event_type or 'event',
            'color': event.color or '#3b82f6',
            'url': '',
        })

    for project in visible_projects.filter(date_of_completion__gte=today, date_of_completion__lte=next_30_days).order_by('date_of_completion')[:8]:
        upcoming_agenda.append({
            'title': project.project_title or 'Untitled Project',
            'date': project.date_of_completion,
            'source': 'project',
            'type': 'deadline',
            'color': '#ef4444',
            'url': f'/projects/{project.id}/',
        })

    for task in visible_tasks.filter(due_date__gte=today, due_date__lte=next_30_days).order_by('due_date')[:8]:
        upcoming_agenda.append({
            'title': task.title or 'Untitled Task',
            'date': task.due_date,
            'source': 'task',
            'type': 'task',
            'color': '#10b981' if task.status == 'completed' else '#3b82f6',
            'url': '',
        })

    upcoming_agenda = sorted(upcoming_agenda, key=lambda item: item['date'])[:10]

    context = {
        'events': events,
        'projects': projects,
        'tasks': tasks,
        'calendar_stats': {
            'events_total': events.count(),
            'upcoming_week': events.filter(start_date__gte=today, start_date__lte=next_7_days).count(),
            'project_deadlines': visible_projects.count(),
            'tasks_due': visible_tasks.count(),
        },
        'upcoming_agenda': upcoming_agenda,
    }
    return render(request, 'administrator/calendar.html', context)


@login_required
def administrator_calendar_events_api(request):
    """API endpoint to get calendar events as JSON"""
    import json
    from django.http import JsonResponse
    
    user = request.user
    events_data = []
    
    # Get calendar events
    events = CalendarEvent.objects.filter(
        Q(created_by=user) | Q(is_public=True) | Q(participants=user)
    ).distinct()
    
    for event in events:
        events_data.append({
            'id': event.id,
            'title': event.title,
            'start': event.start_date.isoformat(),
            'end': event.end_date.isoformat() if event.end_date else event.start_date.isoformat(),
            'color': event.color,
            'extendedProps': {
                'type': event.event_type,
                'description': event.description or '',
                'source': 'event',
            }
        })
    
    # Add project deadlines
    projects = Project.objects.filter(date_of_completion__isnull=False)
    for project in projects:
        events_data.append({
            'id': f'project_{project.id}',
            'title': f'📁 {project.project_title} (Deadline)',
            'start': project.date_of_completion.isoformat(),
            'end': project.date_of_completion.isoformat(),
            'color': '#ef4444',
            'extendedProps': {
                'type': 'deadline',
                'description': f'Project completion deadline for: {project.project_title}',
                'source': 'project',
            },
            'url': f'/projects/{project.id}/',
        })
    
    # Add task due dates
    tasks = Task.objects.filter(due_date__isnull=False)
    for task in tasks:
        events_data.append({
            'id': f'task_{task.id}',
            'title': f'✓ {task.title}',
            'start': task.due_date.isoformat(),
            'end': task.due_date.isoformat(),
            'color': '#3b82f6' if task.status != 'completed' else '#10b981',
            'extendedProps': {
                'type': 'task',
                'description': task.description or f'Task: {task.title}',
                'source': 'task',
            }
        })
    
    return JsonResponse(events_data, safe=False)


@login_required
def administrator_calendar_event_add(request):
    """Add a new calendar event"""
    if request.method == 'POST':
        user = request.user
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        event_type = request.POST.get('event_type', 'task')
        color = request.POST.get('color', '#3b82f6')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date') or start_date
        is_public = request.POST.get('is_public') == 'on'
        
        CalendarEvent.objects.create(
            title=title,
            description=description,
            event_type=event_type,
            color=color,
            start_date=start_date,
            end_date=end_date,
            is_public=is_public,
            created_by=user,
        )
        messages.success(request, 'Event added successfully!')
    return redirect('administrator_calendar_url')


@login_required
def administrator_calendar_event_edit(request, event_id):
    """Edit a calendar event"""
    event = get_object_or_404(CalendarEvent, id=event_id)
    
    if request.method == 'POST':
        event.title = request.POST.get('title', event.title)
        event.description = request.POST.get('description', event.description)
        event.event_type = request.POST.get('event_type', event.event_type)
        event.color = request.POST.get('color', event.color)
        event.start_date = request.POST.get('start_date', event.start_date)
        event.end_date = request.POST.get('end_date') or event.start_date
        event.is_public = request.POST.get('is_public') == 'on'
        event.save()
        messages.success(request, 'Event updated successfully!')
    return redirect('administrator_calendar_url')


@login_required
def administrator_calendar_event_delete(request, event_id):
    """Delete a calendar event"""
    event = get_object_or_404(CalendarEvent, id=event_id)
    event.delete()
    messages.success(request, 'Event deleted successfully!')
    return redirect('administrator_calendar_url')


@login_required
def staff_calendar_view(request):
    """Calendar view for staff"""
    user = request.user
    if user.role != 'dost_staff':
        messages.error(request, 'Access denied.')
        return redirect('index_url')
    
    # Get events visible to staff
    events = CalendarEvent.objects.filter(
        Q(created_by=user) | Q(is_public=True) | Q(participants=user)
    ).distinct()
    
    # Get tasks assigned to this staff
    tasks = Task.objects.filter(assigned_to=user, due_date__isnull=False)
    
    context = {
        'events': events,
        'tasks': tasks,
    }
    return render(request, 'staff/calendar.html', context)


# ===============================
# PROJECT CLONING (#14)
# ===============================
@login_required
def administrator_project_clone_view(request, pk):
    """Clone an existing project as a template"""
    user = request.user
    if user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('index_url')
    
    original_project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        new_title = request.POST.get('title', f"Copy of {original_project.title}")
        
        # Create new project based on original
        new_project = Project.objects.create(
            title=new_title,
            description=original_project.description,
            status='ongoing',
            proponent=original_project.proponent,
            budget=original_project.budget,
            approved_amount=original_project.approved_amount,
            municipality=original_project.municipality,
            barangay=original_project.barangay,
            latitude=original_project.latitude,
            longitude=original_project.longitude,
            date_started=timezone.now().date(),
        )
        
        # Clone milestones if any
        for milestone in original_project.milestones.all():
            ProjectMilestone.objects.create(
                project=new_project,
                title=milestone.title,
                description=milestone.description,
                planned_start=timezone.now().date(),
                planned_end=timezone.now().date() + timedelta(days=30),
                status='pending',
                order=milestone.order,
                created_by=user,
            )
        
        # Log the action
        AuditLog.objects.create(
            user=user,
            action='CREATE',
            model_name='Project',
            object_id=new_project.id,
            changes=f'Cloned from project #{original_project.id}: {original_project.title}'
        )
        
        messages.success(request, f'Project cloned successfully as "{new_title}"!')
        return redirect('administrator_projects_detail_url', pk=new_project.id)
    
    context = {
        'project': original_project,
    }
    return render(request, 'administrator/project_clone.html', context)


# ===============================
# DIGITAL SIGNATURES (#24)
# ===============================
@login_required
def create_digital_signature_view(request):
    """Create a digital signature for a document"""
    import hashlib
    
    if request.method == 'POST':
        user = request.user
        content_type = request.POST.get('content_type')  # proposal, project, extension_request
        object_id = request.POST.get('object_id')
        signature_data = request.POST.get('signature_data')  # Base64 encoded from canvas
        signature_image = request.FILES.get('signature_image')  # Uploaded image
        signature_type = request.POST.get('signature_type', 'approval')
        remarks = request.POST.get('remarks', '')
        
        # Get client IP
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip_address = x_forwarded_for.split(',')[0]
        else:
            ip_address = request.META.get('REMOTE_ADDR')
        
        # Create verification hash
        hash_string = f"{user.id}-{content_type}-{object_id}-{timezone.now().isoformat()}"
        verification_hash = hashlib.sha256(hash_string.encode()).hexdigest()
        
        signature = DigitalSignature.objects.create(
            user=user,
            signature_data=signature_data if signature_data else None,
            signature_image=signature_image if signature_image else None,
            signature_type=signature_type,
            content_type=content_type,
            object_id=object_id,
            ip_address=ip_address,
            device_info=request.META.get('HTTP_USER_AGENT', '')[:255],
            remarks=remarks,
            verification_hash=verification_hash,
        )
        
        messages.success(request, 'Document signed successfully!')
        
        # Redirect based on content type
        if content_type == 'proposal':
            return redirect('administrator_proposals_url')
        elif content_type == 'project':
            return redirect('administrator_projects_detail_url', pk=object_id)
        elif content_type == 'extension_request':
            return redirect('administrator_extension_requests_url')
        elif content_type == 'settings':
            return redirect('administrator_settings_url')
    
    return redirect('administrator_dashboard_url')


@login_required
def verify_digital_signature_view(request, signature_id):
    """Verify a digital signature"""
    from django.http import JsonResponse
    
    signature = get_object_or_404(DigitalSignature, id=signature_id)
    
    return JsonResponse({
        'verified': signature.is_verified,
        'signed_by': signature.user.full_name() or signature.user.username,
        'signed_at': signature.signed_at.isoformat(),
        'signature_type': signature.get_signature_type_display(),
        'verification_hash': signature.verification_hash,
    })


@login_required
def get_signatures_api(request, content_type, object_id):
    """API endpoint to get signatures for a document"""
    from django.http import JsonResponse
    
    signatures = DigitalSignature.objects.filter(
        content_type=content_type,
        object_id=object_id
    ).select_related('user').order_by('-signed_at')
    
    signatures_data = []
    for sig in signatures:
        signatures_data.append({
            'id': sig.id,
            'user_name': sig.user.full_name() if hasattr(sig.user, 'full_name') else sig.user.username,
            'signature_type': sig.get_signature_type_display(),
            'signature_data': sig.signature_data,
            'signed_at': sig.signed_at.strftime('%B %d, %Y at %I:%M %p'),
            'remarks': sig.remarks,
            'is_verified': sig.is_verified,
        })
    
    return JsonResponse({'signatures': signatures_data})


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
def delete_signature_api(request, signature_id):
    """API endpoint to delete a signature"""
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    try:
        signature = DigitalSignature.objects.get(id=signature_id)
        
        # Check permission - only admin or the signature owner can delete
        if request.user.role != 'administrator' and signature.user != request.user:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Store info for response before deletion
        content_type = signature.content_type
        object_id = signature.object_id
        
        # Delete the signature
        signature.delete()
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='Delete',
            model_name='DigitalSignature',
            object_id=signature_id,
            changes=f'Removed signature from {content_type} #{object_id}',
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Signature removed successfully',
            'content_type': content_type,
            'object_id': object_id
        })
        
    except DigitalSignature.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Signature not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ===============================
# GANTT CHART (#18)
# ===============================
@login_required
def project_gantt_view(request, pk):
    """Gantt chart view for a project"""
    project = get_object_or_404(Project, pk=pk)
    milestones = project.milestones.all().order_by('order', 'planned_start')
    tasks = Task.objects.filter(project=project).order_by('due_date')
    
    context = {
        'project': project,
        'milestones': milestones,
        'tasks': tasks,
    }
    return render(request, 'administrator/project_gantt.html', context)


@login_required
def project_milestones_api(request, pk):
    """API endpoint to get project milestones as JSON"""
    from django.http import JsonResponse
    
    project = get_object_or_404(Project, pk=pk)
    milestones = project.milestones.all().order_by('order', 'planned_start')
    
    data = []
    for m in milestones:
        data.append({
            'id': m.id,
            'title': m.title,
            'description': m.description,
            'planned_start': m.planned_start.isoformat(),
            'planned_end': m.planned_end.isoformat(),
            'actual_start': m.actual_start.isoformat() if m.actual_start else None,
            'actual_end': m.actual_end.isoformat() if m.actual_end else None,
            'status': m.status,
            'progress': m.progress_percentage,
            'is_overdue': m.is_overdue,
        })
    
    return JsonResponse(data, safe=False)


@login_required
def project_milestone_add(request, pk):
    """Add a milestone to a project"""
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        description = request.POST.get('description', '')
        planned_start_raw = (request.POST.get('planned_start') or '').strip()
        planned_end_raw = (request.POST.get('planned_end') or '').strip()

        if not title:
            messages.error(request, 'Milestone title is required.')
            return redirect('project_gantt_url', pk=pk)

        try:
            planned_start = datetime.strptime(planned_start_raw, '%Y-%m-%d').date()
            planned_end = datetime.strptime(planned_end_raw, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format. Please use YYYY-MM-DD for milestone dates.')
            return redirect('project_gantt_url', pk=pk)

        if planned_end < planned_start:
            messages.error(request, 'Planned end date cannot be earlier than planned start date.')
            return redirect('project_gantt_url', pk=pk)

        try:
            ProjectMilestone.objects.create(
                project=project,
                title=title,
                description=description,
                planned_start=planned_start,
                planned_end=planned_end,
                created_by=request.user,
            )
            messages.success(request, 'Milestone added successfully!')
        except ValidationError as e:
            error_msg = '; '.join(e.messages) if hasattr(e, 'messages') else str(e)
            messages.error(request, f'Unable to add milestone: {error_msg}')
    
    return redirect('project_gantt_url', pk=pk)


@login_required
def project_milestone_update(request, milestone_id):
    """Update a project milestone"""
    milestone = get_object_or_404(ProjectMilestone, id=milestone_id)
    
    if request.method == 'POST':
        milestone.title = (request.POST.get('title', milestone.title) or milestone.title).strip()
        milestone.description = request.POST.get('description', milestone.description)
        milestone.status = request.POST.get('status', milestone.status)

        progress_raw = request.POST.get('progress', milestone.progress_percentage)
        try:
            progress_value = int(progress_raw)
        except (TypeError, ValueError):
            messages.error(request, 'Progress must be a whole number between 0 and 100.')
            return redirect('project_gantt_url', pk=milestone.project.id)

        if progress_value < 0 or progress_value > 100:
            messages.error(request, 'Progress must be between 0 and 100.')
            return redirect('project_gantt_url', pk=milestone.project.id)

        milestone.progress_percentage = progress_value

        actual_start_raw = (request.POST.get('actual_start') or '').strip()
        actual_end_raw = (request.POST.get('actual_end') or '').strip()

        try:
            actual_start = datetime.strptime(actual_start_raw, '%Y-%m-%d').date() if actual_start_raw else None
            actual_end = datetime.strptime(actual_end_raw, '%Y-%m-%d').date() if actual_end_raw else None
        except ValueError:
            messages.error(request, 'Invalid date format. Please use YYYY-MM-DD for actual milestone dates.')
            return redirect('project_gantt_url', pk=milestone.project.id)

        if actual_start and actual_end and actual_end < actual_start:
            messages.error(request, 'Actual end date cannot be earlier than actual start date.')
            return redirect('project_gantt_url', pk=milestone.project.id)

        milestone.actual_start = actual_start
        milestone.actual_end = actual_end

        try:
            milestone.save()
            messages.success(request, 'Milestone updated successfully!')
        except ValidationError as e:
            error_msg = '; '.join(e.messages) if hasattr(e, 'messages') else str(e)
            messages.error(request, f'Unable to update milestone: {error_msg}')
    
    return redirect('project_gantt_url', pk=milestone.project.id)


# ===============================
# LANGUAGE TOGGLE (#21)
# ===============================
@login_required
def set_language_view(request):
    """Set user's preferred language"""
    from django.http import JsonResponse
    
    if request.method == 'POST':
        language = request.POST.get('language', 'en')
        
        # Get or create user preferences
        preference, created = UserPreference.objects.get_or_create(user=request.user)
        preference.language = language
        preference.save()
        
        # Store in session as well
        request.session['language'] = language
        
        return JsonResponse({'success': True, 'language': language})
    
    return JsonResponse({'success': False})


# ===============================
# GLOBAL SEARCH API
# ===============================
@login_required
def global_search_api(request):
    """Global search - only shows items relevant to the current user's role"""
    from django.http import JsonResponse
    from django.db.models import Q
    
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    results = []
    user = request.user
    role = user.role
    
    try:
        # ========================================
        # ADMIN - Can search everything
        # ========================================
        if role in ['admin', 'administrator']:
            # Search Projects
            projects = Project.objects.filter(
                Q(project_title__icontains=query) | 
                Q(project_code__icontains=query) |
                Q(project_description__icontains=query)
            )[:5]
            print(f"[SEARCH] Role: {role}, Query: {query}, Found {len(projects)} projects")
            
            for project in projects:
                results.append({
                    'title': project.project_title,
                    'type': 'Project',
                    'icon': 'folder',
                    'url': f'/projects/{project.id}/'
                })
            
            # Search Users
            users = User.objects.filter(
                Q(first_name__icontains=query) | 
                Q(last_name__icontains=query) |
                Q(email__icontains=query)
            ).exclude(id=user.id)[:5]
            
            for u in users:
                results.append({
                    'title': f'{u.first_name} {u.last_name}',
                    'type': f'User ({u.role})',
                    'icon': 'person',
                    'url': f'/administrator/users/update/{u.id}/'
                })
            
            # Search Tasks
            tasks = Task.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            )[:5]
            
            for task in tasks:
                results.append({
                    'title': task.title,
                    'type': f'Task ({task.status})',
                    'icon': 'task_alt',
                    'url': f'/projects/{task.project.id}/' if task.project else '/administrator/tasks/'
                })
            
            # Search Proposals
            proposals = Proposal.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            )[:5]
            
            for proposal in proposals:
                results.append({
                    'title': proposal.title,
                    'type': f'Proposal ({proposal.status})',
                    'icon': 'description',
                    'url': f'/administrator/proposals/update/{proposal.id}/'
                })
            
            # Search Budgets (by fund source and fiscal year)
            budgets = Budget.objects.filter(
                Q(fund_source__icontains=query)
            )[:3]
            
            for budget in budgets:
                results.append({
                    'title': f'{budget.fund_source} ({budget.fiscal_year})',
                    'type': 'Budget',
                    'icon': 'account_balance',
                    'url': f'/administrator/budgets/update/{budget.id}/'
                })
            
            # Search Forms
            forms = FormTemplate.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query) |
                Q(category__icontains=query)
            )[:3]
            
            for form in forms:
                results.append({
                    'title': form.title,
                    'type': f'Form ({form.category})',
                    'icon': 'article',
                    'url': '/administrator/forms/'
                })
        
        # ========================================
        # STAFF - Can search projects, tasks, proposals they work with
        # ========================================
        elif role in ['staff', 'dost_staff']:
            # Search Projects (all projects - staff can view all)
            projects = Project.objects.filter(
                Q(project_title__icontains=query) | 
                Q(project_code__icontains=query)
            )[:5]
            
            for project in projects:
                results.append({
                    'title': project.project_title,
                    'type': 'Project',
                    'icon': 'folder',
                    'url': '/staff/projects/'
                })
            
            # Search Tasks assigned to this staff
            tasks = Task.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            ).filter(assigned_to=user)[:5]
            
            for task in tasks:
                results.append({
                    'title': task.title,
                    'type': f'Task ({task.status})',
                    'icon': 'task_alt',
                    'url': '/staff/tasks/'
                })
            
            # Search Proposals
            proposals = Proposal.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            )[:5]
            
            for proposal in proposals:
                results.append({
                    'title': proposal.title,
                    'type': f'Proposal ({proposal.status})',
                    'icon': 'description',
                    'url': f'/staff/proposals/update/{proposal.id}/'
                })
        
        # ========================================
        # PROPONENT - Can search their own projects, proposals, tasks
        # ========================================
        elif role == 'proponent':
            # Search Projects where they are the project leader or proponent
            projects = Project.objects.filter(
                Q(project_title__icontains=query) | 
                Q(project_code__icontains=query)
            ).filter(
                Q(project_leader=user) | Q(proposal__proponent=user)
            ).distinct()[:5]
            
            for project in projects:
                results.append({
                    'title': project.project_title,
                    'type': 'Project',
                    'icon': 'folder',
                    'url': '/proponent/projects/'
                })
            
            # Search Proposals submitted by this proponent
            proposals = Proposal.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            ).filter(proponent=user)[:5]
            
            for proposal in proposals:
                results.append({
                    'title': proposal.title,
                    'type': f'Proposal ({proposal.status})',
                    'icon': 'description',
                    'url': f'/proponent/proposals/update/{proposal.id}/'
                })
            
            # Search Tasks assigned to this proponent
            tasks = Task.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            ).filter(assigned_to=user)[:5]
            
            for task in tasks:
                results.append({
                    'title': task.title,
                    'type': f'Task ({task.status})',
                    'icon': 'task_alt',
                    'url': '/proponent/tasks/'
                })
        
        # ========================================
        # BENEFICIARY - Can only search their own projects
        # ========================================
        elif role == 'beneficiary':
            # Search Projects where they are the beneficiary
            projects = Project.objects.filter(
                Q(project_title__icontains=query) | 
                Q(project_code__icontains=query)
            ).filter(beneficiary=user)[:5]
            
            for project in projects:
                results.append({
                    'title': project.project_title,
                    'type': 'Project',
                    'icon': 'folder',
                    'url': '/beneficiary/projects/'
                })
            
            # Search Proposals where they are the beneficiary
            proposals = Proposal.objects.filter(
                Q(title__icontains=query) | 
                Q(description__icontains=query)
            ).filter(beneficiary=user)[:5]
            
            for proposal in proposals:
                results.append({
                    'title': proposal.title,
                    'type': f'Proposal ({proposal.status})',
                    'icon': 'description',
                    'url': '/beneficiary/proposals/'
                })
        
        print(f"[SEARCH] Returning {len(results)} results for role {role}")
        return JsonResponse({'results': results[:15]})
        
    except Exception as e:
        import traceback
        print(f"[SEARCH ERROR] {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'results': [], 'error': str(e)})


# ===============================
# MENTIONS API
# ===============================
@login_required
def mentions_search_api(request):
    """Search users for @mentions"""
    from django.http import JsonResponse
    from django.db.models import Q
    
    query = request.GET.get('q', '').strip()
    if len(query) < 1:
        return JsonResponse({'users': []})
    
    users = User.objects.filter(
        Q(first_name__icontains=query) | 
        Q(last_name__icontains=query) |
        Q(username__icontains=query)
    ).exclude(id=request.user.id)[:10]
    
    results = []
    for u in users:
        results.append({
            'id': u.id,
            'username': u.username,
            'full_name': f'{u.first_name} {u.last_name}'.strip() or u.username,
            'role': u.role,
            'avatar': u.profile_picture.url if u.profile_picture else None
        })
    
    return JsonResponse({'users': results})


@login_required
def create_mention_view(request):
    """Create a mention and notify the user"""
    from django.http import JsonResponse
    
    if request.method == 'POST':
        mentioned_user_id = request.POST.get('mentioned_user_id')
        content_type = request.POST.get('content_type')
        object_id = request.POST.get('object_id')
        message_preview = request.POST.get('message_preview', '')[:255]
        
        try:
            mentioned_user = User.objects.get(id=mentioned_user_id)
            
            mention = Mention.objects.create(
                mentioned_user=mentioned_user,
                mentioned_by=request.user,
                content_type=content_type,
                object_id=object_id,
                message_preview=message_preview
            )
            
            # Create notification for the mentioned user
            Notification.objects.create(
                user=mentioned_user,
                title='You were mentioned',
                message=f'{request.user.first_name} {request.user.last_name} mentioned you: {message_preview}',
                notification_type='mention',
            )
            
            return JsonResponse({'success': True, 'mention_id': mention.id})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})